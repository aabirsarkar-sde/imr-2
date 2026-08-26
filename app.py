from __future__ import annotations

import calendar
import hashlib
import io
import json
import os
import re
import shutil
import tempfile
import time
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from sqlalchemy import (
    Column,
    Date,
    DateTime,
    Float,
    Integer,
    LargeBinary,
    MetaData,
    Table,
    Text,
    create_engine,
    inspect,
    text,
)
from sqlalchemy.engine import Engine


APP_DIR = Path(__file__).resolve().parent
UPLOAD_DIR_NAME = "uploaded_reports"
MANUAL_DATA_FILE = "manual_readings.csv"
REPORT_GLOBS = ("*.xlsx", "*.xls", "*.csv")
# Month name -> number, including abbreviations ("jan", "sept", "dec") so
# filenames like "master sheet IMR Dec.-25" parse, not just full month names.
MONTHS = {}
for _month_index in range(1, 13):
    MONTHS[calendar.month_name[_month_index].lower()] = _month_index
    MONTHS[calendar.month_abbr[_month_index].lower()] = _month_index
MONTHS["sept"] = 9


def normalize_date_to_month_end(date: pd.Timestamp) -> pd.Timestamp:
    """Snap a date to the last day of its month. Every report represents ONE
    calendar month, but cover-sheet dates can carry an arbitrary day (e.g. July 5)
    while the filename parser uses month-end. Normalizing all report_dates to
    month-end prevents the Portfolio picker from showing duplicate "Jul 2026"
    entries for dates that differ only in day-of-month."""
    if pd.isna(date):
        return date
    day = calendar.monthrange(date.year, date.month)[1]
    return pd.Timestamp(year=date.year, month=date.month, day=day)


METRICS = {
    "Flow Rate": {
        "column": "flow_lph",
        "raw_label": "Total flow liter/hr.",
        "axis_label": "Flow Rate (liter/hr.)",
        "unit": "L/hr",
        "bad_direction": "down",
        "threshold_factor": 0.70,
    },
    "Conductivity": {
        "column": "conductivity_us_cm",
        "raw_label": "Cond. us/cm",
        "axis_label": "Conductivity (uS/cm)",
        "unit": "uS/cm",
        "bad_direction": "up",
        "threshold_factor": 1.30,
    },
}

# Units that mark an operating-parameter cell in the lower/right section of a
# plant sheet. Each reading is laid out as a (tag, value, unit) triple, e.g.
# "PI 1601" | 56 | "bar". The unit normalizes (see normalize_text) to one of
# these keys, which also tells us what kind of parameter it is.
OP_PARAM_KINDS = {
    "bar": "pressure",
    "deg c": "temperature",
    "us cm": "conductivity",
    "lit hrs": "flow",
    "m3 hr": "flow",
}

PARAM_COLUMNS = [
    "source_file",
    "plant_group",
    "plant",
    # The plant's id, resolved at parse time from the sheet's COVER block or its
    # name. Two RO trains at one site share a display name ("Aarti Jhagadia" is
    # both 1957 and 2708), so the SR — not the name — is a plant's identity.
    "plant_sr_no",
    "report_date",
    "tag",
    "kind",
    "value",
    "unit",
]

# Feed pressure to the high-pressure membrane array. The standard instrument
# scheme on these reports tags the array feed as PI 1601. We surface this as the
# operating pressure at which each flow/conductivity reading was taken. Other
# pressure tags are ignored for now.
FEED_TAG = "1601"


@dataclass(frozen=True)
class ReportMeta:
    plant_group: str
    report_date: pd.Timestamp
    # False when no Month-Year was found in the filename and report_date had to
    # be guessed from the file's mtime — the validation gate surfaces this as a
    # warning instead of letting a wrong date land silently.
    date_from_filename: bool = True


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).replace("\n", " ").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_label(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text or "Unknown"


def clean_sheet_name(sheet_name: str) -> str:
    return re.sub(r"\s+", " ", sheet_name).strip()


# A plant group is the set of regions a workbook covers, taken verbatim from the
# filename (e.g. "Ank,JH,Panoli"). A spelling/spacing/case slip in one month's
# filename — "Panol" vs "Panoli" — would otherwise fragment one logical group
# into two on the Portfolio/Dashboard pickers. canonicalize_plant_group() folds
# known variants to one name. To merge a future typo, add its normalized key
# here; do NOT special-case it in the parser.
PLANT_GROUP_ALIASES = {
    "ank,jh,panol": "Ank,JH,Panoli",
    "ank,jh,panoli": "Ank,JH,Panoli",
    # The monthly "master sheet IMR <month>-<yy>" workbooks all cover the Dahej
    # zone; fold their filename-derived label ("master sheet") to that zone name.
    "master sheet": "Dahej",
}


def canonicalize_plant_group(value: object) -> str:
    """Map a raw plant-group label to its canonical name. The lookup key is the
    label with whitespace stripped, comma-spacing collapsed, and lower-cased, so
    "Ank, JH, Panol" and "ank,jh,panoli" both resolve to the same entry. Unknown
    labels pass through cleaned (whitespace collapsed) but otherwise unchanged."""
    text = "" if pd.isna(value) else re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return text
    key = re.sub(r"\s*,\s*", ",", text).lower()
    return PLANT_GROUP_ALIASES.get(key, text)


def canonicalize_zone(value: object) -> str | None:
    """Normalize a zone label to one canonical spelling so casing/spacing/`Zone`
    slips don't fracture one zone into several (e.g. "VADODARA", "vadodara",
    "Vadodara Zone" all -> "Vadodara"). Blank/whitespace -> None so a missing zone
    is never mistaken for a real one. Applied both at ingest and at read time, so
    already-stored mixed-case rows collapse without a re-ingest."""
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    text = re.sub(r"\s*zone\s*$", "", text, flags=re.IGNORECASE).strip()
    return text.title() if text else None


def plant_sr_no_from_name(name: object) -> int | None:
    """Pull the plant id from a sheet name. Prefer the last "(NNNN)" group, e.g.
    "Lupin Ank RO1 (1639)" -> 1639 and "Bharat rasayan RO 6 (3523) ." -> 3523
    (trailing junk after the parens). If there are no parentheses, fall back to a
    trailing 3+ digit run ("BEIL PTHP 3550" -> 3550). The 3-digit floor avoids
    grabbing a stray "(2" from a truncated name. This joins a reading to its MIS row."""
    text = str(name)
    parens = re.findall(r"\((\d+)\)", text)
    if parens:
        return int(parens[-1])
    match = re.search(r"(\d{3,})\s*\.?\s*$", text)
    return int(match.group(1)) if match else None


def current_plant_name(frame: pd.DataFrame) -> str:
    """The name to display for one plant's rows — the one its most recent report
    used. Plants get renamed (Glenmark -> Alivus, Solvay -> Synesqo, Piramal ->
    PGP) and misspelled, so rows sharing a Plant SR No can carry several names;
    showing an arbitrary one would label a plant with a name it has retired."""
    if frame.empty:
        return ""
    if "report_date" in frame.columns and frame["report_date"].notna().any():
        frame = frame.sort_values("report_date")
    return str(frame["plant"].iloc[-1])


def plant_identity_key(frame: pd.DataFrame) -> pd.Series:
    """The value that identifies a plant: its Plant SR No, or its display name
    when no SR could be resolved.

    A site can run several RO trains whose sheets carry ONE display name — "Aarti
    Jhagadia" is both 1957 and 2708, "Ami life karakhadi" is both 3247 and 3251 —
    so any grouping, dedup or join on the name alone silently merges two
    physically separate plants. Every per-plant rollup keys on this instead."""
    sr = pd.to_numeric(frame["plant_sr_no"], errors="coerce").astype("Int64")
    return sr.astype(object).where(sr.notna(), frame["plant"].astype(str))


def parse_report_metadata(path: Path) -> ReportMeta:
    stem = path.stem
    # Longest names first so "march" wins over "mar", "sept" over "sep", etc.
    month_regex = "|".join(sorted(MONTHS, key=len, reverse=True))
    # Month (+ optional abbreviation dot), a spaces/dot/dash separator, then a
    # 4- *or* 2-digit year. Handles "April-2025", "Jan-26", "Dec.-25", "June -25".
    match = re.search(
        rf"\b({month_regex})\b\.?\s*[-./]?\s*((?:19|20)\d{{2}}|\d{{2}})",
        stem,
        flags=re.IGNORECASE,
    )

    if match:
        month_number = MONTHS[match.group(1).lower()]
        year = int(match.group(2))
        if year < 100:  # 2-digit year -> 20xx
            year += 2000
        day = calendar.monthrange(year, month_number)[1]
        report_date = pd.Timestamp(year=year, month=month_number, day=day)
        date_from_filename = True
        # The location label is whatever is NOT the date. The old filenames put it
        # after the date ("IMR <date> <loc>"); the master-sheet naming puts it
        # before ("<loc> IMR <date>"). Take both sides so either convention works.
        remainder = stem[: match.start()] + " " + stem[match.end() :]
    else:
        report_date = pd.Timestamp(path.stat().st_mtime, unit="s").normalize()
        date_from_filename = False
        remainder = stem

    # Strip the "IMR" report-type noise and any leading "NN - " index, then tidy.
    remainder = re.sub(r"\bIMR\b", " ", remainder, flags=re.IGNORECASE)
    remainder = re.sub(r"^\s*\d+\s*-\s*", "", remainder)
    plant_group = re.sub(r"\s+", " ", remainder).strip(" -_,") or stem

    return ReportMeta(
        plant_group=canonicalize_plant_group(plant_group),
        report_date=report_date,
        date_from_filename=date_from_filename,
    )


# A hardened per-plant report carries its own identity in a COVER block at the top
# of each sheet, so plant id / date / zone come from the sheet itself instead of
# the filename + MIS join. The label text (normalized) maps to a field; the value
# is the first non-empty cell to its right. Old master workbooks have no COVER, so
# read_cover_block() returns {} for them and every caller falls back to the
# filename-derived metadata + sheet-name id — i.e. unchanged behavior.
COVER_LABELS = {
    "plant_sr_no": "plant sr no",
    "report_date": "report date",
    "zone": "zone",
    "site_name": "site name",
    "plant_name": "plant name",
    "plant_capacity": "plant capacity",
}


def _value_right(raw: pd.DataFrame, row: int, col: int, span: int = 8) -> object:
    """First non-empty cell to the right of (row, col), within `span` columns.
    Skips the blanks that merged label/value cells leave behind."""
    last = min(col + 1 + span, raw.shape[1])
    for c in range(col + 1, last):
        value = raw.iat[row, c]
        if not pd.isna(value) and str(value).strip() != "":
            return value
    return None


# Some hand-filled workbooks don't use the template's separate label/value cells;
# they pack the identity into single cells as "LABEL :- VALUE" text, often two
# labels per cell, e.g. "SITE NAME :- <site>   Date :- <dd/mm/yyyy>" and
# "PLANT CAPACITY/PLANT SR.NO. :- <cap>/<sr>   Time :- <hh.mm>". Parse those inline
# pairs so the plant id and date come from the sheet instead of being lost (which
# left readings with no sr-no -> zone "Unknown"). Labels are matched longest/most-
# specific first; the site name is intentionally NOT trusted from this layout — a
# site's several RO units share one SITE NAME here, so the unique per-unit sheet
# name stays the display name.
_INLINE_COVER_PATTERNS = [
    ("plant_capacity_sr", r"plant\s*capacity\s*/\s*plant\s*sr\.?\s*no\.?"),
    ("site_name", r"site\s*name"),
    ("plant_name", r"plant\s*name"),
    ("plant_sr_no", r"plant\s*sr\.?\s*no\.?"),
    ("plant_capacity", r"plant\s*capacity"),
    ("report_date", r"date"),
    ("zone", r"zone"),
    ("time", r"time"),
    ("location", r"location"),
]
_INLINE_COVER_EMIT = {"plant_sr_no", "plant_capacity", "report_date", "zone"}


def _split_capacity_sr(value: str) -> tuple[int | None, str | None]:
    """From a combined 'PLANT CAPACITY/PLANT SR.NO.' value like '100W /2977' or
    '2402/150w', pull the SR (the bare 3-4 digit token) and the capacity (the
    token carrying a unit letter). The two orders both occur, so key off shape,
    not position."""
    sr: int | None = None
    cap: str | None = None
    for token in re.split(r"[/\s]+", value.strip()):
        if not token:
            continue
        if re.fullmatch(r"\d{3,4}", token):
            sr = int(token)
        elif re.search(r"[a-zA-Z]", token) and re.search(r"\d", token):
            cap = token
    return sr, cap


def _parse_inline_cover(cell_text: str) -> dict[str, object]:
    """Extract the trusted 'LABEL :- VALUE' pairs packed into one COVER cell. Each
    value runs from its label's ':-' to the start of the next label, so adjacent
    labels in the same cell (Date, Time) act as value terminators."""
    if ":-" not in cell_text:
        return {}

    hits: list[tuple[int, int, str]] = []
    for field, pattern in _INLINE_COVER_PATTERNS:
        for match in re.finditer(pattern + r"\s*:-", cell_text, flags=re.IGNORECASE):
            hits.append((match.start(), match.end(), field))
    if not hits:
        return {}
    hits.sort()

    # The combined 'plant capacity/plant sr no' label contains the standalone
    # 'plant sr no'/'plant capacity' labels; drop those inner overlaps.
    pruned: list[tuple[int, int, str]] = []
    last_value_start = -1
    for start, value_start, field in hits:
        if start < last_value_start:
            continue
        pruned.append((start, value_start, field))
        last_value_start = value_start

    out: dict[str, object] = {}
    for i, (_start, value_start, field) in enumerate(pruned):
        value_end = pruned[i + 1][0] if i + 1 < len(pruned) else len(cell_text)
        value = cell_text[value_start:value_end].strip()
        if not value:
            continue
        if field == "plant_capacity_sr":
            sr, cap = _split_capacity_sr(value)
            if sr is not None:
                out.setdefault("plant_sr_no", sr)
            if cap:
                out.setdefault("plant_capacity", cap)
        elif field in _INLINE_COVER_EMIT:
            out.setdefault(field, value)
    return out


def read_cover_block(raw: pd.DataFrame) -> dict[str, object]:
    """Read a sheet's COVER identity block, if present. Returns {} when none of the
    labels are found (old-format sheets), so callers can fall back cleanly."""
    raw_hits: dict[str, object] = {}
    max_row = min(len(raw), 14)
    max_col = min(raw.shape[1], 12)
    for r in range(max_row):
        for c in range(max_col):
            cell = raw.iat[r, c]
            if pd.isna(cell):
                continue
            cell_text = str(cell)
            # Inline "LABEL :- VALUE" layout (value packed in the same cell).
            for field, value in _parse_inline_cover(cell_text).items():
                raw_hits.setdefault(field, value)
            # Template layout: a label cell with its value in the cell to the right.
            label = normalize_text(cell_text)
            if not label:
                continue
            for field, key in COVER_LABELS.items():
                if field not in raw_hits and key in label:
                    value = _value_right(raw, r, c)
                    if value is not None:
                        raw_hits[field] = value

    cover: dict[str, object] = {}
    if "plant_sr_no" in raw_hits:
        sr = pd.to_numeric(raw_hits["plant_sr_no"], errors="coerce")
        if pd.notna(sr):
            cover["plant_sr_no"] = int(sr)
    if "report_date" in raw_hits:
        value = raw_hits["report_date"]
        if isinstance(value, (pd.Timestamp, datetime)):
            # Real Excel date cells arrive as datetime — unambiguous.
            date = pd.Timestamp(value)
        else:
            # Free-typed text like "11/07/2026": assume dd-mm (the Indian
            # convention on these forms). Parse dayfirst — NOT via pd.Timestamp,
            # which would silently read the string month-first (US) and land on
            # the wrong month.
            date = pd.to_datetime(str(value), errors="coerce", dayfirst=True)
        # A report can't be from a future month — a cover date whose month is
        # ahead of today is a data-entry slip (e.g. "07/09/2026" typed for a July
        # report). Drop it so the caller falls back to the filename-derived month
        # instead of inventing a future month in the trend/pickers.
        if pd.notna(date) and date.to_period("M") > pd.Timestamp.now().to_period("M"):
            date = pd.NaT
        if pd.notna(date):
            # Snap to month-end so all dates for the same month collapse to one
            # timestamp — prevents duplicate "Jul 2026" entries in pickers.
            cover["report_date"] = normalize_date_to_month_end(date)
    for field in ("zone", "site_name", "plant_name", "plant_capacity"):
        if field in raw_hits:
            text = str(raw_hits[field]).strip()
            if text:
                cover[field] = text
    return cover


def sheet_context(
    raw: pd.DataFrame, sheet_name: str, metadata: ReportMeta
) -> tuple[str, int | None, ReportMeta, dict[str, object]]:
    """Resolve a sheet's plant id, sr-no and effective metadata, preferring the
    COVER block and falling back to the sheet name + filename metadata."""
    cover = read_cover_block(raw)
    # A real COVER block is marked by an identity/date cell (sr-no, zone or date).
    # Only then do we take the plant's display name from the sheet's name cell
    # (site_name); old-format sheets fall back to the tab name, even if they
    # happen to carry a stray "SITE NAME" label.
    has_cover = bool(
        cover.get("plant_sr_no") or cover.get("zone") or cover.get("report_date")
    )
    identity = cover.get("plant_name") or cover.get("site_name")
    plant = str(identity) if (has_cover and identity) else clean_sheet_name(sheet_name)
    sr_value = cover.get("plant_sr_no")
    plant_sr_no = int(sr_value) if sr_value is not None else plant_sr_no_from_name(sheet_name)
    report_date = cover.get("report_date") or metadata.report_date
    zone = cover.get("zone")
    plant_group = canonicalize_plant_group(str(zone)) if zone else metadata.plant_group
    effective = ReportMeta(
        plant_group=plant_group,
        report_date=report_date,
        date_from_filename=metadata.date_from_filename if not cover.get("report_date") else True,
    )
    return plant, plant_sr_no, effective, cover


def is_master_om_file(path: Path) -> bool:
    """True if `path` is the fleet-wide 'All plant O & M' register. Reads only the
    first sheet's top rows, so it's cheap enough to screen files at scan time."""
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return False
    try:
        head = pd.read_excel(path, sheet_name=0, header=None, nrows=8)
    except Exception:  # noqa: BLE001
        return False
    return is_master_om_list(head)


def report_paths(root: Path) -> list[Path]:
    source_roots = [root, root / UPLOAD_DIR_NAME]
    paths: list[Path] = []

    for source_root in source_roots:
        if not source_root.exists():
            continue
        paths.extend(
            path
            for pattern in REPORT_GLOBS
            for path in source_root.glob(pattern)
            if path.name != MANUAL_DATA_FILE and not path.name.startswith("~$")
        )

    # The O&M register is not a monthly report — it seeds the editable `plants`
    # table (seed_plants_if_empty), so keep it out of the readings/mis ingest.
    return sorted(p for p in paths if not is_master_om_file(p))


def master_om_paths(root: Path) -> list[Path]:
    """The O&M register file(s) on disk — the seed source for the plant register."""
    found: list[Path] = []
    for source_root in (root, root / UPLOAD_DIR_NAME):
        if not source_root.exists():
            continue
        for pattern in REPORT_GLOBS:
            for path in source_root.glob(pattern):
                if not path.name.startswith("~$") and is_master_om_file(path):
                    found.append(path)
    return sorted(found)


def is_module_header(value: object) -> bool:
    text = normalize_text(value)
    return text in {"mo no", "mod no", "module no", "module number"} or bool(
        re.fullmatch(r"(mo|mod|module)\s*no", text)
    )


def is_install_date_header(value: object) -> bool:
    text = normalize_text(value)
    return "inst" in text and "date" in text


def is_flow_header(value: object) -> bool:
    text = normalize_text(value)
    return "flow" in text and (
        "total" in text or "rate" in text or "liter" in text or "lph" in text
    )


def is_conductivity_header(value: object) -> bool:
    text = normalize_text(value)
    return "cond" in text or "conductivity" in text or "us cm" in text


def is_time_header(value: object) -> bool:
    """The 'Time for ___ ml' column — the raw stopwatch reading (seconds) from
    which flow is computed as (volume_ml / 1000) / (time / 3600) = volume*3.6/time."""
    text = normalize_text(value)
    return "time for" in text or ("time" in text and ("ml" in text or "ltr" in text))


def parse_volume_ml(value: object) -> float | None:
    """Millilitres named in a 'Time for X ml' header. Accepts '500 ml', '1000ml',
    '1 ltr', '2 ltr', '1Ltr' (litres → ×1000). Returns None when no volume is
    printed (the blank template header 'Time for ______ ml')."""
    text = normalize_text(value)
    match = re.search(r"(\d+(?:\.\d+)?)\s*(ml|litre|liter|ltr|l)\b", text)
    if not match:
        return None
    amount = float(match.group(1))
    return amount if match.group(2) == "ml" else amount * 1000.0


def is_plant_group_header(value: object) -> bool:
    text = normalize_text(value)
    return text in {"plant group", "site group", "area group"} or "plant group" in text


def is_plant_header(value: object) -> bool:
    text = normalize_text(value)
    return text in {"plant", "site", "site name", "plant name"} or "site name" in text


def is_stage_header(value: object) -> bool:
    text = normalize_text(value)
    return text in {"stage", "ro stage"}


def is_reading_date_header(value: object) -> bool:
    text = normalize_text(value)
    return (
        text in {"date", "reading date", "report date", "sample date"}
        or "reading date" in text
        or "report date" in text
    )


def find_header_row(raw: pd.DataFrame) -> int | None:
    max_rows = min(len(raw), 80)
    best_row = None
    best_score = 0

    for row_index in range(max_rows):
        values = raw.iloc[row_index].tolist()
        module_count = sum(is_module_header(value) for value in values)
        has_install_date = any(is_install_date_header(value) for value in values)
        has_flow = any(is_flow_header(value) for value in values)
        has_conductivity = any(is_conductivity_header(value) for value in values)

        score = module_count * 3 + has_install_date + has_flow + has_conductivity
        if module_count and has_install_date and has_flow and has_conductivity and score > best_score:
            best_row = row_index
            best_score = score

    return best_row


def find_stage_label(raw: pd.DataFrame, header_row: int, start_col: int, end_col: int) -> str:
    if header_row <= 0:
        return ""

    for row_index in range(header_row - 1, max(-1, header_row - 5), -1):
        values = raw.iloc[row_index, start_col:end_col].dropna().tolist()
        labels = [clean_label(value) for value in values if clean_label(value).lower() != "nan"]
        if labels:
            return labels[0]

    return ""


def find_module_blocks(raw: pd.DataFrame, header_row: int) -> list[dict[str, int | str]]:
    headers = raw.iloc[header_row].tolist()
    module_cols = [col for col, value in enumerate(headers) if is_module_header(value)]
    blocks: list[dict[str, int | str]] = []

    for index, module_col in enumerate(module_cols):
        next_module_col = module_cols[index + 1] if index + 1 < len(module_cols) else len(headers)
        span = range(module_col, next_module_col)
        block: dict[str, int | str] = {"module_col": module_col}

        for col in span:
            header = headers[col]
            if is_install_date_header(header):
                block["install_date_col"] = col
            elif is_flow_header(header):
                block["flow_col"] = col
            elif is_conductivity_header(header):
                block["conductivity_col"] = col
            elif is_time_header(header):
                block["time_col"] = col
                volume = parse_volume_ml(header)
                if volume is not None:
                    block["volume_ml"] = volume

        required = {"module_col", "install_date_col", "flow_col", "conductivity_col"}
        if required.issubset(block):
            block["stage"] = find_stage_label(raw, header_row, module_col, next_module_col)
            block["end_col"] = next_module_col
            blocks.append(block)

    return blocks


def value_at(raw: pd.DataFrame, row: int, col: int | str) -> object:
    if not isinstance(col, int) or row >= len(raw) or col >= raw.shape[1]:
        return np.nan
    return raw.iat[row, col]


def is_bypass_marker(values: list[object]) -> bool:
    """A module is bypassed when any cell in its block reads BY PASS / Bypass.

    A bypassed membrane has a module number but empty flow/conductivity cells,
    so it would otherwise be dropped — yet it is the strongest replacement
    signal we have, so we keep it as data (status="bypass")."""
    for value in values:
        normalized = normalize_text(value)
        if "by pass" in normalized or "bypass" in normalized:
            return True
    return False


def extract_rows_from_sheet(
    raw: pd.DataFrame,
    *,
    report_path: Path,
    sheet_name: str,
    metadata: ReportMeta,
    plant_override: str | None = None,
    plant_sr_no_override: int | None = None,
) -> list[dict[str, object]]:
    header_row = find_header_row(raw)
    if header_row is None:
        return []

    blocks = find_module_blocks(raw, header_row)
    if not blocks:
        return []

    records: list[dict[str, object]] = []
    plant = plant_override or clean_sheet_name(sheet_name)
    plant_sr_no = (
        plant_sr_no_override
        if plant_sr_no_override is not None
        else plant_sr_no_from_name(sheet_name)
    )

    for row_index in range(header_row + 1, len(raw)):
        for block in blocks:
            module_raw = value_at(raw, row_index, block["module_col"])
            module_number = pd.to_numeric(module_raw, errors="coerce")
            if pd.isna(module_number):
                continue

            block_cells = [
                value_at(raw, row_index, col)
                for col in range(int(block["module_col"]), int(block["end_col"]))
            ]
            bypass = is_bypass_marker(block_cells)

            if bypass:
                flow = np.nan
                conductivity = np.nan
            else:
                flow = pd.to_numeric(value_at(raw, row_index, block["flow_col"]), errors="coerce")
                # A live '=<volume*3.6>/<time>' formula written by openpyxl carries
                # no cached result, so pandas reads the flow cell as NaN. Recompute
                # it from the timed reading exactly as the sheet's formula would.
                if pd.isna(flow) and "time_col" in block and block.get("volume_ml"):
                    seconds = pd.to_numeric(value_at(raw, row_index, block["time_col"]), errors="coerce")
                    if pd.notna(seconds) and seconds != 0:
                        flow = float(block["volume_ml"]) * 3.6 / float(seconds)
                conductivity = pd.to_numeric(
                    value_at(raw, row_index, block["conductivity_col"]),
                    errors="coerce",
                )
                if pd.isna(flow) and pd.isna(conductivity):
                    continue

            records.append(
                {
                    "source_file": report_path.name,
                    "plant_group": metadata.plant_group,
                    "plant": plant,
                    "plant_sr_no": plant_sr_no,
                    "stage": block.get("stage", ""),
                    "module_number": float(module_number),
                    "install_date": value_at(raw, row_index, block["install_date_col"]),
                    "report_date": metadata.report_date,
                    "flow_lph": float(flow) if pd.notna(flow) else np.nan,
                    "conductivity_us_cm": float(conductivity) if pd.notna(conductivity) else np.nan,
                    "status": "bypass" if bypass else "active",
                }
            )

    return records


def find_column(columns: pd.Index, predicate) -> str | None:
    for column in columns:
        if predicate(column):
            return str(column)
    return None


def extract_rows_from_structured_table(
    table: pd.DataFrame,
    *,
    report_path: Path,
    sheet_name: str,
    metadata: ReportMeta,
) -> list[dict[str, object]]:
    if table.empty:
        return []

    table = table.rename(columns={column: clean_label(column) for column in table.columns})
    columns = table.columns
    module_col = find_column(columns, is_module_header)
    flow_col = find_column(columns, is_flow_header)
    conductivity_col = find_column(columns, is_conductivity_header)
    time_col = find_column(columns, is_time_header)
    time_volume_ml = parse_volume_ml(time_col) if time_col else None
    reading_date_col = find_column(columns, is_reading_date_header)
    install_date_col = find_column(columns, is_install_date_header)
    plant_group_col = find_column(columns, is_plant_group_header)
    plant_col = find_column(columns, is_plant_header)
    stage_col = find_column(columns, is_stage_header)

    if not all([module_col, flow_col, conductivity_col]):
        return []

    records: list[dict[str, object]] = []
    default_plant = clean_sheet_name(sheet_name) if sheet_name else metadata.plant_group

    for _, row in table.iterrows():
        module_number = pd.to_numeric(row.get(module_col), errors="coerce")
        if pd.isna(module_number):
            continue

        bypass = is_bypass_marker(list(row.values))
        if bypass:
            flow = np.nan
            conductivity = np.nan
        else:
            flow = pd.to_numeric(row.get(flow_col), errors="coerce")
            if pd.isna(flow) and time_col is not None and time_volume_ml:
                seconds = pd.to_numeric(row.get(time_col), errors="coerce")
                if pd.notna(seconds) and seconds != 0:
                    flow = time_volume_ml * 3.6 / float(seconds)
            conductivity = pd.to_numeric(row.get(conductivity_col), errors="coerce")
            if pd.isna(flow) and pd.isna(conductivity):
                continue

        reading_date = metadata.report_date
        if reading_date_col is not None and pd.notna(row.get(reading_date_col)):
            reading_date = row.get(reading_date_col)
        elif install_date_col is not None and pd.notna(row.get(install_date_col)):
            reading_date = row.get(install_date_col)

        install_date = reading_date
        if install_date_col is not None and pd.notna(row.get(install_date_col)):
            install_date = row.get(install_date_col)

        plant_group = metadata.plant_group
        if plant_group_col is not None and pd.notna(row.get(plant_group_col)):
            plant_group = canonicalize_plant_group(row.get(plant_group_col))

        plant = default_plant
        if plant_col is not None and pd.notna(row.get(plant_col)):
            plant = clean_label(row.get(plant_col))

        stage = ""
        if stage_col is not None and pd.notna(row.get(stage_col)):
            stage = clean_label(row.get(stage_col))

        records.append(
            {
                "source_file": report_path.name,
                "plant_group": plant_group,
                "plant": plant,
                "plant_sr_no": plant_sr_no_from_name(plant),
                "stage": stage,
                "module_number": float(module_number),
                "install_date": install_date,
                "report_date": reading_date,
                "flow_lph": float(flow) if pd.notna(flow) else np.nan,
                "conductivity_us_cm": float(conductivity) if pd.notna(conductivity) else np.nan,
                "status": "bypass" if bypass else "active",
            }
        )

    return records


def read_report(path: Path) -> list[dict[str, object]]:
    metadata = parse_report_metadata(path)
    records: list[dict[str, object]] = []

    if path.suffix.lower() == ".csv":
        try:
            structured = pd.read_csv(path)
        except Exception:  # noqa: BLE001
            structured = pd.DataFrame()

        if not structured.empty:
            records.extend(
                extract_rows_from_structured_table(
                    structured,
                    report_path=path,
                    sheet_name=metadata.plant_group,
                    metadata=metadata,
                )
            )
        if records:
            return records

        raw = pd.read_csv(path, header=None)
        records.extend(
            extract_rows_from_sheet(
                raw,
                report_path=path,
                sheet_name=metadata.plant_group,
                metadata=metadata,
            )
        )
        return records

    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        if clean_sheet_name(sheet_name).lower() == "mis":
            continue
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
        # The fleet-wide 'All plant O & M' register carries no readings — it is
        # the plant register only (parsed into MIS by read_report_mis).
        if is_master_om_list(raw):
            return []
        plant, plant_sr_no, eff_meta, _cover = sheet_context(raw, sheet_name, metadata)
        sheet_records = extract_rows_from_sheet(
            raw,
            report_path=path,
            sheet_name=sheet_name,
            metadata=eff_meta,
            plant_override=plant,
            plant_sr_no_override=plant_sr_no,
        )
        if not sheet_records:
            structured = pd.read_excel(path, sheet_name=sheet_name)
            sheet_records = extract_rows_from_structured_table(
                structured,
                report_path=path,
                sheet_name=sheet_name,
                metadata=eff_meta,
            )
        records.extend(sheet_records)

    return records


def extract_operating_parameters(
    raw: pd.DataFrame,
    *,
    report_path: Path,
    sheet_name: str,
    metadata: ReportMeta,
    plant_override: str | None = None,
    plant_sr_no_override: int | None = None,
) -> list[dict[str, object]]:
    """Scan a raw sheet for (tag, value, unit) instrument readings.

    Operating parameters appear in two layouts: a key-value block below the
    module table, or extra columns to the right of it. Both place the unit
    immediately right of the value, with the instrument tag one cell further
    left, so a single scan keyed on the unit cell handles both. The last
    occurrence of a tag wins.
    """
    plant = plant_override or clean_sheet_name(sheet_name)
    plant_sr_no = (
        plant_sr_no_override
        if plant_sr_no_override is not None
        else plant_sr_no_from_name(sheet_name)
    )
    found: dict[str, dict[str, object]] = {}
    rows, cols = raw.shape

    for row_index in range(rows):
        for col_index in range(cols):
            unit = normalize_text(raw.iat[row_index, col_index])
            kind = OP_PARAM_KINDS.get(unit)
            if kind is None or col_index < 2:
                continue

            value = pd.to_numeric(raw.iat[row_index, col_index - 1], errors="coerce")
            tag = normalize_text(raw.iat[row_index, col_index - 2])
            # Some sheets leave a blank spacer column between the tag and its
            # value, e.g. the feed/permeate conductivity block reads
            # "CIS 151" | <blank> | 8200 | "us/cm". When the immediate left cell
            # is empty, look one cell further left for the tag (requiring a
            # letter so a stray number doesn't masquerade as an instrument tag).
            if not tag and col_index >= 3:
                candidate = normalize_text(raw.iat[row_index, col_index - 3])
                if re.search(r"[a-z]", candidate):
                    tag = candidate
            if not tag or pd.isna(value):
                continue

            found[tag] = {
                "source_file": report_path.name,
                "plant_group": metadata.plant_group,
                "plant": plant,
                "plant_sr_no": plant_sr_no,
                "report_date": metadata.report_date,
                "tag": tag.upper(),
                "kind": kind,
                "value": float(value),
                "unit": unit,
            }

    return list(found.values())


def read_report_parameters(path: Path) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        return []

    metadata = parse_report_metadata(path)
    records: list[dict[str, object]] = []
    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        if clean_sheet_name(sheet_name).lower() == "mis":
            continue
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
        if is_master_om_list(raw):
            return []
        plant, plant_sr_no, eff_meta, _cover = sheet_context(raw, sheet_name, metadata)
        records.extend(
            extract_operating_parameters(
                raw,
                plant_override=plant,
                plant_sr_no_override=plant_sr_no,
                report_path=path,
                sheet_name=sheet_name,
                metadata=eff_meta,
            )
        )
    return records


def find_mis_header_row(raw: pd.DataFrame) -> int | None:
    for row_index in range(min(len(raw), 8)):
        joined = " ".join(normalize_text(value) for value in raw.iloc[row_index].tolist())
        if "plant sr no" in joined and "site name" in joined:
            return row_index
    return None


def map_mis_columns(header_cells: list[object]) -> dict[str, int]:
    """Map MIS field name -> column index by matching the normalized header.
    'plant sr no' is checked before the plain 'sr no' so the two don't collide."""
    mapping: dict[str, int] = {}
    for col, cell in enumerate(header_cells):
        normalized = normalize_text(cell)
        if not normalized:
            continue
        if "plant sr no" in normalized:
            mapping["plant_sr_no"] = col
        elif normalized == "zone":
            mapping["zone"] = col
        elif "zm name" in normalized:
            mapping["zm_name"] = col
        elif "site name" in normalized:
            mapping["site_name"] = col
        elif normalized == "status":
            mapping["status"] = col
        elif "membrane required" in normalized:
            mapping["membrane_required"] = col
        elif "remark" in normalized:
            mapping["remarks"] = col
    return mapping


def find_master_om_header_row(raw: pd.DataFrame) -> int | None:
    """Row index of the fleet-wide 'All plant O & M' register header (Site Name /
    Plant Sr No / Installed Capacity). The 'installed capacity' column is what
    distinguishes this authoritative fleet register from a per-workbook MIS sheet,
    which instead carries ZONE / ZM NAME / STATUS columns."""
    for row_index in range(min(len(raw), 8)):
        joined = " ".join(normalize_text(v) for v in raw.iloc[row_index].tolist())
        if "site name" in joined and "plant sr no" in joined and "installed capacity" in joined:
            return row_index
    return None


def is_master_om_list(raw: pd.DataFrame) -> bool:
    return find_master_om_header_row(raw) is not None


def extract_master_om_rows(
    raw: pd.DataFrame, *, report_path: Path, report_date: object
) -> list[dict[str, object]]:
    """Parse the fleet-wide 'All plant O & M' register into MIS rows.

    Layout: a Site Name / Plant Sr No / Installed Capacity table whose rows are
    grouped by zone — each zone begins with a lone '<Zone> Zone' banner row (only
    the site-name cell filled, no Plant Sr No), and zone is forward-filled down its
    group. This is the authoritative plant register (every zone and plant), so its
    rows are stamped with report_date such that the latest MIS row per plant — the
    one build_zone_by_sr() and the submission roster use — resolves to this file
    over any older per-workbook MIS.
    """
    header_row = find_master_om_header_row(raw)
    if header_row is None:
        return []

    mapping = map_mis_columns(raw.iloc[header_row].tolist())
    site_col = mapping.get("site_name")
    sr_col = mapping.get("plant_sr_no")
    if site_col is None or sr_col is None:
        return []
    cap_col = next(
        (c for c, cell in enumerate(raw.iloc[header_row].tolist())
         if "capacity" in normalize_text(cell)),
        None,
    )

    records: list[dict[str, object]] = []
    current_zone: str | None = None
    for row_index in range(header_row + 1, len(raw)):
        site = value_at(raw, row_index, site_col)
        site_text = "" if pd.isna(site) else str(site).strip()
        plant_sr_no = pd.to_numeric(value_at(raw, row_index, sr_col), errors="coerce")

        if pd.isna(plant_sr_no):
            # A lone '<Zone> Zone' banner row (no Plant Sr No) opens a new group.
            if re.search(r"\bzone\b", site_text, re.IGNORECASE):
                current_zone = canonicalize_zone(site_text)
            continue

        capacity = value_at(raw, row_index, cap_col) if cap_col is not None else None
        cap_text = None if pd.isna(capacity) else str(capacity).strip() or None
        records.append(
            {
                "source_file": report_path.name,
                "report_date": report_date,
                "zone": current_zone,
                "zm_name": None,
                "plant_sr_no": int(plant_sr_no),
                "site_name": site_text or None,
                "installed_capacity": cap_text,
                "status": None,
                "membrane_required": None,
                "remarks": None,
            }
        )

    return records


def extract_mis_rows(
    raw: pd.DataFrame, *, report_path: Path, metadata: ReportMeta
) -> list[dict[str, object]]:
    """Parse the MIS register sheet (one per workbook).

    ZONE and ZM NAME are merged cells that are blank on continuation rows, so we
    forward-fill them down. Rows without a PLANT SR NO are skipped.
    """
    header_row = find_mis_header_row(raw)
    if header_row is None:
        return []

    mapping = map_mis_columns(raw.iloc[header_row].tolist())
    if "plant_sr_no" not in mapping:
        return []

    def cell_text(row_index: int, field: str) -> str | None:
        value = value_at(raw, row_index, mapping.get(field))
        if pd.isna(value):
            return None
        return str(value).strip() or None

    records: list[dict[str, object]] = []
    last_zone: str | None = None
    last_zm: str | None = None

    for row_index in range(header_row + 1, len(raw)):
        zone = cell_text(row_index, "zone")
        if zone:  # merged cell: only set on the group's first row
            last_zone = zone
        zm_name = cell_text(row_index, "zm_name")
        if zm_name:
            last_zm = zm_name

        plant_sr_no = pd.to_numeric(value_at(raw, row_index, mapping["plant_sr_no"]), errors="coerce")
        if pd.isna(plant_sr_no):
            continue

        records.append(
            {
                "source_file": report_path.name,
                "report_date": metadata.report_date,
                "zone": last_zone,
                "zm_name": last_zm,
                "plant_sr_no": int(plant_sr_no),
                "site_name": cell_text(row_index, "site_name"),
                "status": cell_text(row_index, "status"),
                "membrane_required": pd.to_numeric(
                    value_at(raw, row_index, mapping.get("membrane_required")), errors="coerce"
                ),
                "remarks": cell_text(row_index, "remarks"),
            }
        )

    return records


def read_report_mis(path: Path) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        return []

    metadata = parse_report_metadata(path)
    records: list[dict[str, object]] = []
    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
        # The fleet-wide 'All plant O & M' register IS the plant register — the
        # whole file maps to MIS rows. Stamp today so it wins as the latest MIS
        # row per plant over any older per-workbook MIS/COVER zone.
        if is_master_om_list(raw):
            return extract_master_om_rows(
                raw, report_path=path, report_date=datetime.now(timezone.utc).date()
            )
        if clean_sheet_name(sheet_name).lower() == "mis":
            records.extend(extract_mis_rows(raw, report_path=path, metadata=metadata))
            continue
        # Per-plant template: no MIS sheet, but each sheet's COVER block carries
        # its zone/site/sr-no. Synthesize an MIS row so zone resolution (which
        # joins readings.plant_sr_no -> mis.zone) works without a register sheet.
        cover = read_cover_block(raw)
        sr_value = cover.get("plant_sr_no")
        # Canonicalize so casing slips don't fork one zone; a blank/whitespace
        # cover zone becomes None (skip) rather than a phantom "" zone that would
        # later clobber the register's real zone for this plant.
        zone = canonicalize_zone(cover.get("zone"))
        if sr_value is not None and zone:
            records.append(
                {
                    "source_file": path.name,
                    "report_date": cover.get("report_date") or metadata.report_date,
                    "zone": zone,
                    "zm_name": None,
                    "plant_sr_no": int(sr_value),
                    "site_name": cover.get("site_name"),
                    "status": None,
                    "membrane_required": None,
                    "remarks": None,
                }
            )
    return records


def safe_upload_name(filename: str) -> str:
    name = Path(filename).name
    name = re.sub(r"[^A-Za-z0-9., _()\-]+", "_", name).strip()
    return name or "uploaded_report.xlsx"


# --------------------------------------------------------------------------- #
# Persistence (PostgreSQL)
#
# Reports are parsed exactly once, at ingest. Each report's normalized rows are
# written to the `readings` and `parameters` tables and the file is recorded in
# `ingested_files` with a content hash. On every startup we re-scan the report
# folder, but only files whose hash is new or changed are parsed again — so the
# dashboard never re-parses Excel to render; it just queries the database.
#
# The schema uses SQLAlchemy's generic column types, so the same code runs on
# Postgres (production) and SQLite (tests) without change.
# --------------------------------------------------------------------------- #

DB_METADATA = MetaData()

READINGS_TABLE = Table(
    "readings",
    DB_METADATA,
    Column("source_file", Text),
    Column("plant_group", Text),
    Column("plant", Text),
    # Trailing "(NNNN)" id parsed from the plant sheet name; joins to mis.plant_sr_no.
    Column("plant_sr_no", Integer),
    Column("stage", Text),
    Column("module_number", Float),
    Column("module_label", Text),
    Column("install_date", Date),
    Column("report_date", Date),
    Column("flow_lph", Float),
    Column("conductivity_us_cm", Float),
    Column("status", Text),
)

PARAMETERS_TABLE = Table(
    "parameters",
    DB_METADATA,
    Column("source_file", Text),
    Column("plant_group", Text),
    Column("plant", Text),
    # Same id as readings.plant_sr_no — a plant's real identity. Sites run more
    # than one RO train under one display name, so plant-level analytics key on
    # this, never on `plant`.
    Column("plant_sr_no", Integer),
    Column("report_date", Date),
    Column("tag", Text),
    Column("kind", Text),
    Column("value", Float),
    Column("unit", Text),
)

# The MIS sheet (one per workbook) is the plant register: real ZONE, owner
# (ZM NAME), per-plant STATUS, membrane-required quantity, and PLANT SR NO,
# which joins to readings.plant_sr_no.
MIS_TABLE = Table(
    "mis",
    DB_METADATA,
    Column("source_file", Text),
    Column("report_date", Date),
    Column("zone", Text),
    Column("zm_name", Text),
    Column("plant_sr_no", Integer),
    Column("site_name", Text),
    Column("status", Text),
    Column("membrane_required", Float),
    Column("remarks", Text),
)

INGESTED_FILES_TABLE = Table(
    "ingested_files",
    DB_METADATA,
    Column("filename", Text, primary_key=True),
    Column("content_hash", Text, nullable=False),
    Column("ingested_at", DateTime),
    Column("n_readings", Integer),
    Column("n_parameters", Integer),
    Column("n_mis", Integer),
)

# Durable copy of each uploaded workbook's raw bytes, so a report can be
# re-parsed later (after a parser improvement) without depending on the
# ephemeral Cloud disk. New table -> create_all makes it; no migration needed.
REPORT_FILES_TABLE = Table(
    "report_files",
    DB_METADATA,
    Column("filename", Text, primary_key=True),
    Column("content_hash", Text, nullable=False),
    Column("file_bytes", LargeBinary, nullable=False),
    Column("uploaded_at", DateTime),
    Column("status", Text),
)

# The authoritative, user-editable plant register: one row per plant, the single
# source of truth for a plant's zone and whether it is still operating. Seeded
# once from the 'All plant O & M' list, then maintained in-app on the Plant
# Register page (add new clients, mark shut-down plants inactive, fix a zone).
# Zone resolution treats this as authority over any per-workbook cover zone, so a
# mis-typed cover ZONE can no longer invent a phantom zone.
PLANTS_TABLE = Table(
    "plants",
    DB_METADATA,
    Column("plant_sr_no", Integer, primary_key=True),
    Column("site_name", Text),
    Column("zone", Text),
    Column("installed_capacity", Text),
    Column("status", Text),  # "active" (expected to report) / "inactive" (shut down)
    Column("plant_type", Text),  # PLANT_TYPE_OPTIONS; defaults to "PT"
    Column("updated_at", DateTime),
)

READINGS_COLUMNS = [c.name for c in READINGS_TABLE.columns]
MIS_COLUMNS = [c.name for c in MIS_TABLE.columns]
PLANTS_COLUMNS = [c.name for c in PLANTS_TABLE.columns]


def database_url() -> str | None:
    """Resolve the DB connection string from DATABASE_URL or st.secrets, and
    normalize it to the psycopg2 driver SQLAlchemy expects."""
    url = os.environ.get("DATABASE_URL")
    if not url:
        try:
            url = st.secrets["postgres"]["url"]
        except Exception:  # noqa: BLE001 - secrets file may be absent
            url = None
    if not url:
        return None
    if url.startswith("postgres://"):
        url = "postgresql+psycopg2://" + url[len("postgres://") :]
    elif url.startswith("postgresql://"):
        url = "postgresql+psycopg2://" + url[len("postgresql://") :]
    return url


@st.cache_resource
def get_engine() -> Engine | None:
    url = database_url()
    if not url:
        return None
    return create_engine(url, pool_pre_ping=True)


def init_db(engine: Engine) -> None:
    DB_METADATA.create_all(engine)
    migrate_added_columns(engine)
    migrate_renamed_statuses(engine)


# Register statuses that have been relabelled. Reads all go through
# `canonical_status()`, which still resolves the old spellings, so this is only to
# stop the retired label from lingering in the stored data.
_RENAMED_STATUSES = {"STRO": "STPT RO"}


def migrate_renamed_statuses(engine: Engine) -> None:
    inspector = inspect(engine)
    if not inspector.has_table("plants"):
        return
    with engine.begin() as conn:
        for old, new in _RENAMED_STATUSES.items():
            conn.execute(
                text("UPDATE plants SET status = :new WHERE status = :old"),
                {"new": new, "old": old},
            )


# Tables whose live schema is patched up at boot. `plants` is maintained in-app,
# so no re-ingest would ever back-fill it. `parameters` is here because it is read
# on every page but only rewritten when a report is re-parsed — the ALTER has to
# land before the re-parse that fills the new column, and init_db() runs ahead of
# both ingest passes in main(), so this is the one place that ordering holds.
MIGRATED_TABLES = ("plants", "parameters")


def migrate_added_columns(engine: Engine) -> None:
    """`create_all()` creates MISSING tables but never ALTERs an existing one, so a
    column added to a table that is already live in prod has to be patched on. Adds
    any column the Table definition has and the live table lacks. Portable across
    Postgres and SQLite (neither is given a DEFAULT — the app treats NULL as blank
    and normalizes on save)."""
    inspector = inspect(engine)
    for name in MIGRATED_TABLES:
        table = DB_METADATA.tables.get(name)
        if table is None or not inspector.has_table(name):
            continue
        live = {c["name"] for c in inspector.get_columns(name)}
        for col in table.columns:
            if col.name in live:
                continue
            col_type = col.type.compile(engine.dialect)
            with engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE {name} ADD COLUMN {col.name} {col_type}"))
                # Existing plants predate the type column — seed them to the default
                # so the register never shows a blank type.
                if (name, col.name) == ("plants", "plant_type"):
                    conn.execute(text(
                        "UPDATE plants SET plant_type = :t WHERE plant_type IS NULL"
                    ), {"t": DEFAULT_PLANT_TYPE})
                # A column on an ingest-built table is only filled by re-parsing.
                # Clearing the hashes makes the startup ingest treat every report
                # as new, so it re-parses and back-fills. Rows are replaced
                # DELETE-then-insert per source_file, so this never duplicates —
                # and a report that can no longer be parsed simply keeps the rows
                # it already has.
                if (name, col.name) == ("parameters", "plant_sr_no"):
                    conn.execute(text("DELETE FROM ingested_files"))


def file_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


# Conductivity readings above this are data-entry errors (the raw sheets contain
# values like 2,984,050 uS/cm); blank them rather than let them skew the analysis.
MAX_PLAUSIBLE_CONDUCTIVITY_US_CM = 50000.0

# Each RO module houses this many membrane elements; fleet membrane count is this
# times the module count.
MEMBRANES_PER_MODULE = 184


def normalize_reading_records(records: list[dict[str, object]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=READINGS_COLUMNS)

    df = pd.DataFrame(records)
    df["install_date"] = pd.to_datetime(df["install_date"], errors="coerce")
    df["report_date"] = pd.to_datetime(df["report_date"], errors="coerce")
    df["flow_lph"] = pd.to_numeric(df["flow_lph"], errors="coerce")
    df["conductivity_us_cm"] = pd.to_numeric(df["conductivity_us_cm"], errors="coerce")
    df["module_number"] = pd.to_numeric(df["module_number"], errors="coerce")
    df["module_label"] = df["module_number"].map(format_module_number)
    if "plant_sr_no" not in df.columns:
        df["plant_sr_no"] = pd.NA
    df["plant_sr_no"] = pd.to_numeric(df["plant_sr_no"], errors="coerce").astype("Int64")
    if "status" not in df.columns:
        df["status"] = "active"
    df["status"] = df["status"].fillna("active")
    # Implausibly high conductivity is a data-entry error, not a reading.
    df.loc[df["conductivity_us_cm"] > MAX_PLAUSIBLE_CONDUCTIVITY_US_CM, "conductivity_us_cm"] = np.nan
    df = df.dropna(subset=["report_date", "module_number"])
    # Keep a row if it has a module number AND (it is bypassed OR has at least
    # one of flow/conductivity). Bypassed membranes carry no flow/conductivity
    # but are the strongest replacement signal, so they must be retained.
    keep = (
        (df["status"] == "bypass")
        | df["flow_lph"].notna()
        | df["conductivity_us_cm"].notna()
    )
    df = df[keep]
    # One reading per (file, plant, stage, module, month); a re-ingest of the
    # same file replaces its rows wholesale, so file-local dedup is enough.
    df = df.drop_duplicates(
        subset=["source_file", "plant", "stage", "module_number", "report_date"],
        keep="last",
    )
    return df.reindex(columns=READINGS_COLUMNS)


def normalize_parameter_records(records: list[dict[str, object]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=PARAM_COLUMNS)

    df = pd.DataFrame(records)
    df["report_date"] = pd.to_datetime(df["report_date"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    if "plant_sr_no" not in df.columns:
        df["plant_sr_no"] = pd.NA
    df["plant_sr_no"] = pd.to_numeric(df["plant_sr_no"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["report_date", "value"])
    # Dedup on the SR as well as the name: one workbook can hold two sheets for
    # two RO trains at the same site, which share a display name. Keying on the
    # name alone silently dropped the second train's whole parameter block.
    df = df.drop_duplicates(
        subset=["source_file", "plant", "plant_sr_no", "report_date", "tag"],
        keep="last",
    )
    return df.reindex(columns=PARAM_COLUMNS)


def normalize_mis_records(records: list[dict[str, object]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=MIS_COLUMNS)

    df = pd.DataFrame(records)
    df["report_date"] = pd.to_datetime(df["report_date"], errors="coerce")
    df["plant_sr_no"] = pd.to_numeric(df["plant_sr_no"], errors="coerce")
    df["membrane_required"] = pd.to_numeric(df["membrane_required"], errors="coerce")
    df = df.dropna(subset=["plant_sr_no"])
    df["plant_sr_no"] = df["plant_sr_no"].astype(int)
    df = df.drop_duplicates(subset=["source_file", "plant_sr_no"], keep="last")
    return df.reindex(columns=MIS_COLUMNS)


@dataclass
class ParseResult:
    """The normalized output of parsing one report, before it is committed.
    Carries the file's metadata plus a quality signal the gate surfaces."""
    meta: ReportMeta
    readings: pd.DataFrame
    parameters: pd.DataFrame
    mis: pd.DataFrame
    n_cond_blanked: int  # raw conductivity values dropped as out-of-range


def parse_report_file(path: Path) -> ParseResult:
    """Parse a report on disk into normalized DataFrames WITHOUT committing.
    The same parse the old ingest did inline, now reusable by the upload gate."""
    meta = parse_report_metadata(path)
    raw_readings = read_report(path)
    n_blanked = 0
    for record in raw_readings:
        value = pd.to_numeric(record.get("conductivity_us_cm"), errors="coerce")
        if pd.notna(value) and value > MAX_PLAUSIBLE_CONDUCTIVITY_US_CM:
            n_blanked += 1
    return ParseResult(
        meta=meta,
        readings=normalize_reading_records(raw_readings),
        parameters=normalize_parameter_records(read_report_parameters(path)),
        mis=normalize_mis_records(read_report_mis(path)),
        n_cond_blanked=n_blanked,
    )


def parse_report_bytes(filename: str, data: bytes) -> ParseResult:
    """Parse uploaded bytes without a durable disk write. The bytes are written
    to a transient temp dir under their ORIGINAL (sanitized) filename — group and
    report_date are derived from the name — then parsed via parse_report_file and
    the temp dir removed. Use safe_upload_name(filename) as the canonical key for
    staging / commit / byte storage so it matches the source_file stamped in rows."""
    tmpdir = Path(tempfile.mkdtemp(prefix="imr_stage_"))
    try:
        target = tmpdir / safe_upload_name(filename)
        target.write_bytes(data)
        return parse_report_file(target)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def commit_parse_result(
    engine: Engine, filename: str, content_hash: str, result: ParseResult
) -> dict[str, int]:
    """Write a parsed result to the live tables (DELETE-then-insert by
    source_file, idempotent) and record it in ingested_files. Returns counts."""
    readings, parameters, mis = result.readings, result.parameters, result.mis
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM readings WHERE source_file = :f"), {"f": filename})
        conn.execute(text("DELETE FROM parameters WHERE source_file = :f"), {"f": filename})
        conn.execute(text("DELETE FROM mis WHERE source_file = :f"), {"f": filename})
        readings.to_sql(
            "readings", conn, if_exists="append", index=False,
            method="multi", chunksize=1000,
        )
        if not parameters.empty:
            parameters.to_sql(
                "parameters", conn, if_exists="append", index=False,
                method="multi", chunksize=1000,
            )
        if not mis.empty:
            mis.to_sql(
                "mis", conn, if_exists="append", index=False,
                method="multi", chunksize=1000,
            )
        conn.execute(
            text(
                """
                INSERT INTO ingested_files
                    (filename, content_hash, ingested_at, n_readings, n_parameters, n_mis)
                VALUES (:f, :h, :ts, :nr, :np, :nm)
                ON CONFLICT (filename) DO UPDATE SET
                    content_hash = excluded.content_hash,
                    ingested_at = excluded.ingested_at,
                    n_readings = excluded.n_readings,
                    n_parameters = excluded.n_parameters,
                    n_mis = excluded.n_mis
                """
            ),
            {
                "f": filename,
                "h": content_hash,
                "ts": datetime.now(timezone.utc),
                "nr": int(len(readings)),
                "np": int(len(parameters)),
                "nm": int(len(mis)),
            },
        )
    return {
        "readings": int(len(readings)),
        "parameters": int(len(parameters)),
        "mis": int(len(mis)),
    }


def store_report_bytes(
    engine: Engine, filename: str, content_hash: str, data: bytes, status: str = "committed"
) -> None:
    """Persist an uploaded workbook's raw bytes so it can be re-parsed later
    without the original file on disk. Bytes are bound as a parameter (not via
    to_sql) — psycopg2 adapts to bytea, sqlite3 to BLOB."""
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO report_files (filename, content_hash, file_bytes, uploaded_at, status)
                VALUES (:f, :h, :b, :ts, :s)
                ON CONFLICT (filename) DO UPDATE SET
                    content_hash = excluded.content_hash,
                    file_bytes = excluded.file_bytes,
                    uploaded_at = excluded.uploaded_at,
                    status = excluded.status
                """
            ),
            {
                "f": filename,
                "h": content_hash,
                "b": data,
                "ts": datetime.now(timezone.utc),
                "s": status,
            },
        )


def load_report_bytes(engine: Engine, filename: str) -> bytes | None:
    """Return the stored raw bytes for a report, or None if not stored."""
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT file_bytes FROM report_files WHERE filename = :f"), {"f": filename}
        ).first()
    if row is None or row[0] is None:
        return None
    return bytes(row[0])  # psycopg2 returns a memoryview; normalize to bytes


def report_files_index(engine: Engine) -> list[str]:
    """Filenames that have durable bytes stored (eligible for re-parse)."""
    with engine.connect() as conn:
        return [
            r[0]
            for r in conn.execute(
                text("SELECT filename FROM report_files ORDER BY filename")
            ).all()
        ]


def ingest_reports(engine: Engine, data_dir: str) -> dict[str, list[str]]:
    """Parse and store any report on disk whose content hash is new or changed.

    Returns a summary of what was ingested / skipped (already up to date) /
    failed. Already-ingested files are never re-parsed. This is the startup path
    for repo-committed files; uploads go through the validation gate instead.
    """
    root = Path(data_dir)
    summary: dict[str, list[str]] = {"ingested": [], "skipped": [], "failed": []}

    with engine.connect() as conn:
        known = dict(
            conn.execute(text("SELECT filename, content_hash FROM ingested_files")).all()
        )

    for path in report_paths(root):
        try:
            fingerprint = file_fingerprint(path)
        except OSError as exc:
            summary["failed"].append(f"{path.name}: {exc}")
            continue

        if known.get(path.name) == fingerprint:
            summary["skipped"].append(path.name)
            continue

        try:
            result = parse_report_file(path)
        except Exception as exc:  # noqa: BLE001 - one bad report shouldn't sink the rest
            summary["failed"].append(f"{path.name}: {exc}")
            continue

        # A readings workbook must yield a module table, but a register-only file
        # (the 'All plant O & M' list) has no readings by design — reject only when
        # there is NOTHING to store, else its MIS/parameter rows get silently lost.
        if result.readings.empty and result.mis.empty and result.parameters.empty:
            summary["failed"].append(f"{path.name}: no data found")
            continue

        commit_parse_result(engine, path.name, fingerprint, result)
        summary["ingested"].append(path.name)

    return summary


def ingest_report_files(engine: Engine) -> dict[str, list[str]]:
    """Startup safety net: re-commit any durably-stored report whose content_hash
    is not already recorded in ingested_files — e.g. the derived tables were wiped
    but report_files survived a Cloud restart. Normally a no-op (hashes match)."""
    summary: dict[str, list[str]] = {"ingested": [], "skipped": [], "failed": []}
    with engine.connect() as conn:
        known = dict(
            conn.execute(text("SELECT filename, content_hash FROM ingested_files")).all()
        )
        stored = conn.execute(
            text("SELECT filename, content_hash FROM report_files")
        ).all()

    for filename, content_hash in stored:
        if known.get(filename) == content_hash:
            summary["skipped"].append(filename)
            continue
        data = load_report_bytes(engine, filename)
        if data is None:
            summary["failed"].append(f"{filename}: no stored bytes")
            continue
        try:
            result = parse_report_bytes(filename, data)
        except Exception as exc:  # noqa: BLE001
            summary["failed"].append(f"{filename}: {exc}")
            continue
        if result.readings.empty and result.mis.empty and result.parameters.empty:
            summary["failed"].append(f"{filename}: no data found")
            continue
        commit_parse_result(engine, filename, content_hash, result)
        summary["ingested"].append(filename)

    return summary


DEFAULT_ZONES = [
    "Ahmedabad", "Ankleshwar", "Dahej", "Jhagadia", "Panoli", "Vadodara", "Vapi",
]


def known_zones(plants: pd.DataFrame) -> list[str]:
    """The canonical zones the register defines (sorted). Falls back to the known
    business zones only when the register is empty, so the editor always offers a
    sensible zone list."""
    if plants is None or plants.empty:
        return list(DEFAULT_ZONES)
    zs = {z for z in plants["zone"].map(canonicalize_zone).dropna().unique()}
    return sorted(zs) if zs else list(DEFAULT_ZONES)


# A plant's register status. "Active" (or a blank status) = a normal RO plant
# expected to submit a monthly IMR. Every other value marks an exception — a
# non-operating state or a non-module RO type — that is NOT expected to report, so
# it's excluded from the IMR Tracker's roster. (SPRO = spiral RO, UF RO =
# ultrafiltration RO: these report differently, not as per-module IMRs.)
PLANT_STATUS_OPTIONS = [
    "Active", "Inactive", "STPT RO", "SPRO", "UF RO",
    "Stand By RO", "Not in Rochem Scope", "Plant Shutdown",
]
_STATUS_ALIASES = {
    "active": "Active",
    "inactive": "Inactive",
    # "STRO" was the old label for this status; rows already stored under it (and
    # workbooks that spell it either way) still resolve to the current one.
    "stpt ro": "STPT RO", "stptro": "STPT RO", "stpt": "STPT RO",
    "stro": "STPT RO", "st ro": "STPT RO", "st-ro": "STPT RO",
    "spro": "SPRO", "spiral ro": "SPRO", "sp ro": "SPRO",
    "uf ro": "UF RO", "ufro": "UF RO",
    "stand by ro": "Stand By RO", "standby ro": "Stand By RO", "standby": "Stand By RO",
    "not in rochem scope": "Not in Rochem Scope", "out of scope": "Not in Rochem Scope",
    "plant shutdown": "Plant Shutdown", "shutdown": "Plant Shutdown",
}


def canonical_status(status: object) -> str | None:
    """Map a raw register status to one of PLANT_STATUS_OPTIONS, or None for a
    blank/unrecognized status (treated as a normal reporting plant)."""
    if status is None or (isinstance(status, float) and pd.isna(status)):
        return None
    key = re.sub(r"\s+", " ", str(status)).strip().lower()
    if not key:
        return None
    return _STATUS_ALIASES.get(key)


# The kind of system installed at a plant, picked per plant on the Plant Register.
# Everything starts as PT (the default) until someone marks the exceptions.
PLANT_TYPE_OPTIONS = [
    "PT", "HP", "SP", "STPT", "UF", "LP", "Wringer", "Hybrid RO", "Chemical",
    "Utility", "Aqua",
]
DEFAULT_PLANT_TYPE = "PT"
_PLANT_TYPE_BY_KEY = {t.lower(): t for t in PLANT_TYPE_OPTIONS}
_PLANT_TYPE_ALIASES = {
    "hybrid": "Hybrid RO", "hybridro": "Hybrid RO", "hybrid r o": "Hybrid RO",
    "st pt": "STPT", "st-pt": "STPT",
    "uf ro": "UF", "ufro": "UF",
    "chemicals": "Chemical", "chem": "Chemical",
    "utilities": "Utility",
}


def canonical_plant_type(plant_type: object) -> str | None:
    """Map a raw plant type to one of PLANT_TYPE_OPTIONS, or None for a blank or
    unrecognized value (the caller decides whether to fall back to the default)."""
    if plant_type is None or (isinstance(plant_type, float) and pd.isna(plant_type)):
        return None
    key = re.sub(r"\s+", " ", str(plant_type)).strip().lower()
    if not key:
        return None
    return _PLANT_TYPE_BY_KEY.get(key) or _PLANT_TYPE_ALIASES.get(key)


def status_is_reporting(status: object) -> bool:
    """True if the plant is expected to submit a monthly IMR — i.e. Active (or a
    blank status). Every exception flag (Inactive, shutdown, standby, out-of-scope,
    non-module RO type) is not expected to report."""
    canon = canonical_status(status)
    return canon is None or canon == "Active"


def seed_plants_if_empty(engine: Engine, data_dir: str) -> int:
    """Populate the editable plant register once, only if it's empty. Seeds from
    the O&M list on disk; if that's not present, falls back to any zoned rows
    already in `mis` (the register may have been ingested before this table
    existed). Returns the number of rows seeded (0 if it already had rows)."""
    with engine.connect() as conn:
        if (conn.execute(text("SELECT COUNT(*) FROM plants")).scalar() or 0) > 0:
            return 0

    rows: list[dict[str, object]] = []
    for path in master_om_paths(Path(data_dir)):
        raw = pd.read_excel(path, sheet_name=0, header=None)
        for rec in extract_master_om_rows(
            raw, report_path=path, report_date=datetime.now(timezone.utc).date()
        ):
            rows.append({
                "plant_sr_no": int(rec["plant_sr_no"]),
                "site_name": rec.get("site_name"),
                "zone": canonicalize_zone(rec.get("zone")),
                "installed_capacity": rec.get("installed_capacity"),
            })
        if rows:
            break

    if not rows:  # fall back to zoned MIS rows already in the DB
        mis = pd.read_sql("SELECT * FROM mis", engine, parse_dates=["report_date"])
        if not mis.empty:
            zoned = mis.copy()
            zoned["zone"] = zoned["zone"].map(canonicalize_zone)
            zoned = zoned[zoned["zone"].notna()].dropna(subset=["plant_sr_no"])
            zoned = zoned.sort_values("report_date").drop_duplicates("plant_sr_no", keep="last")
            for _, r in zoned.iterrows():
                rows.append({
                    "plant_sr_no": int(r["plant_sr_no"]),
                    "site_name": r.get("site_name"),
                    "zone": r["zone"],
                    "installed_capacity": None,
                })

    if not rows:
        return 0

    frame = pd.DataFrame({r["plant_sr_no"]: r for r in rows}.values())  # de-dupe by SR
    frame["status"] = "active"
    frame["plant_type"] = DEFAULT_PLANT_TYPE
    frame["updated_at"] = datetime.now(timezone.utc)
    frame = frame.reindex(columns=PLANTS_COLUMNS)
    frame.to_sql("plants", engine, if_exists="append", index=False, method="multi", chunksize=500)
    return len(frame)


def save_plants(
    engine: Engine, edited: pd.DataFrame, scope_srs: set[int] | None = None
) -> int:
    """Persist the edited register rows. Rows without an SR are dropped, duplicate
    SRs collapse to the last, zones are canonicalized, status defaults to "Active".

    `scope_srs` bounds what the save may delete: when the editor shows a FILTERED
    view, pass the SR set that was on screen — only those (plus the ones being
    written) are replaced, so plants hidden by the filter are left untouched. Pass
    None to replace the whole table (row deletions in a full-table view stick).
    Returns the number of rows written."""
    clean = edited.copy()
    clean["plant_sr_no"] = pd.to_numeric(clean.get("plant_sr_no"), errors="coerce")
    clean = clean.dropna(subset=["plant_sr_no"])
    clean["plant_sr_no"] = clean["plant_sr_no"].astype(int)
    clean = clean.drop_duplicates("plant_sr_no", keep="last")
    clean["zone"] = clean.get("zone").map(canonicalize_zone) if "zone" in clean else None
    # Store a canonical status (PLANT_STATUS_OPTIONS); blank -> "Active".
    status = clean["status"] if "status" in clean else None
    clean["status"] = pd.Series(status, index=clean.index).map(
        lambda v: canonical_status(v) or "Active"
    )
    # Same for the system type: blank/unrecognized -> the default (PT).
    ptype = clean["plant_type"] if "plant_type" in clean else None
    clean["plant_type"] = pd.Series(ptype, index=clean.index).map(
        lambda v: canonical_plant_type(v) or DEFAULT_PLANT_TYPE
    )
    for col in ("site_name", "installed_capacity"):
        if col in clean:
            clean[col] = clean[col].map(lambda v: None if pd.isna(v) else (str(v).strip() or None))
    clean["updated_at"] = datetime.now(timezone.utc)
    clean = clean.reindex(columns=PLANTS_COLUMNS)

    kept = set(clean["plant_sr_no"].tolist())
    with engine.begin() as conn:
        if scope_srs is None:
            conn.execute(text("DELETE FROM plants"))
        else:
            to_delete = scope_srs | kept  # in-view (removals) + rows being rewritten
            for sr in to_delete:
                conn.execute(text("DELETE FROM plants WHERE plant_sr_no = :s"), {"s": int(sr)})
        if not clean.empty:
            clean.to_sql("plants", conn, if_exists="append", index=False,
                         method="multi", chunksize=500)
    return len(clean)


def add_plant(
    engine: Engine,
    plant_sr_no: int,
    site_name: str | None,
    zone: str | None,
    installed_capacity: str | None,
    status: str | None,
    plant_type: str | None = None,
) -> None:
    """Insert one new plant into the register (onboarding a new client's plant).

    Raises ValueError if the SR No is already taken — an existing plant is edited
    in the table, never silently overwritten from the add form."""
    sr = int(plant_sr_no)
    with engine.begin() as conn:
        taken = conn.execute(
            text("SELECT site_name FROM plants WHERE plant_sr_no = :s"), {"s": sr}
        ).first()
        if taken is not None:
            name = (taken[0] or "").strip()
            raise ValueError(
                f"SR No {sr} already exists in the register"
                + (f" ({name})." if name else ".")
                + " Edit that row in the table instead."
            )
        row = pd.DataFrame([{
            "plant_sr_no": sr,
            "site_name": (str(site_name).strip() or None) if site_name else None,
            "zone": canonicalize_zone(zone),
            "installed_capacity": (
                (str(installed_capacity).strip() or None) if installed_capacity else None
            ),
            "status": canonical_status(status) or "Active",
            "plant_type": canonical_plant_type(plant_type) or DEFAULT_PLANT_TYPE,
            "updated_at": datetime.now(timezone.utc),
        }]).reindex(columns=PLANTS_COLUMNS)
        row.to_sql("plants", conn, if_exists="append", index=False)


@dataclass
class Issue:
    """One line in a report's data-quality report. level is ERROR (blocks the
    commit), WARN (commit allowed, but something is off), or INFO (FYI)."""
    level: str
    code: str
    message: str


def looks_like_raw_filename(plant_group: str) -> bool:
    """Heuristic: a plant_group that still looks like an undigested filename
    (underscores, long digit runs, or many tokens) — a sign the date/label split
    didn't cleanly isolate the location."""
    if "_" in plant_group:
        return True
    if re.search(r"\d{4,}", plant_group):
        return True
    return len(plant_group.split()) > 4


def validate_parse(engine: Engine, filename: str, result: ParseResult) -> list[Issue]:
    """Produce a data-quality report for a parsed-but-not-yet-committed report.
    The upload gate shows these and blocks Confirm if any ERROR is present."""
    issues: list[Issue] = []
    readings = result.readings

    if readings.empty:
        # A register-only file (the 'All plant O & M' list) legitimately has no
        # module table — commit its MIS rows instead of rejecting it. Only a file
        # with nothing at all to store is a real error.
        if not result.mis.empty:
            issues.append(Issue("INFO", "register_only",
                f"Plant register — no module readings; {len(result.mis)} register "
                f"row(s) (zones/plants) will be committed."))
            return issues
        issues.append(Issue("ERROR", "no_data",
            "No module table or register rows found — there is nothing to commit."))
        return issues

    if not result.meta.date_from_filename:
        issues.append(Issue("WARN", "date_fallback",
            f"No month/year in the filename; report_date was guessed as "
            f"{result.meta.report_date:%d %b %Y}. Rename like '… April-2025 …' to fix."))

    if looks_like_raw_filename(result.meta.plant_group):
        issues.append(Issue("WARN", "raw_group",
            f"Plant group '{result.meta.plant_group}' looks like a raw filename, "
            f"not a clean location label."))

    missing = sorted(
        str(p) for p in
        readings.loc[readings["plant_sr_no"].isna(), "plant"].dropna().unique()
    )
    if missing:
        shown = ", ".join(missing[:8]) + (" …" if len(missing) > 8 else "")
        issues.append(Issue("WARN", "no_sr_no",
            f"{len(missing)} sheet(s) have no plant id (NNNN) — their readings get "
            f"zone 'Unknown': {shown}"))

    # Zone match: of the distinct plant_sr_no in this file, how many resolve to a
    # zone via this file's MIS or the existing DB register.
    sr_nos = {int(x) for x in readings["plant_sr_no"].dropna().unique()}
    if sr_nos:
        known_sr = {int(x) for x in result.mis["plant_sr_no"].dropna().unique()}
        with engine.connect() as conn:
            known_sr |= {
                int(r[0]) for r in
                conn.execute(text("SELECT DISTINCT plant_sr_no FROM mis")).all()
                if r[0] is not None
            }
        matched = len(sr_nos & known_sr)
        if matched / len(sr_nos) < 0.5:
            issues.append(Issue("WARN", "low_zone_match",
                f"Only {matched}/{len(sr_nos)} plants match a zone in the MIS register; "
                f"most readings will show zone 'Unknown'."))

    if result.n_cond_blanked:
        issues.append(Issue("INFO", "cond_blanked",
            f"{result.n_cond_blanked} conductivity value(s) above "
            f"{int(MAX_PLAUSIBLE_CONDUCTIVITY_US_CM):,} µS/cm dropped as data-entry errors."))

    with engine.connect() as conn:
        already = conn.execute(
            text("SELECT 1 FROM ingested_files WHERE filename = :f"), {"f": filename}
        ).first()
    if already:
        issues.append(Issue("INFO", "will_replace",
            f"'{filename}' is already in the database — committing replaces it."))

    return issues


def ingested_report_summary(engine: Engine) -> pd.DataFrame:
    """List every ingested report with its stored row counts, for the
    manage-data UI. One row per file in `ingested_files`."""
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                "SELECT filename, n_readings, n_parameters, n_mis, ingested_at "
                "FROM ingested_files ORDER BY filename"
            )
        ).all()
    return pd.DataFrame(
        rows,
        columns=["filename", "n_readings", "n_parameters", "n_mis", "ingested_at"],
    )


def remove_report(engine: Engine, root: Path, filename: str) -> dict[str, int]:
    """Completely remove one report from the system.

    Deletes its rows from `readings`/`parameters`/`mis`, drops its `ingested_files`
    record, and deletes the file on disk. The on-disk delete is essential: if the
    file survived, the next `ingest_reports()` run would re-parse it (its hash is
    no longer known) and restore everything we just deleted. Returns per-table
    delete counts.
    """
    counts: dict[str, int] = {}
    with engine.begin() as conn:
        for table in ("readings", "parameters", "mis"):
            result = conn.execute(
                text(f"DELETE FROM {table} WHERE source_file = :f"), {"f": filename}
            )
            counts[table] = result.rowcount or 0
        conn.execute(
            text("DELETE FROM ingested_files WHERE filename = :f"), {"f": filename}
        )
        # Drop the durable byte copy too, else a re-parse could resurrect it.
        conn.execute(
            text("DELETE FROM report_files WHERE filename = :f"), {"f": filename}
        )

    # Delete the source file wherever it lives so it is not re-ingested on refresh.
    for source_root in (root, root / UPLOAD_DIR_NAME):
        candidate = source_root / filename
        if candidate.exists():
            candidate.unlink()

    return counts


@st.cache_data(show_spinner="Loading readings from the database...")
def load_readings() -> pd.DataFrame:
    engine = get_engine()
    if engine is None:
        return pd.DataFrame(columns=READINGS_COLUMNS)
    df = pd.read_sql(
        "SELECT * FROM readings",
        engine,
        parse_dates=["install_date", "report_date"],
    )
    if df.empty:
        return df
    df["module_label"] = df["module_label"].astype("string")
    # Fold filename-typo group variants ("Panol" vs "Panoli") at read time too, so
    # rows ingested before canonicalize_plant_group() existed are merged without a
    # forced re-parse.
    df["plant_group"] = df["plant_group"].map(canonicalize_plant_group)
    # Snap report_date to month-end: cover-sheet dates may carry an exact day
    # (e.g. July 5) while filename dates use month-end. Without this, the
    # Portfolio picker shows duplicate "Jul 2026" entries for the same month.
    df["report_date"] = df["report_date"].map(normalize_date_to_month_end)
    return df.sort_values(["plant_group", "plant", "module_number", "report_date"]).reset_index(
        drop=True
    )


@st.cache_data(show_spinner="Loading operating parameters from the database...")
def load_parameters() -> pd.DataFrame:
    engine = get_engine()
    if engine is None:
        return pd.DataFrame(columns=PARAM_COLUMNS)
    df = pd.read_sql("SELECT * FROM parameters", engine, parse_dates=["report_date"])
    if df.empty:
        return pd.DataFrame(columns=PARAM_COLUMNS)
    df["plant_group"] = df["plant_group"].map(canonicalize_plant_group)
    # Same month-end normalization as readings — keeps salt passage, permeate
    # conductivity, and permeate flow sections aligned with the fleet status.
    df["report_date"] = df["report_date"].map(normalize_date_to_month_end)
    return df.sort_values(["plant_group", "plant", "tag", "report_date"]).reset_index(drop=True)


@st.cache_data(show_spinner="Loading the plant register (MIS) from the database...")
def load_mis() -> pd.DataFrame:
    engine = get_engine()
    if engine is None:
        return pd.DataFrame(columns=MIS_COLUMNS)
    df = pd.read_sql("SELECT * FROM mis", engine, parse_dates=["report_date"])
    # Fold in the authoritative plant register (register zones win; unrecognized
    # cover zones dropped) so every zone consumer resolves zones the same way.
    df = apply_register_authority(df, load_plants())
    if df.empty:
        return pd.DataFrame(columns=MIS_COLUMNS)
    # Month-end normalization, same as readings/parameters.
    df["report_date"] = df["report_date"].map(normalize_date_to_month_end)
    return df.sort_values(["zone", "plant_sr_no", "report_date"]).reset_index(drop=True)


@st.cache_data(show_spinner="Loading the plant register...")
def load_plants() -> pd.DataFrame:
    engine = get_engine()
    if engine is None:
        return pd.DataFrame(columns=PLANTS_COLUMNS)
    df = pd.read_sql("SELECT * FROM plants", engine)
    if df.empty:
        return pd.DataFrame(columns=PLANTS_COLUMNS)
    df["zone"] = df["zone"].map(canonicalize_zone)
    if "status" in df:
        # Keep the stored casing (e.g. "Plant Shutdown"); logic normalizes as needed.
        df["status"] = df["status"].fillna("active").astype(str).str.strip()
    if "plant_type" in df:
        # Rows written before the type column existed read as blank -> the default.
        df["plant_type"] = df["plant_type"].map(
            lambda v: canonical_plant_type(v) or DEFAULT_PLANT_TYPE
        )
    return df.sort_values(["zone", "plant_sr_no"], na_position="last").reset_index(drop=True)


# Far-future sentinel (pandas max is 2262-04-11): register rows are stamped with
# it so, in the "latest MIS row per plant" zone resolution, the register always
# wins over any per-workbook cover zone.
REGISTER_REPORT_DATE = pd.Timestamp("2262-01-01")
REGISTER_SOURCE = "__register__"


def apply_register_authority(mis: pd.DataFrame, plants: pd.DataFrame) -> pd.DataFrame:
    """Make the plant register authoritative over the raw MIS rows for zones:

    1. Drop any cover-derived MIS row whose zone the register doesn't recognize —
       so a mis-typed cover ZONE (e.g. a site/town name) can't invent a phantom
       zone. 2. Append every register plant as a MIS row stamped far in the future,
       so build_zone_by_sr() picks the register's zone over any cover zone.

    A no-op until the register is seeded, so the app still works pre-seed."""
    base = mis.copy() if (mis is not None and not mis.empty) else pd.DataFrame(columns=MIS_COLUMNS)
    if plants is None or plants.empty:
        return base

    reg = plants.dropna(subset=["plant_sr_no"]).copy()
    reg["zone"] = reg["zone"].map(canonicalize_zone)
    reg = reg[reg["zone"].notna()]
    known = set(reg["zone"].unique())
    if not known:
        return base

    if not base.empty:
        cover_zone = base["zone"].map(canonicalize_zone)
        base = base[cover_zone.isna() | cover_zone.isin(known)]

    reg_rows = pd.DataFrame(
        {
            "source_file": REGISTER_SOURCE,
            "report_date": REGISTER_REPORT_DATE,
            "zone": reg["zone"].to_numpy(),
            "zm_name": pd.Series([None] * len(reg), dtype=object),
            "plant_sr_no": reg["plant_sr_no"].astype(int).to_numpy(),
            "site_name": (
                reg["site_name"].astype(object).to_numpy() if "site_name" in reg
                else pd.Series([None] * len(reg), dtype=object)
            ),
            "status": pd.Series([None] * len(reg), dtype=object),
            "membrane_required": pd.Series([np.nan] * len(reg), dtype=float),
            "remarks": pd.Series([None] * len(reg), dtype=object),
        }
    ).reindex(columns=MIS_COLUMNS)
    if base.empty:
        return reg_rows
    return pd.concat([base, reg_rows], ignore_index=True)


def feed_pressure_series(
    params: pd.DataFrame, plant_group: str, plant: str, plant_sr_no: int | None = None
) -> pd.DataFrame:
    """The plant's PI 1601 feed pressure per month.

    Matched on Plant SR No wherever we have one: a site's two RO trains share a
    display name (so the name alone overlaid one train's pressure on the other's
    readings), and a renamed plant's older rows carry a name the caller never
    passes. Rows ingested before `parameters` carried an SR have none to match on,
    so the name lookup stays as the fallback."""
    pressures = params[params["kind"] == "pressure"]
    by_sr = pressures.iloc[0:0]
    if plant_sr_no is not None and "plant_sr_no" in pressures.columns:
        by_sr = pressures[pressures["plant_sr_no"] == plant_sr_no]
    pressures = (
        by_sr
        if not by_sr.empty
        else pressures[
            (pressures["plant_group"] == plant_group) & (pressures["plant"] == plant)
        ]
    )
    if pressures.empty:
        return pd.DataFrame(columns=["report_date", "value"])

    rows: list[dict[str, object]] = []
    for report_date, group in pressures.groupby("report_date"):
        feed = group[group["tag"].str.contains(FEED_TAG, regex=False)]["value"]
        if feed.empty:
            continue
        rows.append({"report_date": report_date, "value": float(feed.iloc[0])})

    series = pd.DataFrame(rows)
    if series.empty:
        return pd.DataFrame(columns=["report_date", "value"])
    return series.sort_values("report_date").reset_index(drop=True)


def format_module_number(value: object) -> str:
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return "Unknown"
    if float(number).is_integer():
        return str(int(number))
    return f"{number:g}"


def aggregate_series(df: pd.DataFrame, metric_col: str) -> pd.DataFrame:
    metric_df = df.dropna(subset=[metric_col]).copy()
    if metric_df.empty:
        return metric_df

    grouped = (
        metric_df.groupby("report_date", as_index=False)
        .agg(
            value=(metric_col, "mean"),
            install_date=("install_date", "first"),
            source_file=("source_file", "first"),
            stage=("stage", "first"),
        )
        .sort_values("report_date")
    )
    return grouped


# Teal, distinct from the blue reading line, so the operating-pressure context
# reads clearly on its own right-hand axis.
PRESSURE_LINE_COLOR = "#0d9488"


def add_feed_pressure_axis(fig: go.Figure, pressure: pd.DataFrame | None) -> None:
    """Overlay PI 1601 feed pressure as a dashed line on a secondary right-hand axis.

    Shared by the single-module and comparison charts so the operating pressure
    each reading was taken at is visible at a glance, not just on hover. Pressure
    is a plant-level series, so in the comparison view this is one shared line
    behind all the module lines.
    """
    if pressure is None or pressure.empty:
        return
    ordered = pressure.dropna(subset=["value"]).sort_values("report_date")
    if ordered.empty:
        return
    fig.add_trace(
        go.Scatter(
            x=ordered["report_date"],
            y=ordered["value"],
            mode="lines+markers",
            name="Feed pressure (PI 1601)",
            yaxis="y2",
            line={"color": PRESSURE_LINE_COLOR, "width": 2, "dash": "dot"},
            marker={"size": 6, "symbol": "diamond"},
            hovertemplate="Feed pressure (PI 1601)<br>%{x|%d %b %Y}<br>%{y:,.2f} bar<extra></extra>",
        )
    )
    fig.update_layout(
        yaxis2={
            "title": "Feed Pressure (bar, PI 1601)",
            "overlaying": "y",
            "side": "right",
            "showgrid": False,
            "automargin": True,
        }
    )


def make_chart(
    series: pd.DataFrame,
    *,
    metric_name: str,
    metric_config: dict[str, object],
    pressure: pd.DataFrame | None = None,
) -> go.Figure:
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=series["report_date"],
            y=series["value"],
            mode="lines+markers",
            name="Historical Reading",
            line={"color": "#2563eb", "width": 3},
            marker={"size": 8},
            hovertemplate="%{x|%d %b %Y}<br>%{y:,.2f}<extra></extra>",
        )
    )

    add_feed_pressure_axis(fig, pressure)

    fig.update_layout(
        title=f"{metric_name} History",
        xaxis_title="Report Date",
        yaxis_title=str(metric_config["axis_label"]),
        hovermode="x unified",
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        template="plotly_white",
        height=470,
    )
    return fig


def plant_no_column(values: pd.Series) -> pd.Series:
    """Format plant SR numbers for the "Plant No" column shown beside a plant name.

    Plant sheet names are inconsistent about carrying their SR ("Aarti Jhagadia
    HP Ro (1959)" vs "Aarti Industries Ud."), so every plant table gets the SR as
    its own column. Rendered as text, not int: Streamlit would otherwise print an
    id as "1,959"."""
    return values.map(lambda v: f"{int(v)}" if pd.notna(v) else "—")


def plant_label_with_sr(plant: str, plant_sr_no: int | None) -> str:
    """"Name (SR)" for a heading, where a table would instead get its own column.

    Many sheet names already carry the SR ("Aarti Jhagadia HP Ro (1959)", "Aarti
    Alchemie PTRO 2497"), so it is only appended when the name doesn't already
    read as that SR. Takes the SR from the caller rather than looking it up by
    name: one name can cover two RO trains, so a name alone cannot say which."""
    if plant_sr_no is None:
        return plant
    return plant if plant_sr_no_from_name(plant) == plant_sr_no else f"{plant} ({plant_sr_no})"


def metric_card(
    title: str,
    value: str,
    subtitle: str = "",
    color: str = "#0f172a",
    anchor: str | None = None,
) -> None:
    """A KPI tile. Pass `anchor` (a section's `anchor=` id) to make the whole card a
    link that jumps to the detail behind the number — same pattern as `jump_nav()`,
    so it is a plain anchor with no rerun."""
    card = f"""
        <div class="metric-title">{title}</div>
        <div class="metric-value" style="color:{color};">{value}</div>
        <div class="metric-subtitle">{subtitle}</div>
    """
    if anchor:
        body = f'<a class="metric-card metric-card-link" href="#{anchor}">{card}</a>'
    else:
        body = f'<div class="metric-card">{card}</div>'
    st.markdown(body, unsafe_allow_html=True)


def hero_card(title: str, value: str, subtitle: str = "", color: str = "#dc2626") -> None:
    """Oversized single-number card for the page's headline metric."""
    st.markdown(
        f"""
        <div class="hero-card">
            <div class="hero-title">{title}</div>
            <div class="hero-value" style="color:{color};">{value}</div>
            <div class="hero-subtitle">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def jump_nav(items: list[tuple[str, str, str]]) -> None:
    """Render a row of clickable cards that scroll to in-page section anchors.

    Each item is (label, anchor, subtitle); `anchor` must match the `anchor=`
    set on the target section's subheader. Pure anchor links — no rerun.
    """
    cards = "".join(
        f'<a class="nav-card" href="#{anchor}">'
        f'<span class="nav-label">{label}</span>'
        f'<span class="nav-sub">{subtitle}</span></a>'
        for label, anchor, subtitle in items
    )
    st.markdown(f'<div class="nav-grid">{cards}</div>', unsafe_allow_html=True)


def rerun_app() -> None:
    if hasattr(st, "rerun"):
        st.rerun()
    else:
        st.experimental_rerun()


def sidebar_filters(df: pd.DataFrame) -> tuple[pd.DataFrame, str, str, str, str]:
    st.sidebar.header("Filters")

    plant_groups = sorted(df["plant_group"].dropna().unique())
    if len(plant_groups) > 1:
        selected_group = st.sidebar.selectbox("Plant Group", plant_groups)
    else:
        selected_group = plant_groups[0]
        st.sidebar.caption(f"Plant group: {selected_group}")

    group_df = df[df["plant_group"] == selected_group]
    plant_key, selected_plant, selected_sr = select_plant(group_df, key="dash_plant")

    plant_df = rows_for_plant(group_df, plant_key)
    module_labels = sorted(
        plant_df["module_label"].dropna().unique(),
        key=lambda item: pd.to_numeric(item, errors="coerce"),
    )
    selected_module = st.sidebar.selectbox("Module Number", module_labels)

    metric_name = st.sidebar.radio("View Metric", list(METRICS), horizontal=False)

    filtered = plant_df[plant_df["module_label"] == selected_module].copy()
    return (
        filtered, selected_group, plant_key, selected_plant, selected_sr,
        selected_module, metric_name,
    )


def plant_options(frame: pd.DataFrame) -> list[tuple[object, str, int | None]]:
    """The plants to choose between, as (plant_key, display name, SR), by name.

    One entry per PLANT, not per plant name. Keying the picker on the name was
    wrong in both directions: several sites run two RO trains whose sheets share a
    name ("Aarti Jhagadia" is both 1957 and 2708), so one entry silently merged two
    separate plants; and a plant that was renamed or misspelled across months
    ("Glenmark Ank RO 1 (1748)" is now "Alivus Life science LTD") appeared as two
    entries, each holding only part of its own history."""
    data = frame.dropna(subset=["plant"]).copy()
    if data.empty:
        return []
    data["plant_key"] = plant_identity_key(data)
    options: list[tuple[object, str, int | None]] = []
    for plant_key, rows in data.groupby("plant_key"):
        srs = rows["plant_sr_no"].dropna()
        options.append(
            (plant_key, current_plant_name(rows), int(srs.iloc[0]) if not srs.empty else None)
        )
    return sorted(options, key=lambda option: (option[1], option[2] or 0))


def plant_option_label(
    option: tuple[object, str, int | None], options: list[tuple[object, str, int | None]]
) -> str:
    """"Name" normally; "Name (SR)" when two plants still share a display name, so
    the two trains can be told apart in the picker."""
    _key, name, sr = option
    shared = sum(1 for _, other, _ in options if other == name) > 1
    if not shared:
        return name
    return f"{name} ({sr})" if sr is not None else f"{name} (no SR)"


def select_plant(frame: pd.DataFrame, key: str) -> tuple[object, str, int | None]:
    """Sidebar plant picker over whole plants. Returns (plant_key, name, SR)."""
    options = plant_options(frame)
    return st.sidebar.selectbox(
        "Plant",
        options,
        format_func=lambda option: plant_option_label(option, options),
        key=key,
    )


def rows_for_plant(frame: pd.DataFrame, plant_key: object) -> pd.DataFrame:
    """Every row of one plant — matched on its identity key, so a sibling train
    sharing the display name is excluded and a spell of rows filed under the
    plant's older name is still included."""
    if frame.empty:
        return frame
    return frame[plant_identity_key(frame) == plant_key]


def stage_uploaded_files(engine: Engine, files: list[object]) -> None:
    """Parse + validate each uploaded file into st.session_state['staged']
    WITHOUT committing. Keyed by sanitized filename so a re-upload overwrites."""
    staged = st.session_state.setdefault("staged", {})
    for uploaded in files:
        data = uploaded.getvalue()
        name = safe_upload_name(getattr(uploaded, "name", "uploaded_report.xlsx"))
        try:
            result = parse_report_bytes(name, data)
            issues = validate_parse(engine, name, result)
        except Exception as exc:  # noqa: BLE001 - surface parse failures in the gate
            staged[name] = {"error": str(exc)}
            continue
        staged[name] = {
            "result": result,
            "issues": issues,
            "content_hash": hashlib.sha256(data).hexdigest(),
            "data": data,
        }


# Map an Issue level to the Streamlit callout that renders it.
ISSUE_RENDERERS = {"ERROR": st.error, "WARN": st.warning, "INFO": st.info}


def render_parse_preview(result: ParseResult) -> None:
    """Show the actual rows a report parsed into — readings, the plant register
    (MIS), and operating parameters — so an uploader can eyeball the extraction
    (SR numbers, zones, flow/conductivity) BEFORE committing anything."""
    readings = result.readings
    mis = result.mis
    parameters = result.parameters

    if not readings.empty:
        cols = [
            "plant", "plant_sr_no", "stage", "module_label", "report_date",
            "install_date", "flow_lph", "conductivity_us_cm", "status",
        ]
        preview = readings.reindex(columns=cols).rename(
            columns={
                "plant": "Plant", "plant_sr_no": "SR No", "stage": "Stage",
                "module_label": "Module", "report_date": "Report",
                "install_date": "Installed", "flow_lph": "Flow (LPH)",
                "conductivity_us_cm": "Cond (µS/cm)", "status": "Status",
            }
        )
        st.caption(f"Readings — {len(preview)} rows")
        st.dataframe(preview, width="stretch", hide_index=True, height=240)

    if not mis.empty:
        # For a register file (the O&M list) this is the whole point of the
        # preview — show the zone tally so 28-vs-4 style gaps are obvious here.
        zone_counts = (
            mis["zone"].map(canonicalize_zone).fillna("Unknown").value_counts().sort_index()
        )
        tally = " · ".join(f"{z}: {n}" for z, n in zone_counts.items())
        st.caption(f"Plant register (MIS) — {len(mis)} rows · {tally}")
        st.dataframe(
            mis.reindex(columns=["zone", "plant_sr_no", "site_name", "status"]).rename(
                columns={"zone": "Zone", "plant_sr_no": "SR No",
                         "site_name": "Site", "status": "Status"}
            ),
            width="stretch", hide_index=True, height=240,
        )

    if not parameters.empty:
        st.caption(f"Operating parameters — {len(parameters)} rows")
        st.dataframe(
            parameters.reindex(columns=["plant", "report_date", "tag", "kind", "value", "unit"]).rename(
                columns={"plant": "Plant", "report_date": "Report", "tag": "Tag",
                         "kind": "Kind", "value": "Value", "unit": "Unit"}
            ),
            width="stretch", hide_index=True, height=200,
        )

    if readings.empty and mis.empty and parameters.empty:
        st.caption("Nothing was extracted from this file.")


def render_staged_review(engine: Engine) -> None:
    """Main-area review of just-uploaded reports: an Excel-like preview of each
    workbook, its quality flags, and the rows that will be committed — with a
    Confirm/Discard pair. Nothing is written until Confirm. Renders nothing when
    no uploads are pending."""
    staged = st.session_state.get("staged", {})
    if not staged:
        return

    st.title("📥 Review uploaded reports")
    st.caption(
        "Preview each workbook exactly as uploaded, check the quality flags, then "
        "Confirm to commit or Discard. Nothing is saved to the database until you Confirm."
    )

    for name in list(staged):
        entry = staged[name]
        with st.container(border=True):
            st.subheader(name)
            if "error" in entry:
                st.error(f"Could not parse: {entry['error']}")
                if st.button("Discard", key=f"discard_{name}"):
                    staged.pop(name, None)
                    rerun_app()
                continue

            result: ParseResult = entry["result"]
            issues: list[Issue] = entry["issues"]
            readings = result.readings
            bypass = int((readings["status"] == "bypass").sum()) if "status" in readings else 0
            dates = readings["report_date"].dropna()
            span = ""
            if not dates.empty:
                lo, hi = dates.min(), dates.max()
                span = f" · {lo:%b %Y}" + (f"–{hi:%b %Y}" if hi != lo else "")
            st.caption(
                f"{result.meta.plant_group} · {result.meta.report_date:%d %b %Y} · "
                f"{plant_identity_key(readings).nunique()} plants · {len(readings)} readings · "
                f"{bypass} bypass{span}"
            )
            for issue in issues:
                ISSUE_RENDERERS.get(issue.level, st.info)(issue.message)

            # The star of the review: the workbook rendered like a real spreadsheet.
            render_excel_workbook(entry["data"], name, key=f"staged_{name}")

            with st.expander("Rows that will be committed to the database"):
                render_parse_preview(result)

            blocked = any(i.level == "ERROR" for i in issues)
            ok_col, no_col = st.columns(2)
            if ok_col.button(
                "✓ Confirm & commit", key=f"commit_{name}", type="primary", disabled=blocked
            ):
                commit_parse_result(engine, name, entry["content_hash"], result)
                store_report_bytes(engine, name, entry["content_hash"], entry["data"])
                load_readings.clear()
                load_parameters.clear()
                load_mis.clear()
                load_plants.clear()
                compute_fleet_status.clear()
                staged.pop(name, None)
                st.success(f"Committed {name}.")
                rerun_app()
            if no_col.button("Discard", key=f"discard_{name}"):
                staged.pop(name, None)
                rerun_app()


def module_series_map(
    plant_df: pd.DataFrame, module_labels: list[str], metric_col: str
) -> dict[str, pd.DataFrame]:
    series_map: dict[str, pd.DataFrame] = {}
    for module_label in module_labels:
        module_df = plant_df[plant_df["module_label"] == module_label]
        series = aggregate_series(module_df, metric_col)
        if not series.empty:
            series_map[module_label] = series
    return series_map


def make_comparison_chart(
    series_map: dict[str, pd.DataFrame],
    *,
    metric_config: dict[str, object],
    average: pd.DataFrame | None,
    pressure: pd.DataFrame | None = None,
) -> go.Figure:
    fig = go.Figure()
    for module_label, series in series_map.items():
        ordered = series.sort_values("report_date")
        fig.add_trace(
            go.Scatter(
                x=ordered["report_date"],
                y=ordered["value"],
                mode="lines+markers",
                name=f"Module {module_label}",
                marker={"size": 7},
                hovertemplate=f"Module {module_label}<br>%{{x|%d %b %Y}}<br>%{{y:,.2f}}<extra></extra>",
            )
        )

    if average is not None and not average.empty:
        ordered = average.sort_values("report_date")
        fig.add_trace(
            go.Scatter(
                x=ordered["report_date"],
                y=ordered["value"],
                mode="lines",
                name="Plant average",
                line={"color": "#475569", "width": 2, "dash": "dash"},
                hovertemplate="Plant average<br>%{x|%d %b %Y}<br>%{y:,.2f}<extra></extra>",
            )
        )

    add_feed_pressure_axis(fig, pressure)

    fig.update_layout(
        title="Module Comparison",
        xaxis_title="Report Date",
        yaxis_title=str(metric_config["axis_label"]),
        hovermode="x unified",
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        template="plotly_white",
        height=470,
    )
    return fig


def build_comparison_table(
    series_map: dict[str, pd.DataFrame], metric_config: dict[str, object]
) -> pd.DataFrame:
    unit = str(metric_config["unit"])
    rows: list[dict[str, object]] = []

    for module_label, series in series_map.items():
        ordered = series.sort_values("report_date")
        latest_row = ordered.iloc[-1]
        latest = float(latest_row["value"])
        rows.append(
            {
                "Module": module_label,
                "Latest": f"{latest:,.2f} {unit}",
                "Latest Date": pd.Timestamp(latest_row["report_date"]).strftime("%d %b %Y"),
                "_sort": pd.to_numeric(module_label, errors="coerce"),
            }
        )

    table = pd.DataFrame(rows)
    if table.empty:
        return table
    return table.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)


def render_comparison_section(
    plant_df: pd.DataFrame,
    selected_module: str,
    metric_name: str,
    metric_config: dict[str, object],
    pressure: pd.DataFrame | None = None,
) -> None:
    module_labels = sorted(
        plant_df["module_label"].dropna().unique(),
        key=lambda item: pd.to_numeric(item, errors="coerce"),
    )
    if len(module_labels) < 2:
        return

    st.markdown("---")
    st.subheader("Compare Modules")
    st.caption(
        f"Overlay multiple modules for {metric_name.lower()} to spot outliers against their peers. "
        "The dashed teal line (right axis) is the shared plant feed pressure (PI 1601) "
        "all modules were operating at."
    )

    default_selection = [selected_module] if selected_module in module_labels else module_labels[:1]
    chosen = st.multiselect(
        "Modules to compare",
        options=module_labels,
        default=default_selection,
        key="compare_modules",
    )
    show_average = st.checkbox("Show plant average", value=True, key="compare_average")

    if not chosen:
        st.info("Select at least one module to compare.")
        return

    metric_col = str(metric_config["column"])
    series_map = module_series_map(plant_df, chosen, metric_col)
    if not series_map:
        st.warning("No valid readings for the selected modules and metric.")
        return

    average = aggregate_series(plant_df, metric_col) if show_average else None

    st.plotly_chart(
        make_comparison_chart(
            series_map, metric_config=metric_config, average=average, pressure=pressure
        ),
        width="stretch",
    )

    table = build_comparison_table(series_map, metric_config)
    if not table.empty:
        st.dataframe(table, width="stretch", hide_index=True)


# The replacement page offers two ways to flag a membrane. "Peer outlier" is
# relative — who stands out above the rest of this stage. "Absolute limit" is a
# fixed operator-set cutoff — who is over an acceptable conductivity outright,
# which catches a whole stage that is uniformly degraded (where a relative test
# finds nothing, because the peer baseline itself is elevated).
PEER_METHOD = "Peer outlier (IQR)"
ABSOLUTE_METHOD = "Absolute limit (uS/cm)"
OUTLIER_METHODS = (PEER_METHOD, ABSOLUTE_METHOD)
DEFAULT_ABSOLUTE_LIMIT = 500.0

# Month-over-month escape hatch (Portfolio degraded flag). The IQR peer test is
# blind to a stage that degrades in lockstep — when every module rises together,
# none stands out as an outlier. So a module the IQR clears is ALSO flagged if
# its conductivity jumps unusually versus its OWN prior-month reading: at least
# +MOM_JUMP_RATIO relative AND +MOM_MIN_DELTA absolute (the floor stops trivial
# rises on near-zero permeate readings, e.g. 60 -> 100 uS/cm, from tripping it).
MOM_JUMP_RATIO = 1.5
MOM_MIN_DELTA = 150.0

# Whole-plant conductivity salt passage. The plant-level instrument block records
# feed conductivity and combined permeate conductivity (us/cm) once per plant
# sheet. Salt passage % = permeate / feed * 100 — the share of feed salinity that
# leaks through the entire RO train. A healthy train passes only a few percent, so
# a plant above this fraction is surfaced for review.
#
# The tag numbering varies by skid: the leading digit is the train (CIS 151, CIS
# 251, CIS 351), and the last two digits name the point — 41/51/71 on the feed
# side, 80 for combined permeate. So BOTH patterns anchor on the "CIS" (in-line
# conductivity) prefix and match the whole tag. Matching a bare "180" anywhere in
# the tag — as this once did — also matched the permeate FLOW tag FIS 180, so a
# flow value mis-typed as us/cm was read as permeate conductivity (Ami Life 3247
# ranked at 89% salt passage on its 5600 lit/hr flow reading). Matching a bare
# "151" likewise missed every plant tagged 141/171/241/251/351, which simply
# never appeared in the ranking.
FEED_COND_TAG_RE = r"^CIS\s*\d*(?:41|51|71)$"
PERMEATE_COND_TAG_RE = r"^CIS\s*\d*80$"
SALT_PASSAGE_FLAG_PCT = 10.0

# Whole-plant permeate (product) flow, logged once per plant sheet as FIS 180 in
# either m3/hr or litres/hr. A sustained month-over-month fall in permeate flow is
# a classic fouling/membrane-loss signal, so the Portfolio surfaces the biggest
# drops. PERMEATE_FLOW_M3HR_PER_LPH converts litres/hr -> m3/hr so plants logged
# in either unit rank on one scale.
# Anchored for the same reason as the conductivity tags: "FIS"/"FI"/"FT" are the
# flow instruments, so a bare "180" would also sweep in PI 180 (pressure) or CIS
# 180 (conductivity) rows whose unit cell was mis-typed as a flow unit.
PERMEATE_FLOW_TAG_RE = r"^F(?:IS|I|T)\s*\d*80$"
PERMEATE_FLOW_M3HR_PER_LPH = 0.001


def iqr_outlier_stats(values: pd.Series, sensitivity: float) -> dict[str, object]:
    """High-side IQR (Tukey) cutoff plus a per-module 'IQRs above Q3' score.

    Conductivity degradation is a *rise*, so only the upper tail matters. The
    median and upper fence are returned in uS/cm (for the chart line and the flag
    test). IQR is robust: the very outliers we are hunting don't inflate the
    cutoff the way a mean + std-dev would. Below 4 readings the quartiles are too
    unstable, so the fence is left NaN and the caller flags nothing.
    """
    clean = pd.to_numeric(values, errors="coerce").dropna()
    n = int(clean.shape[0])
    stats: dict[str, object] = {
        "n": n,
        "median": float(clean.median()) if n else float("nan"),
        "upper_fence": float("nan"),
        "scores": pd.Series(np.zeros(len(values)), index=values.index),
    }
    if n < 4:
        return stats

    q1 = float(clean.quantile(0.25))
    q3 = float(clean.quantile(0.75))
    iqr = q3 - q1
    stats["upper_fence"] = q3 + sensitivity * iqr
    if iqr > 0:
        stats["scores"] = (values - q3) / iqr
    return stats


def make_outlier_chart(
    readings: pd.DataFrame,
    *,
    median: float,
    fence: float,
    unit: str,
    title: str,
    fence_label: str = "Cutoff",
) -> go.Figure:
    colors = ["#dc2626" if flag else "#2563eb" for flag in readings["flag"]]
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=readings["module_label"].astype(str),
            y=readings["conductivity"],
            marker_color=colors,
            customdata=readings["pct_vs_median"],
            hovertemplate=(
                "Module %{x}<br>%{y:,.0f} " + unit
                + "<br>%{customdata:+.0f}% vs stage median<extra></extra>"
            ),
        )
    )
    if pd.notna(median):
        fig.add_hline(
            y=median,
            line_dash="dot",
            line_color="#64748b",
            annotation_text="Stage median",
            annotation_position="top left",
        )
    if pd.notna(fence):
        fig.add_hline(
            y=fence,
            line_dash="dash",
            line_color="#dc2626",
            annotation_text=fence_label,
            annotation_position="top right",
        )
    fig.update_layout(
        title=title,
        xaxis_title="Module",
        yaxis_title=f"Conductivity ({unit})",
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        template="plotly_white",
        height=420,
        showlegend=False,
    )
    fig.update_xaxes(
        type="category",
        categoryorder="array",
        categoryarray=readings["module_label"].astype(str).tolist(),
    )
    return fig


def evaluate_stage_readings(
    readings: pd.DataFrame, *, method: str, sensitivity: float, limit: float
) -> tuple[pd.DataFrame, float, float, bool]:
    """Annotate one stage's per-module readings with flag / score / pct_vs_median.

    Returns (readings, stage_median, cutoff_fence, too_few). The peer baseline is
    per stage, so this is called once per stage when consolidating a whole plant.
    """
    readings = readings.copy()
    median = float(pd.to_numeric(readings["conductivity"], errors="coerce").median())
    n_modules = int(readings["conductivity"].notna().sum())
    too_few = False

    if method == PEER_METHOD:
        stats = iqr_outlier_stats(readings["conductivity"], sensitivity)
        fence = float(stats["upper_fence"])
        readings["score"] = pd.Series(stats["scores"]).to_numpy()
        too_few = n_modules < 4
    else:  # ABSOLUTE_METHOD
        fence = float(limit)
        readings["score"] = readings["conductivity"] / fence if fence else np.nan

    readings["pct_vs_median"] = (
        (readings["conductivity"] / median - 1.0) * 100.0 if median else np.nan
    )
    readings["flag"] = readings["conductivity"] > fence if pd.notna(fence) else False
    readings = readings.sort_values("conductivity", ascending=False).reset_index(drop=True)
    return readings, median, fence, too_few


def render_outlier_section(plant_df: pd.DataFrame) -> None:
    metric_config = METRICS["Conductivity"]
    col = str(metric_config["column"])
    unit = str(metric_config["unit"])

    data = plant_df.dropna(subset=[col]).copy()
    if data.empty:
        st.info("No conductivity readings for this plant.")
        return

    data["stage_label"] = data["stage"].fillna("").map(
        lambda value: str(value).strip() if str(value).strip() else "Unspecified"
    )

    date_options = sorted(data["report_date"].dropna().unique(), reverse=True)
    if not date_options:
        st.info("No dated conductivity readings for this plant.")
        return

    controls = st.columns([2, 3])
    with controls[0]:
        report_date = st.selectbox(
            "Report month",
            date_options,
            format_func=lambda value: pd.Timestamp(value).strftime("%b %Y"),
            key="outlier_date",
        )
    with controls[1]:
        method = st.selectbox("Detection method", OUTLIER_METHODS, key="outlier_method")

    if method == PEER_METHOD:
        sensitivity = st.slider(
            "Sensitivity (k)",
            min_value=1.0,
            max_value=5.0,
            value=1.5,
            step=0.5,
            help="Stricter = fewer flags. IQR cutoff = Q3 + k·IQR (1.5 standard, 3.0 = extreme).",
            key="outlier_k",
        )
        limit = DEFAULT_ABSOLUTE_LIMIT
        score_label = "IQR above Q3"
        score_fmt = lambda value: f"{value:,.1f}"
        cutoff_label = "IQR Fence"
    else:  # ABSOLUTE_METHOD
        sensitivity = 1.5
        limit = st.number_input(
            f"Replace above ({unit})",
            min_value=0.0,
            value=DEFAULT_ABSOLUTE_LIMIT,
            step=50.0,
            help="Flag every module over this fixed conductivity, regardless of its peers. "
            "Use when a whole stage is degraded and the peer test finds nothing.",
            key="outlier_limit",
        )
        score_label = "x Limit"
        score_fmt = lambda value: f"{value:,.2f}x"
        cutoff_label = "Replacement Limit"

    snapshot = data[data["report_date"] == report_date]
    month_text = pd.Timestamp(report_date).strftime("%b %Y")
    stages_present = sorted(snapshot["stage_label"].unique())

    # Evaluate every stage (each against its own peers) and collect the flags.
    per_stage: dict[str, tuple[pd.DataFrame, float, float, bool]] = {}
    flagged_records: list[dict[str, object]] = []
    total_modules = 0
    too_few_stages: list[str] = []

    for stage in stages_present:
        stage_readings = snapshot[snapshot["stage_label"] == stage].groupby(
            "module_label", as_index=False
        ).agg(
            conductivity=(col, "mean"),
            flow=("flow_lph", "mean"),
            install_date=("install_date", "first"),
        )
        if stage_readings.empty:
            continue
        evaluated, median, fence, too_few = evaluate_stage_readings(
            stage_readings, method=method, sensitivity=sensitivity, limit=limit
        )
        per_stage[stage] = (evaluated, median, fence, too_few)
        total_modules += int(evaluated["conductivity"].notna().sum())
        if too_few:
            too_few_stages.append(stage)
        for _, row in evaluated[evaluated["flag"]].iterrows():
            flagged_records.append(
                {
                    "Stage": stage,
                    "Module": row["module_label"],
                    "conductivity": float(row["conductivity"]),
                    "pct_vs_median": float(row["pct_vs_median"]),
                    "score": float(row["score"]),
                    "install_date": row["install_date"],
                }
            )

    # ----- Consolidated headline + table (the whole plant, this month) -----
    n_flagged = len(flagged_records)
    card1, card2, card3 = st.columns(3)
    with card1:
        flag_color = "#dc2626" if n_flagged else "#16a34a"
        metric_card("Modules Needing Replacement", str(n_flagged), month_text, flag_color)
    with card2:
        metric_card("Modules Evaluated", str(total_modules), f"{len(stages_present)} stage(s)")
    with card3:
        cutoff_text = f"{limit:,.0f} {unit}" if method == ABSOLUTE_METHOD else f"k = {sensitivity:g}"
        metric_card("Method", method.split(" (")[0], cutoff_text)

    if too_few_stages:
        st.info(
            "Too few modules (<4) for a reliable peer test in: "
            + ", ".join(too_few_stages)
            + ". Use an absolute limit to flag those stages."
        )

    if flagged_records:
        consolidated = pd.DataFrame(flagged_records).sort_values(
            ["Stage", "conductivity"], ascending=[True, False]
        )
        display_table = pd.DataFrame(
            {
                "Stage": consolidated["Stage"],
                "Module": consolidated["Module"],
                f"Conductivity ({unit})": consolidated["conductivity"].map(lambda v: f"{v:,.0f}"),
                "vs Stage Median": consolidated["pct_vs_median"].map(lambda v: f"{v:+,.0f}%"),
                score_label: consolidated["score"].map(score_fmt),
                "Install Date": consolidated["install_date"].map(
                    lambda v: pd.Timestamp(v).strftime("%d %b %Y") if pd.notna(v) else "Unknown"
                ),
            }
        )
        st.markdown(
            f"**{n_flagged} module(s) need replacement** across "
            f"{consolidated['Stage'].nunique()} stage(s) — {month_text}:"
        )
        st.dataframe(display_table, width="stretch", hide_index=True)
        st.download_button(
            "Download replacement list (CSV)",
            display_table.to_csv(index=False).encode("utf-8"),
            file_name=f"replacement_candidates_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )
    else:
        st.success(f"No modules need replacement in any stage for {month_text}.")

    # ----- Per-stage drill-down chart -----
    if per_stage:
        st.markdown("---")
        st.markdown("**Stage detail**")
        stage = st.selectbox("Stage to chart", list(per_stage), key="outlier_stage")
        evaluated, median, fence, _ = per_stage[stage]
        st.plotly_chart(
            make_outlier_chart(
                evaluated,
                median=median,
                fence=fence,
                unit=unit,
                title=f"{stage} conductivity by module — {month_text}",
                fence_label=cutoff_label,
            ),
            width="stretch",
        )


def select_group_plant(df: pd.DataFrame) -> tuple[str, object, str, int | None]:
    """Sidebar plant picker for the standalone Replacement page (mirrors the
    dashboard's group/plant selectors, with its own widget keys)."""
    st.sidebar.header("Plant")
    groups = sorted(df["plant_group"].dropna().unique())
    if len(groups) > 1:
        group = st.sidebar.selectbox("Plant Group", groups, key="rep_group")
    else:
        group = groups[0]
        st.sidebar.caption(f"Plant group: {group}")
    return group, *select_plant(df[df["plant_group"] == group], key="rep_plant")


def render_replacement_page(df: pd.DataFrame) -> None:
    st.title("Membrane Replacement Candidates")
    st.caption(
        "For the selected plant and month, every stage is checked and the flagged "
        "membranes are consolidated into one list with a total count — flagged either "
        "for standing out above their stage peers (conductivity outliers) or for being "
        "over a fixed acceptable limit. Each stage is judged on its own readings, so a "
        "naturally high-TDS stage isn't penalised against a cleaner one."
    )
    group, plant_key, plant, plant_sr_no = select_group_plant(df)
    st.caption(f"{group} | {plant_label_with_sr(plant, plant_sr_no)}")
    plant_df = rows_for_plant(df[df["plant_group"] == group], plant_key)
    render_outlier_section(plant_df)


@st.cache_data(show_spinner="Building the fleet view...")
def compute_fleet_status(readings: pd.DataFrame, mis: pd.DataFrame) -> pd.DataFrame:
    """Annotate every reading with bypass / degraded / need, for all months.

    A module is degraded if EITHER signal fires:
    - **Peer outlier (IQR):** its conductivity exceeds the stage's Tukey upper
      fence (Q3 + 1.5·IQR), computed against peers in the same stage that month
      via evaluate_stage_readings (k=1.5). Robust down to ~4 modules per stage.
    - **Month-over-month jump:** its conductivity rose unusually versus the SAME
      module's prior reading (>= MOM_JUMP_RATIO and >= MOM_MIN_DELTA). This is a
      per-module time-series test, so it catches a stage degrading in lockstep —
      the case the within-month IQR peer test is structurally blind to.

    A module "needs attention" if it is bypassed or degraded. Returns one row per
    (plant, stage, module, month), with degraded_iqr / degraded_mom kept split so
    callers can show *why* a module was flagged.
    """
    col = str(METRICS["Conductivity"]["column"])
    flow_col = str(METRICS["Flow Rate"]["column"])
    columns = [
        "plant_group", "plant", "plant_sr_no", "plant_key", "zone", "stage_label",
        "module_label", "report_date", "conductivity", "prev_conductivity",
        "flow", "prev_flow", "install_date", "status", "degraded_iqr",
        "degraded_mom", "degraded", "need", "cutoff", "source_file",
    ]
    if readings.empty:
        return pd.DataFrame(columns=columns)

    data = readings.copy()
    data["stage_label"] = data["stage"].fillna("").map(
        lambda value: str(value).strip() if str(value).strip() else "Unspecified"
    )
    # A plant IS its Plant SR No. Several sites log two RO trains under one sheet
    # name ("Aarti Jhagadia" is both 1957 and 2708), so every grouping below keys
    # on plant_key. Keying on the name pooled two separate trains into one stage
    # peer group — the IQR cutoff that decides "degraded" was computed across
    # modules from different plants — and stamped both with whichever SR happened
    # to come first. The name is kept for display only.
    data["plant_key"] = plant_identity_key(data)

    # plant_sr_no -> zone via the shared resolver (canonicalized, latest MIS row
    # that names a zone) so the dashboard's zones match the rest of the app and a
    # blank later report can't blank a plant's known zone.
    zone_by_sr = build_zone_by_sr(mis)

    plant_meta: dict[object, tuple[object, str, object, str]] = {}
    for plant_key, pdf in data.groupby("plant_key"):
        plant_group = pdf["plant_group"].iloc[0]
        plant = current_plant_name(pdf)
        srs = pdf["plant_sr_no"].dropna()
        plant_sr = int(srs.iloc[0]) if not srs.empty else None
        zone = zone_by_sr.get(plant_sr) if plant_sr is not None else None
        plant_meta[plant_key] = (plant_group, plant, plant_sr, zone if zone else "Unknown")

    records: list[dict[str, object]] = []
    for (plant_key, stage_label, report_date), group in data.groupby(
        ["plant_key", "stage_label", "report_date"]
    ):
        agg = group.groupby("module_label", as_index=False).agg(
            conductivity=(col, "mean"),
            flow=(flow_col, "mean"),
            install_date=("install_date", "first"),
            is_bypass=("status", lambda s: bool((s == "bypass").any())),
            # The workbook this month's reading was parsed from, so a module row can
            # link straight back to its own IMR in the preview.
            source_file=("source_file", "first"),
        )
        active = agg[~agg["is_bypass"]]
        cutoff = float("nan")
        iqr_map: dict[object, bool] = {}
        if not active.empty:
            evaluated, _, cutoff, _ = evaluate_stage_readings(
                active, method=PEER_METHOD, sensitivity=1.5, limit=0.0
            )
            iqr_map = dict(zip(evaluated["module_label"], evaluated["flag"]))

        plant_group, plant, plant_sr, zone = plant_meta[plant_key]
        for _, row in agg.iterrows():
            bypass = bool(row["is_bypass"])
            degraded_iqr = (not bypass) and bool(iqr_map.get(row["module_label"], False))
            records.append(
                {
                    "plant_group": plant_group,
                    "plant": plant,
                    "plant_sr_no": plant_sr,
                    "plant_key": plant_key,
                    "zone": zone,
                    "stage_label": stage_label,
                    "module_label": row["module_label"],
                    "report_date": report_date,
                    "conductivity": row["conductivity"],
                    "flow": row["flow"],
                    "install_date": row["install_date"],
                    "status": "bypass" if bypass else "active",
                    "degraded_iqr": degraded_iqr,
                    "cutoff": cutoff,
                    "source_file": row["source_file"],
                }
            )

    result = pd.DataFrame(records)
    if result.empty:
        return pd.DataFrame(columns=columns)

    # Second signal: an unusually high month-over-month jump vs the SAME module's
    # prior reading. Computed per (plant_key, stage, module) time series — blind to
    # within-stage peers, so it catches a stage that degrades in lockstep, the
    # case IQR misses. shift() compares each reading to that module's previous
    # available report (no flag on a module's first-ever month).
    result = result.sort_values(["plant_key", "stage_label", "module_label", "report_date"])
    grouped = result.groupby(["plant_key", "stage_label", "module_label"])
    prev = grouped["conductivity"].shift()
    result["prev_conductivity"] = prev
    result["prev_flow"] = grouped["flow"].shift()
    result["degraded_mom"] = (
        (result["status"] == "active")
        & prev.notna()
        & (result["conductivity"] >= prev * MOM_JUMP_RATIO)
        & (result["conductivity"] - prev >= MOM_MIN_DELTA)
    )
    result["degraded"] = result["degraded_iqr"] | result["degraded_mom"]
    result["need"] = (result["status"] == "bypass") | result["degraded"]

    return result.reindex(columns=columns).reset_index(drop=True)


def build_zone_by_sr(mis: pd.DataFrame) -> dict[object, object]:
    """plant_sr_no -> zone, from the most recent MIS row for each plant that
    actually names a zone. Rows with a blank/missing zone are dropped BEFORE
    picking the latest, so a later report that omits the zone can't overwrite the
    register's known zone with "Unknown". Zones are canonicalized so casing slips
    ("VADODARA" vs "Vadodara") collapse to one."""
    zone_by_sr: dict[object, object] = {}
    if mis is None or mis.empty:
        return zone_by_sr
    zoned = mis.copy()
    zoned["zone"] = zoned["zone"].map(canonicalize_zone)
    zoned = zoned[zoned["zone"].notna()]
    if zoned.empty:
        return zone_by_sr
    latest_mis = zoned.sort_values("report_date").drop_duplicates("plant_sr_no", keep="last")
    for _, row in latest_mis.iterrows():
        zone_by_sr[row["plant_sr_no"]] = row["zone"]
    return zone_by_sr


def resolve_plant_sr(plants: pd.Series, readings: pd.DataFrame | None) -> pd.Series:
    """Fallback plant_sr_no for a Series of plant display names.

    Only for rows ingested before `parameters` carried its own plant_sr_no —
    `plant_level_frame()` prefers the stored SR and calls this for the rest. It
    looks up the SR the readings already resolved for that name (the SAME SR the
    rest of the app joins on), then falls back to parsing it out of the name;
    parsing alone fails for the newer sheet-naming ("Cadila pharma. ANK", "AARTI
    PHASE I VAPI"), which is what left the zone and Plant No columns blank.

    A name is resolved ONLY when it maps to exactly one SR. Several sites run two
    RO trains logged under one display name — "Aarti Jhagadia" is both 1957 and
    2708 — and this used to keep whichever row Postgres happened to return last,
    stamping one train's readings with the other's Plant SR No. An ambiguous name
    now resolves to NA (shown as a blank Plant No), because no number at all beats
    a number the client would act on."""
    sr_by_plant_name: dict[object, int] = {}
    if readings is not None and not readings.empty and "plant_sr_no" in readings:
        resolved = readings.dropna(subset=["plant_sr_no"])
        unique_srs = resolved.groupby("plant")["plant_sr_no"].unique()
        sr_by_plant_name = {
            name: int(srs[0]) for name, srs in unique_srs.items() if len(srs) == 1
        }

    def lookup(name: object) -> object:
        sr = sr_by_plant_name.get(name)
        if sr is not None:
            return sr
        parsed = plant_sr_no_from_name(name)
        return parsed if parsed is not None else pd.NA

    return plants.map(lookup).astype("Int64")


def plant_level_frame(
    frame: pd.DataFrame, mis: pd.DataFrame, readings: pd.DataFrame | None
) -> pd.DataFrame:
    """Give a plant-level parameter frame its identity columns: a trustworthy
    `plant_sr_no`, a `plant_key` to group on, and the plant's `zone`.

    A plant IS its SR number, not its display name. Sites run more than one RO
    train under a single name, so every plant-level rollup (salt passage, permeate
    flow, permeate conductivity) must dedup and group on `plant_key` — the SR
    where we have one, the name only as a last resort for rows whose SR we could
    never resolve. Grouping on the name merged two physically separate trains into
    one row, mixing one plant's feed with another's permeate.

    Rows carry their own plant_sr_no once they have been re-ingested; anything
    older falls back to the name-based lookup."""
    out = frame.copy()
    stored = (
        pd.to_numeric(out["plant_sr_no"], errors="coerce").astype("Int64")
        if "plant_sr_no" in out.columns
        else pd.Series(pd.NA, index=out.index, dtype="Int64")
    )
    out["plant_sr_no"] = stored.fillna(resolve_plant_sr(out["plant"], readings))
    out["plant_key"] = plant_identity_key(out)

    zone_by_sr = build_zone_by_sr(mis)
    out["zone"] = out["plant_sr_no"].map(
        lambda sr: zone_by_sr.get(int(sr)) if pd.notna(sr) else None
    ).fillna("Unknown")
    return out


def compute_salt_passage(
    parameters: pd.DataFrame, mis: pd.DataFrame, readings: pd.DataFrame | None = None
) -> pd.DataFrame:
    """Per-plant per-month conductivity salt passage from the CIS feed/permeate tags.

    Feed conductivity (CIS ?41/?51/?71) and combined permeate conductivity
    (CIS ?80) are recorded once per plant sheet. Salt passage % = permeate / feed
    * 100 — the share of feed salinity that leaks through the whole RO train.
    Returns one row per (plant, month) that carries BOTH readings, with the
    plant's zone joined in, sorted by passage descending.

    Every step keys on `plant_key` (the Plant SR No), never the display name —
    two RO trains at one site share a name, and grouping by name paired one
    train's feed with the other's permeate and dropped the loser's row entirely."""
    columns = [
        "plant_group", "plant", "plant_sr_no", "zone",
        "report_date", "feed", "permeate", "passage_pct",
    ]
    if parameters is None or parameters.empty:
        return pd.DataFrame(columns=columns)

    cond = parameters[parameters["kind"] == "conductivity"].copy()
    if cond.empty:
        return pd.DataFrame(columns=columns)

    tag = cond["tag"].astype(str).str.strip().str.upper()
    cond["role"] = pd.NA
    cond.loc[tag.str.match(FEED_COND_TAG_RE), "role"] = "feed"
    cond.loc[tag.str.match(PERMEATE_COND_TAG_RE), "role"] = "permeate"
    cond = cond.dropna(subset=["role"])
    if cond.empty:
        return pd.DataFrame(columns=columns)

    cond = plant_level_frame(cond, mis, readings)
    # One value per (plant, month, role); last write wins on any duplicate.
    cond = cond.drop_duplicates(["plant_key", "report_date", "role"], keep="last")
    # Pivot on plant_key alone, then merge the identity columns back on. Carrying
    # plant_sr_no through the pivot index would drop every plant whose SR is still
    # unresolved, since a NA index entry is silently dropped by the groupby.
    wide = cond.pivot(
        index=["plant_key", "report_date"], columns="role", values="value"
    ).reset_index()
    # Identity per (plant_key, month), so each row is labelled with the name its
    # OWN report used rather than one picked from some other month.
    identity = cond.drop_duplicates(["plant_key", "report_date"], keep="last")[
        ["plant_key", "report_date", "plant_group", "plant", "plant_sr_no", "zone"]
    ]
    wide = wide.merge(identity, on=["plant_key", "report_date"], how="left")
    if "feed" not in wide.columns or "permeate" not in wide.columns:
        return pd.DataFrame(columns=columns)
    wide = wide.dropna(subset=["feed", "permeate"])
    wide = wide[wide["feed"] > 0]
    if wide.empty:
        return pd.DataFrame(columns=columns)
    wide["passage_pct"] = wide["permeate"] / wide["feed"] * 100.0

    return (
        wide.reindex(columns=columns)
        .sort_values("passage_pct", ascending=False)
        .reset_index(drop=True)
    )


def compute_permeate_flow(
    parameters: pd.DataFrame, mis: pd.DataFrame, readings: pd.DataFrame | None = None
) -> pd.DataFrame:
    """Per-plant per-month permeate (product) flow from the FIS 180 tag, in m3/hr.

    Permeate flow is logged once per plant sheet as FIS ?80, in either m3/hr or
    litres/hr; litres/hr is converted to m3/hr so every plant ranks on one scale.
    Returns one row per (plant, month) with the month-over-month change vs the
    plant's OWN prior reading (`fall_m3hr` positive = a drop), zone joined in.
    """
    columns = [
        "plant_group", "plant", "plant_sr_no", "zone", "report_date",
        "flow_m3hr", "prev_flow_m3hr", "fall_m3hr", "change_pct",
    ]
    if parameters is None or parameters.empty:
        return pd.DataFrame(columns=columns)

    flow = parameters[parameters["kind"] == "flow"].copy()
    if flow.empty:
        return pd.DataFrame(columns=columns)
    flow = flow[flow["tag"].astype(str).str.strip().str.upper().str.match(PERMEATE_FLOW_TAG_RE)]
    if flow.empty:
        return pd.DataFrame(columns=columns)

    # Normalize litres/hr -> m3/hr (m3/hr passes through unchanged).
    unit = flow["unit"].astype(str)
    factor = pd.Series(1.0, index=flow.index)
    factor[unit.str.contains("lit")] = PERMEATE_FLOW_M3HR_PER_LPH
    flow["flow_m3hr"] = pd.to_numeric(flow["value"], errors="coerce") * factor
    flow = flow.dropna(subset=["flow_m3hr"])
    flow = flow[flow["flow_m3hr"] > 0]
    if flow.empty:
        return pd.DataFrame(columns=columns)

    # One value per (plant, month); last write wins, then a per-plant time series.
    # Keyed on plant_key (the SR) so two RO trains sharing a site name keep their
    # own series instead of collapsing into one and shifting against each other.
    flow = plant_level_frame(flow, mis, readings)
    flow = flow.drop_duplicates(["plant_key", "report_date"], keep="last")
    flow = flow.sort_values(["plant_key", "report_date"])
    prev = flow.groupby("plant_key")["flow_m3hr"].shift()
    flow["prev_flow_m3hr"] = prev
    flow["fall_m3hr"] = prev - flow["flow_m3hr"]
    flow["change_pct"] = (flow["flow_m3hr"] / prev - 1.0) * 100.0

    return flow.reindex(columns=columns).reset_index(drop=True)


def compute_permeate_conductivity(
    parameters: pd.DataFrame, mis: pd.DataFrame, readings: pd.DataFrame | None = None
) -> pd.DataFrame:
    """Per-plant per-month permeate conductivity (CIS ?80) with its MoM rise.

    Permeate (product) conductivity is logged once per plant sheet as CIS ?80
    (180 on train 1, 280 on train 2, ...).
    Returns one row per (plant, month) with the month-over-month change vs the
    plant's OWN prior reading (`rise` positive = product water got saltier — a
    plant-wide membrane / feed signal), zone joined in.
    """
    columns = [
        "plant_group", "plant", "plant_sr_no", "zone", "report_date",
        "permeate", "prev_permeate", "rise", "change_pct",
    ]
    if parameters is None or parameters.empty:
        return pd.DataFrame(columns=columns)

    cond = parameters[parameters["kind"] == "conductivity"].copy()
    if cond.empty:
        return pd.DataFrame(columns=columns)
    cond = cond[cond["tag"].astype(str).str.strip().str.upper().str.match(PERMEATE_COND_TAG_RE)]
    if cond.empty:
        return pd.DataFrame(columns=columns)

    cond["permeate"] = pd.to_numeric(cond["value"], errors="coerce")
    cond = cond.dropna(subset=["permeate"])
    cond = cond[cond["permeate"] > 0]
    if cond.empty:
        return pd.DataFrame(columns=columns)

    # One value per (plant, month); last write wins, then a per-plant time series.
    # Keyed on plant_key (the SR) for the same reason as the flow series.
    cond = plant_level_frame(cond, mis, readings)
    cond = cond.drop_duplicates(["plant_key", "report_date"], keep="last")
    cond = cond.sort_values(["plant_key", "report_date"])
    prev = cond.groupby("plant_key")["permeate"].shift()
    cond["prev_permeate"] = prev
    cond["rise"] = cond["permeate"] - prev
    cond["change_pct"] = (cond["permeate"] / prev - 1.0) * 100.0

    return cond.reindex(columns=columns).reset_index(drop=True)


def build_plant_ranking(snapshot: pd.DataFrame, latest: pd.Timestamp) -> pd.DataFrame:
    """One row per PLANT — keyed on plant_key (the Plant SR No), so a site running
    two RO trains under one sheet name ranks as the two plants it is, each with its
    own Plant No, instead of one merged row stamped with an arbitrary SR."""
    rows: list[dict[str, object]] = []
    for plant_key, pdf in snapshot.groupby("plant_key"):
        plant = current_plant_name(pdf)
        module_count = len(pdf)
        bypassed = int((pdf["status"] == "bypass").sum())
        degraded = int(pdf["degraded"].sum())
        need = int(pdf["need"].sum())
        avg_cond = pdf.loc[pdf["status"] == "active", "conductivity"].mean()
        ages = (
            pd.Timestamp(latest) - pd.to_datetime(pdf["install_date"], errors="coerce")
        ).dt.days / 365.25
        rows.append(
            {
                "Plant": plant,
                "Plant No": plant_no_column(pdf["plant_sr_no"]).iloc[0],
                "plant_key": plant_key,
                "Zone": pdf["zone"].iloc[0],
                "Modules": module_count,
                "Bypassed": bypassed,
                "Degraded": degraded,
                "Need": need,
                "Need %": round(need / module_count * 100, 1) if module_count else 0.0,
                "Avg Cond (uS/cm)": round(float(avg_cond), 0) if pd.notna(avg_cond) else np.nan,
                "Median Age (yrs)": round(float(ages.median()), 1) if ages.notna().any() else np.nan,
            }
        )
    ranking = pd.DataFrame(rows)
    if ranking.empty:
        return ranking
    return ranking.sort_values("Need", ascending=False).reset_index(drop=True)


def make_fleet_trend_chart(status: pd.DataFrame) -> go.Figure:
    months = sorted(status["report_date"].dropna().unique())
    degraded_pct: list[float] = []
    avg_cond: list[float] = []
    for month in months:
        sub = status[status["report_date"] == month]
        total = len(sub)
        degraded_pct.append(float(sub["degraded"].sum()) / total * 100 if total else np.nan)
        avg_cond.append(sub.loc[sub["status"] == "active", "conductivity"].mean())

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=months, y=degraded_pct, name="Degraded %", mode="lines+markers",
            line={"color": "#dc2626", "width": 3},
            hovertemplate="%{x|%b %Y}<br>%{y:.1f}% degraded<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=months, y=avg_cond, name="Avg conductivity (uS/cm)", mode="lines+markers",
            yaxis="y2", line={"color": "#2563eb", "width": 2, "dash": "dash"},
            hovertemplate="%{x|%b %Y}<br>%{y:,.0f} uS/cm avg<extra></extra>",
        )
    )
    fig.update_layout(
        title="Fleet deterioration over time",
        xaxis_title="Report month",
        yaxis_title="Degraded %",
        yaxis2={"title": "Avg conductivity (uS/cm)", "overlaying": "y", "side": "right", "showgrid": False},
        hovermode="x unified",
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        template="plotly_white",
        height=420,
    )
    return fig


def make_zone_rollup_chart(snapshot: pd.DataFrame) -> go.Figure:
    rollup = (
        snapshot.groupby("zone")
        .agg(
            bypassed=("status", lambda s: int((s == "bypass").sum())),
            degraded=("degraded", "sum"),
        )
        .reset_index()
    )
    rollup["total"] = rollup["bypassed"] + rollup["degraded"]
    rollup = rollup.sort_values("total", ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(x=rollup["zone"], y=rollup["degraded"], name="Degraded", marker_color="#d97706"))
    fig.add_trace(go.Bar(x=rollup["zone"], y=rollup["bypassed"], name="Bypassed", marker_color="#dc2626"))
    fig.update_layout(
        barmode="stack",
        title="Modules needing attention by zone",
        xaxis_title="Zone",
        yaxis_title="Modules",
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        template="plotly_white",
        height=380,
    )
    return fig


def make_age_profile_chart(snapshot: pd.DataFrame, latest: pd.Timestamp) -> go.Figure | None:
    years = pd.to_datetime(snapshot["install_date"], errors="coerce").dt.year.dropna()
    if years.empty:
        return None
    counts = years.astype(int).value_counts().sort_index()
    latest_year = pd.Timestamp(latest).year
    labels = [str(year) for year in counts.index]
    colors = ["#dc2626" if (latest_year - year) > 5 else "#2563eb" for year in counts.index]
    hover = "%{x}<br>%{customdata:,} modules<extra></extra>"

    fig = go.Figure()
    # Full-height backdrop behind each column. A year with 2 modules next to one with
    # 500 is a hairline bar that is near-impossible to hit; the backdrop makes the whole
    # column a click target, and it reports the same year/count as the real bar.
    fig.add_trace(
        go.Bar(
            x=labels,
            y=[int(counts.max())] * len(counts),
            customdata=counts.values,
            marker_color="rgba(37, 99, 235, 0.07)",
            hovertemplate=hover,
            width=0.9,
        )
    )
    fig.add_trace(
        go.Bar(
            x=labels,
            y=counts.values,
            customdata=counts.values,
            marker_color=colors,
            hovertemplate=hover,
            text=[f"{v:,}" for v in counts.values],
            textposition="outside",
            cliponaxis=False,
            width=0.9,
        )
    )
    fig.update_layout(
        title="Membrane age profile — install year of fitted modules (>5 yrs old in red)",
        xaxis_title="Install year",
        yaxis_title="Modules",
        margin={"l": 20, "r": 20, "t": 60, "b": 20},
        template="plotly_white",
        height=380,
        showlegend=False,
        barmode="overlay",
        bargap=0.12,
        hovermode="closest",
        # Click-to-select is the point of this chart; drag-to-zoom only gets in the way.
        dragmode=False,
    )
    fig.update_xaxes(type="category")
    # Headroom so the tallest bar's count label is not clipped by the plot edge.
    fig.update_yaxes(range=[0, int(counts.max()) * 1.12])
    return fig


# st.switch_page() needs the actual Page object, and the pages are built inside
# main(); this lets one page navigate to another by name. Repopulated every run.
PAGES: dict[str, object] = {}


def render_open_imr_button(row: pd.Series, *, key: str) -> None:
    """For one selected module row, offer a jump to the IMR Preview showing the
    workbook that reading was parsed from.

    `row` must come from the fleet-status frame (not a display table), so it still
    carries `source_file` / `plant_sr_no`. The request is handed to the preview page
    through session state; that page decides whether the file is still previewable.
    """
    plant = str(row.get("plant") or "")
    sr = row.get("plant_sr_no")
    source = row.get("source_file")
    stage = str(row.get("stage_label") or "")
    module = str(row.get("module_label") or "")
    when = row.get("report_date")
    month = pd.Timestamp(when).strftime("%b %Y") if pd.notna(when) else ""

    who = f"{plant}{f' (SR {int(sr)})' if pd.notna(sr) else ''}"
    st.markdown(f"**{who}** · {stage} · Module {module}{f' · {month}' if month else ''}")

    if not source or pd.isna(source):
        st.info("No source workbook is recorded for this reading, so there is nothing to open.")
        return

    if st.button(
        f"📗 Open {source} in IMR Preview",
        key=f"open_imr_{key}",
        type="primary",
        help="Opens the raw workbook this reading came from, with the plant pre-searched.",
    ):
        st.session_state["preview_open_request"] = {
            "source_file": str(source),
            # Search by SR No when we have one — it is unambiguous where a site name
            # may be spelled differently across the sheet, MIS, and register.
            "query": str(int(sr)) if pd.notna(sr) else plant,
        }
        target = PAGES.get("preview")
        if target is None:  # navigation not built yet (shouldn't happen in-app)
            st.error("The IMR Preview page is unavailable.")
        else:
            st.switch_page(target)


def flag_reason(row: pd.Series) -> str:
    """Why this module is on a flagged list — bypassed, or which degradation
    signal(s) fired. Shared by every drill-down so the wording never diverges."""
    if row["status"] == "bypass":
        return "Bypassed"
    tags = []
    if row["degraded_iqr"]:
        tags.append("Peer outlier (IQR)")
    if row["degraded_mom"]:
        tags.append("MoM jump")
    return " + ".join(tags) if tags else "Degraded"


def build_flagged_detail_table(rows: pd.DataFrame, latest: pd.Timestamp) -> pd.DataFrame:
    """Every number behind a flag, one row per module, so a drill-down list can be
    read and exported without opening the workbook. Bypassed modules have no flow or
    conductivity by definition — those cells read "—" rather than being dropped."""
    def num(series: pd.Series, fmt: str) -> pd.Series:
        return series.map(lambda v: format(v, fmt) if pd.notna(v) else "—")

    install = pd.to_datetime(rows["install_date"], errors="coerce")
    age_years = (pd.Timestamp(latest) - install).dt.days / 365.25
    delta = rows["conductivity"] - rows["prev_conductivity"]

    return pd.DataFrame(
        {
            "Plant": rows["plant"],
            "Plant No": plant_no_column(rows["plant_sr_no"]),
            "Zone": rows["zone"],
            "Stage": rows["stage_label"],
            "Module": rows["module_label"],
            "Reason": rows.apply(flag_reason, axis=1),
            "Conductivity (uS/cm)": num(rows["conductivity"], ",.0f"),
            "Prev Month (uS/cm)": num(rows["prev_conductivity"], ",.0f"),
            "Δ Conductivity": delta.map(lambda v: f"{v:+,.0f}" if pd.notna(v) else "—"),
            "Stage Fence (uS/cm)": num(rows["cutoff"], ",.0f"),
            "Flow (LPH)": num(rows["flow"], ",.1f"),
            "Prev Flow (LPH)": num(rows["prev_flow"], ",.1f"),
            "Install Date": install.dt.strftime("%d %b %Y").fillna("Unknown"),
            "Age (yrs)": age_years.map(lambda v: f"{v:,.1f}" if pd.notna(v) else "—"),
            "Report": rows["source_file"],
        }
    )


def render_flagged_module_list(
    snapshot: pd.DataFrame, latest: pd.Timestamp, month_text: str, *, bypassed: bool
) -> None:
    """The full fleet-wide list behind the Degraded / Bypassed KPI cards.

    `bypassed=True` lists offline modules; otherwise the degraded (still-active)
    ones. Sorted worst-first by conductivity for degraded, and by plant for
    bypassed (which have no reading to rank on).
    """
    if bypassed:
        rows = snapshot[snapshot["status"] == "bypass"]
        title, anchor = "Bypassed modules", "bypassed-modules"
        caption = (
            "Every module taken offline this month. A bypass is the strongest "
            "replacement signal there is — the plant has already stopped using it."
        )
        order, ascending = ["plant", "stage_label", "module_label"], True
        empty = "No bypassed modules"
    else:
        rows = snapshot[snapshot["degraded"]]
        title, anchor = "Degraded modules", "degraded-modules"
        caption = (
            "Every still-active module flagged this month — either a peer outlier "
            "against its own stage (IQR) or an unusual month-over-month jump."
        )
        order, ascending = ["conductivity"], False
        empty = "No degraded modules"

    st.markdown("---")
    st.subheader(f"{title} — {month_text}", anchor=anchor)
    if rows.empty:
        st.success(f"{empty} in {month_text}.")
        return

    st.caption(f"{len(rows):,} module(s). {caption}")
    # Sort the source frame so a selected row index still points at the same module.
    rows = rows.sort_values(order, ascending=ascending).reset_index(drop=True)
    table = build_flagged_detail_table(rows, latest)

    st.caption("Select a row to open the IMR that module's reading came from.")
    key = "bypassed" if bypassed else "degraded"
    event = st.dataframe(
        table,
        width="stretch",
        hide_index=True,
        height=440,
        on_select="rerun",
        selection_mode="single-row",
        key=f"portfolio_{key}_list",
    )
    picked = event.selection.get("rows", []) if event else []
    if picked:
        render_open_imr_button(rows.iloc[picked[0]], key=f"{key}_list")

    st.download_button(
        f"Download {title.lower()} (CSV)",
        table.to_csv(index=False).encode("utf-8"),
        file_name=f"{key}_modules_{month_text.replace(' ', '_')}.csv",
        mime="text/csv",
    )


def render_plant_flagged_modules(
    snapshot: pd.DataFrame, plant_key: object, plant: str, month_text: str
) -> None:
    """List every module flagged (degraded or bypassed) for one plant this month.

    `snapshot` is the already-zone/month-filtered fleet status, so degraded/need
    and the per-stage IQR cutoff are read straight off it — no recomputation.
    Selected by `plant_key` (the Plant SR No), so a site with two RO trains under
    one name lists only the train whose row was clicked.
    """
    plant_rows = snapshot[snapshot["plant_key"] == plant_key]
    srs = plant_rows["plant_sr_no"].dropna().unique()
    heading = plant_label_with_sr(plant, int(srs[0]) if len(srs) else None)

    flagged = plant_rows[plant_rows["need"]].copy()
    st.markdown(f"#### {heading} — flagged modules · {month_text}")
    if flagged.empty:
        st.success(f"No modules flagged for {plant} in {month_text}.")
        return

    n_degraded = int(flagged["degraded"].sum())
    n_bypass = int((flagged["status"] == "bypass").sum())
    n_iqr = int(flagged["degraded_iqr"].sum())
    n_mom = int(flagged["degraded_mom"].sum())
    st.caption(
        f"{len(flagged):,} module(s) need attention — {n_degraded:,} degraded "
        f"({n_iqr:,} peer outlier · {n_mom:,} month-over-month jump) "
        f"+ {n_bypass:,} bypassed."
    )

    flagged = flagged.sort_values(
        ["status", "conductivity"], ascending=[True, False]
    ).reset_index(drop=True)
    detail = pd.DataFrame(
        {
            "Stage": flagged["stage_label"],
            "Module": flagged["module_label"],
            "Reason": flagged.apply(flag_reason, axis=1),
            "Conductivity (uS/cm)": flagged["conductivity"].map(
                lambda v: f"{v:,.0f}" if pd.notna(v) else "—"
            ),
            "Prev Month (uS/cm)": flagged["prev_conductivity"].map(
                lambda v: f"{v:,.0f}" if pd.notna(v) else "—"
            ),
            "Stage Fence (uS/cm)": flagged["cutoff"].map(
                lambda v: f"{v:,.0f}" if pd.notna(v) else "—"
            ),
            "Install Date": flagged["install_date"].map(
                lambda v: pd.Timestamp(v).strftime("%d %b %Y") if pd.notna(v) else "Unknown"
            ),
        }
    )
    st.caption("Select a row to open the IMR that module's reading came from.")
    detail_event = st.dataframe(
        detail,
        width="stretch",
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        # Keyed per plant so switching plants doesn't carry a stale row selection.
        key=f"flagged_modules::{plant}",
    )
    detail_rows = detail_event.selection.get("rows", []) if detail_event else []
    if detail_rows:
        render_open_imr_button(flagged.iloc[detail_rows[0]], key=f"flagged::{plant}")
    st.download_button(
        "Download flagged modules (CSV)",
        detail.to_csv(index=False).encode("utf-8"),
        file_name=f"{plant.replace(' ', '_')}_flagged_{month_text.replace(' ', '_')}.csv",
        mime="text/csv",
        key="portfolio_plant_flagged_csv",
    )


def render_portfolio_page(df: pd.DataFrame, params: pd.DataFrame, mis: pd.DataFrame) -> None:
    st.title("Fleet Portfolio")
    st.caption(
        "Fleet-wide membrane health across every plant — no plant selection needed. "
        "A module needs attention if it is bypassed, or degraded — an active module "
        "that is a peer outlier in its stage (Q3 + 1.5·IQR around the stage median) "
        "or shows an unusually high month-over-month jump in conductivity."
    )

    status = compute_fleet_status(df, mis)
    if status.empty:
        st.info("No readings are available to build the fleet view.")
        return

    # Month + zone slicers — the whole page renders for the chosen month and the
    # chosen zones, defaulting to the most recent month and all zones. Every
    # downstream section keys off `snapshot`/`selected`/`status_zoned`.
    months = sorted(status["report_date"].dropna().unique(), reverse=True)
    month_labels = [pd.Timestamp(m).strftime("%b %Y") for m in months]
    zones = sorted(status["zone"].dropna().astype(str).unique().tolist())
    picker_col, zone_col = st.columns([1, 2])
    with picker_col:
        picked = st.selectbox("Report month", month_labels, index=0)
    with zone_col:
        selected_zones = st.multiselect(
            "Zones",
            zones,
            default=zones,
            help="Filter the whole page to one or more zones. Clear the box to show all zones.",
        )
    selected = months[month_labels.index(picked)]
    month_text = picked
    active_zones = selected_zones or zones  # empty selection = all zones
    status_zoned = status[status["zone"].isin(active_zones)]
    snapshot = status_zoned[status_zoned["report_date"] == selected]

    # Counted on plant_key: a site running two RO trains under one sheet name is
    # two plants, and the name would count it as one.
    total_plants = int(snapshot["plant_key"].nunique())
    total_modules = len(snapshot)
    active_modules = int((snapshot["status"] == "active").sum())
    degraded = int(snapshot["degraded"].sum())
    bypassed = int((snapshot["status"] == "bypass").sum())
    need = int(snapshot["need"].sum())
    need_pct = need / total_modules * 100 if total_modules else 0.0

    # Last month's demand (modules to replace) for the same zones, and the delta.
    # The previous month is the next entry in `months` (sorted newest-first).
    picked_index = month_labels.index(picked)
    prev_month = months[picked_index + 1] if picked_index + 1 < len(months) else None
    if prev_month is not None:
        prev_need = int(
            status_zoned.loc[status_zoned["report_date"] == prev_month, "need"].sum()
        )
        prev_month_text = pd.Timestamp(prev_month).strftime("%b %Y")
        demand_delta = need - prev_need
    else:
        prev_need = None
        prev_month_text = "—"
        demand_delta = None

    # ----- 1. Headline KPI: modules to replace, paired with the membrane count -----
    # Each module holds a fixed number of membranes, so the membranes to replace is
    # the modules-to-replace figure (the hero number) times that per-module count.
    membranes = MEMBRANES_PER_MODULE * need
    hero_col, mem_col = st.columns([3, 1])
    with hero_col:
        hero_card(
            "Modules to Replace",
            f"{need:,}",
            f"{month_text} — {degraded:,} degraded + {bypassed:,} bypassed "
            f"across {total_plants:,} plants ({need_pct:.1f}% of {total_modules:,} modules)",
        )
    with mem_col:
        # A calmer companion to the hero — same card, blue accent, smaller number.
        st.markdown(
            f"""
            <div class="hero-card" style="border-color:#bfdbfe;
                 background:linear-gradient(180deg,#eff6ff 0%,#ffffff 70%);height:100%;">
                <div class="hero-title">Membranes to Replace</div>
                <div class="hero-value" style="color:#2563eb;
                     font-size:clamp(1.9rem,3.5vw,3rem);">{membranes:,}</div>
                <div class="hero-subtitle">{MEMBRANES_PER_MODULE} × {need:,} modules</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ----- 2. Supporting KPI cards -----
    st.markdown("")
    kpis = st.columns(6)
    with kpis[0]:
        metric_card("Plants", f"{total_plants:,}", "in fleet")
    with kpis[1]:
        metric_card("Active Modules", f"{active_modules:,}", f"{total_modules:,} total this month")
    with kpis[2]:
        metric_card(
            "Degraded", f"{degraded:,}", "click for the full list", "#d97706",
            anchor="degraded-modules",
        )
    with kpis[3]:
        metric_card(
            "Bypassed", f"{bypassed:,}", "click for the full list", "#dc2626",
            anchor="bypassed-modules",
        )
    with kpis[4]:
        metric_card(
            "Last Month Demand",
            f"{prev_need:,}" if prev_need is not None else "—",
            f"modules to replace · {prev_month_text}",
        )
    with kpis[5]:
        if demand_delta is None:
            metric_card("Δ vs Last Month", "—", "no prior month")
        else:
            # More modules to replace than last month is bad (red); fewer is good (green).
            delta_color = (
                "#dc2626" if demand_delta > 0 else "#16a34a" if demand_delta < 0 else "#0f172a"
            )
            metric_card(
                "Δ vs Last Month",
                f"{demand_delta:+,}",
                f"vs {prev_need:,} in {prev_month_text}",
                delta_color,
            )

    # ----- Jump navigation: clickable cards that scroll to each section -----
    st.markdown("")
    jump_nav(
        [
            ("Plant ranking", "plant-ranking", "per-plant attention"),
            ("Degraded modules", "degraded-modules", "full list"),
            ("Bypassed modules", "bypassed-modules", "full list"),
            ("Worst 50 modules", "worst-modules", "highest conductivity"),
            ("Conductivity rises", "cond-rises", "module MoM"),
            ("Flow drops", "flow-drops", "module MoM"),
            ("Salt passage", "salt-passage", "sites > 10%"),
            ("Permeate cond. rises", "permeate-cond-rises", "site MoM"),
            ("Permeate flow drops", "permeate-flow-drops", "site MoM"),
            ("Age profile", "age-profile", "install years"),
        ]
    )

    # ----- 2. Plant ranking (centerpiece) -----
    st.markdown("---")
    st.subheader("Plant ranking", anchor="plant-ranking")
    st.caption(
        "One row per plant, sorted by modules needing attention. Click a column header "
        "to re-sort, or select a row to see that plant's flagged modules below."
    )
    ranking = build_plant_ranking(snapshot, selected)
    # plant_key is the join back to the snapshot for the drill-down, not something
    # to show — the reader already has "Plant No". Row positions are unchanged, so
    # a selected row still indexes into `ranking`.
    ranking_display = ranking.drop(columns=["plant_key"], errors="ignore")
    ranking_event = st.dataframe(
        ranking_display,
        width="stretch",
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key="portfolio_ranking",
    )
    st.download_button(
        "Download plant ranking (CSV)",
        ranking_display.to_csv(index=False).encode("utf-8"),
        file_name=f"fleet_ranking_{month_text.replace(' ', '_')}.csv",
        mime="text/csv",
    )

    # Drill-down: clicking a plant row lists every module flagged for it this month.
    selected_rows = ranking_event.selection.get("rows", []) if ranking_event else []
    if selected_rows and not ranking.empty:
        chosen = ranking.iloc[selected_rows[0]]
        render_plant_flagged_modules(
            snapshot, chosen["plant_key"], str(chosen["Plant"]), month_text
        )

    # ----- 2b. The lists behind the Degraded / Bypassed KPI cards -----
    render_flagged_module_list(snapshot, selected, month_text, bypassed=False)
    render_flagged_module_list(snapshot, selected, month_text, bypassed=True)

    # ----- 3. Fleet trend -----
    st.markdown("---")
    st.subheader("Fleet trend")
    st.caption("Month-over-month deterioration across the selected zones.")
    st.plotly_chart(make_fleet_trend_chart(status_zoned), width="stretch")

    # ----- 4 & 5. Zone rollup + worst modules -----
    left, right = st.columns(2)
    with left:
        st.subheader("By zone")
        st.plotly_chart(make_zone_rollup_chart(snapshot), width="stretch")
    with right:
        st.subheader(f"Worst 50 modules — {month_text}", anchor="worst-modules")
        active_snap = snapshot[snapshot["status"] == "active"].dropna(subset=["conductivity"])
        worst = active_snap.sort_values("conductivity", ascending=False).head(50)
        worst_table = pd.DataFrame(
            {
                "Plant": worst["plant"],
                "Plant No": plant_no_column(worst["plant_sr_no"]),
                "Stage": worst["stage_label"],
                "Module": worst["module_label"],
                "Conductivity (uS/cm)": worst["conductivity"].map(lambda v: f"{v:,.0f}"),
                "Install Date": worst["install_date"].map(
                    lambda v: pd.Timestamp(v).strftime("%d %b %Y") if pd.notna(v) else "Unknown"
                ),
            }
        )
        st.dataframe(worst_table, width="stretch", hide_index=True, height=600)

    # ----- 5b. Biggest month-over-month conductivity rises -----
    st.markdown("---")
    st.subheader(f"Top 50 conductivity rises — {month_text}", anchor="cond-rises")
    st.caption(
        "Active modules whose conductivity rose the most versus their own prior "
        "reading. A jump shared across a whole plant/stage usually points to a feed "
        "or cleaning issue rather than individual membranes."
    )
    jumps = active_snap.dropna(subset=["prev_conductivity"]).copy()
    jumps["delta"] = jumps["conductivity"] - jumps["prev_conductivity"]
    jumps = jumps[jumps["delta"] > 0].sort_values("delta", ascending=False).head(50)
    if jumps.empty:
        st.info("No month-over-month increases to show (no prior-month readings yet).")
    else:
        jump_table = pd.DataFrame(
            {
                "Plant": jumps["plant"],
                "Plant No": plant_no_column(jumps["plant_sr_no"]),
                "Stage": jumps["stage_label"],
                "Module": jumps["module_label"],
                "Prev (uS/cm)": jumps["prev_conductivity"].map(lambda v: f"{v:,.0f}"),
                "Now (uS/cm)": jumps["conductivity"].map(lambda v: f"{v:,.0f}"),
                "Change (uS/cm)": jumps["delta"].map(lambda v: f"+{v:,.0f}"),
                "Change (x)": (jumps["conductivity"] / jumps["prev_conductivity"]).map(
                    lambda v: f"{v:,.1f}x"
                ),
            }
        )
        st.dataframe(jump_table, width="stretch", hide_index=True, height=600)
        st.download_button(
            "Download conductivity rises (CSV)",
            jump_table.to_csv(index=False).encode("utf-8"),
            file_name=f"mom_cond_rises_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ----- 5c. Biggest month-over-month flow drops -----
    st.markdown("---")
    st.subheader(f"Top 50 flow drops — {month_text}", anchor="flow-drops")
    st.caption(
        "Active modules whose flow fell the most versus their own prior reading. "
        "A falling permeate flow is a classic fouling / membrane-loss signal."
    )
    drops = active_snap.dropna(subset=["flow", "prev_flow"]).copy()
    drops["delta"] = drops["prev_flow"] - drops["flow"]
    drops = drops[drops["delta"] > 0].sort_values("delta", ascending=False).head(50)
    if drops.empty:
        st.info("No month-over-month flow drops to show (no prior-month readings yet).")
    else:
        drop_table = pd.DataFrame(
            {
                "Plant": drops["plant"],
                "Plant No": plant_no_column(drops["plant_sr_no"]),
                "Stage": drops["stage_label"],
                "Module": drops["module_label"],
                "Prev (L/hr)": drops["prev_flow"].map(lambda v: f"{v:,.0f}"),
                "Now (L/hr)": drops["flow"].map(lambda v: f"{v:,.0f}"),
                "Change (L/hr)": drops["delta"].map(lambda v: f"-{v:,.0f}"),
                "Change (%)": (drops["flow"] / drops["prev_flow"] - 1.0).map(
                    lambda v: f"{v * 100:,.0f}%"
                ),
            }
        )
        st.dataframe(drop_table, width="stretch", hide_index=True, height=600)
        st.download_button(
            "Download flow drops (CSV)",
            drop_table.to_csv(index=False).encode("utf-8"),
            file_name=f"mom_flow_drops_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ----- 5d. Highest conductivity salt passage (feed vs permeate) -----
    st.markdown("---")
    st.subheader(
        f"Top 15 sites — salt passage > {SALT_PASSAGE_FLAG_PCT:.0f}% — {month_text}",
        anchor="salt-passage",
    )
    st.caption(
        "Conductivity salt passage = permeate conductivity (CIS 180) ÷ feed "
        "conductivity (CIS 151) × 100 — the share of feed salinity leaking through "
        "the whole RO train. A healthy train passes only a few percent; a high "
        "figure means the membranes are letting salts through plant-wide."
    )
    passage = compute_salt_passage(params, mis, df)
    passage = passage[passage["zone"].isin(active_zones)]
    passage = passage[passage["report_date"] == selected]
    passage = passage[passage["passage_pct"] > SALT_PASSAGE_FLAG_PCT]
    passage = passage.sort_values("passage_pct", ascending=False).head(15)
    if passage.empty:
        st.info(
            f"No site exceeds {SALT_PASSAGE_FLAG_PCT:.0f}% salt passage this month "
            "(or feed/permeate conductivity wasn't recorded in these workbooks)."
        )
    else:
        passage_table = pd.DataFrame(
            {
                "Plant": passage["plant"],
                "Plant No": plant_no_column(passage["plant_sr_no"]),
                "Zone": passage["zone"],
                "Feed (uS/cm)": passage["feed"].map(lambda v: f"{v:,.0f}"),
                "Permeate (uS/cm)": passage["permeate"].map(lambda v: f"{v:,.0f}"),
                "Salt passage": passage["passage_pct"].map(lambda v: f"{v:.1f}%"),
            }
        )
        st.dataframe(passage_table, width="stretch", hide_index=True)
        st.download_button(
            "Download salt passage (CSV)",
            passage_table.to_csv(index=False).encode("utf-8"),
            file_name=f"salt_passage_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ----- 5e. Highest month-over-month rise in permeate conductivity -----
    st.markdown("---")
    st.subheader(
        f"Top 10 sites — permeate conductivity rise — {month_text}",
        anchor="permeate-cond-rises",
    )
    st.caption(
        "Whole-plant permeate (product) conductivity from CIS 180, versus the "
        "plant's own prior month. A site-wide rise means the product water got "
        "saltier across the train — a feed, cleaning, or membrane-integrity signal."
    )
    perm_rise = compute_permeate_conductivity(params, mis, df)
    perm_rise = perm_rise[perm_rise["zone"].isin(active_zones)]
    perm_rise = perm_rise[perm_rise["report_date"] == selected]
    perm_rise = perm_rise.dropna(subset=["prev_permeate"])
    perm_rise = perm_rise[perm_rise["rise"] > 0]
    perm_rise = perm_rise.sort_values("rise", ascending=False).head(10)
    if perm_rise.empty:
        st.info(
            "No site shows a permeate-conductivity rise this month "
            "(or CIS 180 wasn't recorded with a prior month to compare)."
        )
    else:
        perm_table = pd.DataFrame(
            {
                "Plant": perm_rise["plant"],
                "Plant No": plant_no_column(perm_rise["plant_sr_no"]),
                "Zone": perm_rise["zone"],
                "Prev (uS/cm)": perm_rise["prev_permeate"].map(lambda v: f"{v:,.0f}"),
                "Now (uS/cm)": perm_rise["permeate"].map(lambda v: f"{v:,.0f}"),
                "Rise (uS/cm)": perm_rise["rise"].map(lambda v: f"+{v:,.0f}"),
                "Change": perm_rise["change_pct"].map(lambda v: f"{v:,.0f}%"),
            }
        )
        st.dataframe(perm_table, width="stretch", hide_index=True)
        st.download_button(
            "Download permeate conductivity rises (CSV)",
            perm_table.to_csv(index=False).encode("utf-8"),
            file_name=f"permeate_cond_rises_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ----- 5f. Biggest fall in permeate flow (month over month) -----
    st.markdown("---")
    st.subheader(
        f"Top 10 sites — biggest permeate-flow drop — {month_text}",
        anchor="permeate-flow-drops",
    )
    st.caption(
        "Whole-plant permeate (product) flow from FIS 180, normalized to m3/hr, "
        "versus the plant's own prior month. A sustained fall in permeate flow is "
        "a classic fouling / membrane-loss signal."
    )
    flow_mom = compute_permeate_flow(params, mis, df)
    flow_mom = flow_mom[flow_mom["zone"].isin(active_zones)]
    flow_mom = flow_mom[flow_mom["report_date"] == selected]
    flow_mom = flow_mom.dropna(subset=["prev_flow_m3hr"])
    flow_mom = flow_mom[flow_mom["fall_m3hr"] > 0]
    flow_mom = flow_mom.sort_values("fall_m3hr", ascending=False).head(10)
    if flow_mom.empty:
        st.info(
            "No site shows a permeate-flow drop this month "
            "(or FIS 180 flow wasn't recorded with a prior month to compare)."
        )
    else:
        flow_table = pd.DataFrame(
            {
                "Plant": flow_mom["plant"],
                "Plant No": plant_no_column(flow_mom["plant_sr_no"]),
                "Zone": flow_mom["zone"],
                "Prev (m3/hr)": flow_mom["prev_flow_m3hr"].map(lambda v: f"{v:,.2f}"),
                "Now (m3/hr)": flow_mom["flow_m3hr"].map(lambda v: f"{v:,.2f}"),
                "Fall (m3/hr)": flow_mom["fall_m3hr"].map(lambda v: f"-{v:,.2f}"),
                "Change": flow_mom["change_pct"].map(lambda v: f"{v:,.1f}%"),
            }
        )
        st.dataframe(flow_table, width="stretch", hide_index=True)
        st.download_button(
            "Download permeate-flow drops (CSV)",
            flow_table.to_csv(index=False).encode("utf-8"),
            file_name=f"permeate_flow_drops_{month_text.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ----- 6. Membrane age profile (click a bar to list its modules) -----
    st.markdown("---")
    st.subheader("Membrane age profile", anchor="age-profile")
    age_chart = make_age_profile_chart(snapshot, selected)
    if age_chart is None:
        st.info("No install dates available to build the age profile.")
    else:
        st.caption(
            "Click anywhere in a year's column — not just the bar itself — to list the "
            "modules installed that year. Shift-click to add more years."
        )
        age_event = st.plotly_chart(
            age_chart,
            width="stretch",
            on_select="rerun",
            selection_mode="points",
            key="portfolio_age_profile",
        )
        picked_points = age_event.selection.get("points", []) if age_event else []
        picked_years = {str(p.get("x")) for p in picked_points if p.get("x") is not None}
        if picked_years:
            install_year = pd.to_datetime(
                snapshot["install_date"], errors="coerce"
            ).dt.year
            # Sort the source frame, not the display table, so a selected row index
            # still points at the same module in `cohort` (which carries source_file).
            cohort = (
                snapshot[install_year.astype("Int64").astype(str).isin(picked_years)]
                .sort_values(["plant", "stage_label", "module_label"])
                .reset_index(drop=True)
            )
            year_text = ", ".join(sorted(picked_years))
            st.markdown(f"**Modules installed in {year_text}** — {len(cohort):,} fitted")
            st.caption("Select a row to open the IMR that module's reading came from.")
            cohort_table = pd.DataFrame(
                {
                    "Plant": cohort["plant"],
                    "Plant No": plant_no_column(cohort["plant_sr_no"]),
                    "Zone": cohort["zone"],
                    "Stage": cohort["stage_label"],
                    "Module": cohort["module_label"],
                    "Install Date": pd.to_datetime(
                        cohort["install_date"], errors="coerce"
                    ).dt.strftime("%d %b %Y"),
                    "Status": cohort["status"],
                }
            )
            cohort_event = st.dataframe(
                cohort_table,
                width="stretch",
                hide_index=True,
                height=360,
                on_select="rerun",
                selection_mode="single-row",
                key="portfolio_age_cohort",
            )
            cohort_rows = cohort_event.selection.get("rows", []) if cohort_event else []
            if cohort_rows:
                render_open_imr_button(cohort.iloc[cohort_rows[0]], key="age_cohort")
            st.download_button(
                "Download these modules (CSV)",
                cohort_table.to_csv(index=False).encode("utf-8"),
                file_name=f"modules_installed_{year_text.replace(', ', '_')}.csv",
                mime="text/csv",
            )


def render_dashboard(df: pd.DataFrame, params: pd.DataFrame) -> None:
    (
        filtered, selected_group, plant_key, selected_plant, selected_sr,
        selected_module, metric_name,
    ) = sidebar_filters(df)
    metric_config = METRICS[metric_name]
    metric_col = str(metric_config["column"])
    series = aggregate_series(filtered, metric_col)

    st.title("RO Plant Membrane Health Dashboard")
    st.caption(f"{selected_group} | {plant_label_with_sr(selected_plant, selected_sr)}")

    if series.empty:
        st.warning("No valid numeric readings are available for this module and metric.")
        return

    # PI 1601 feed pressure is the operating pressure each reading was taken at.
    # It is plant-level (shared across modules), overlaid on a secondary axis.
    pressure_by_date = feed_pressure_series(
        params, selected_group, selected_plant, selected_sr
    )

    ordered = series.sort_values("report_date")
    latest_row = ordered.iloc[-1]
    first_row = ordered.iloc[0]
    latest = float(latest_row["value"])
    first = float(first_row["value"])
    unit = str(metric_config["unit"])
    latest_date = pd.Timestamp(latest_row["report_date"]).strftime("%d %b %Y")

    change = latest - first
    if first:
        change_subtitle = f"{change / first * 100:+,.1f}% since {pd.Timestamp(first_row['report_date']).strftime('%b %Y')}"
    else:
        change_subtitle = f"since {pd.Timestamp(first_row['report_date']).strftime('%b %Y')}"

    col1, col2, col3 = st.columns(3)
    with col1:
        metric_card("Latest Reading", f"{latest:,.2f} {unit}", latest_date)
    with col2:
        metric_card("Change Since First", f"{change:+,.2f} {unit}", change_subtitle)
    with col3:
        metric_card("Readings on Record", str(len(ordered)), "report months")

    st.plotly_chart(
        make_chart(
            series,
            metric_name=metric_name,
            metric_config=metric_config,
            pressure=pressure_by_date,
        ),
        width="stretch",
    )

    with st.expander("Cleaned readings"):
        display = filtered[
            [
                "source_file",
                "report_date",
                "plant",
                "plant_sr_no",
                "stage",
                "module_label",
                "install_date",
                "flow_lph",
                "conductivity_us_cm",
            ]
        ].assign(
            plant_sr_no=lambda d: plant_no_column(d["plant_sr_no"])
        ).rename(
            columns={
                "source_file": "Source File",
                "report_date": "Report Date",
                "plant": "Plant",
                "plant_sr_no": "Plant No",
                "stage": "Stage",
                "module_label": "Module Number",
                "install_date": "Install Date",
                "flow_lph": "Total flow liter/hr.",
                "conductivity_us_cm": "Cond. us/cm",
            }
        )
        st.dataframe(display, width="stretch", hide_index=True)

    plant_df = rows_for_plant(df[df["plant_group"] == selected_group], plant_key)
    render_comparison_section(
        plant_df, selected_module, metric_name, metric_config, pressure_by_date
    )


def render_data_manager(engine: Engine) -> None:
    """Sidebar admin panel: upload reports through a validation gate, review the
    pending queue, and permanently remove a bad report. Shown on every page."""
    with st.sidebar.expander("🗂 Manage Data"):
        # --- Upload + validate (no commit until Confirm) ---
        uploaded_files = st.file_uploader(
            "Upload monthly Excel report",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            help="Files are parsed and checked; nothing is saved until you Confirm.",
            key="data_uploader",
        )
        if st.button("Validate & preview", disabled=not uploaded_files):
            stage_uploaded_files(engine, uploaded_files or [])
            rerun_app()

        pending = st.session_state.get("staged", {})
        if pending:
            st.info(
                f"📥 {len(pending)} file(s) awaiting review in the main panel → "
                "check the Excel preview, then Confirm or Discard."
            )

        # --- Re-parse a stored report through the same gate ---
        stored = report_files_index(engine)
        if stored:
            st.markdown("---")
            reparse_choice = st.selectbox(
                "Re-parse a stored report",
                stored,
                index=None,
                placeholder="Pick a file to re-parse…",
                help="Re-runs the current parser on the stored bytes, back into the review queue.",
                key="reparse_choice",
            )
            if reparse_choice and st.button("Re-parse", key="reparse_btn"):
                data = load_report_bytes(engine, reparse_choice)
                if data is None:
                    st.warning("No stored bytes for that file.")
                else:
                    result = parse_report_bytes(reparse_choice, data)
                    st.session_state.setdefault("staged", {})[reparse_choice] = {
                        "result": result,
                        "issues": validate_parse(engine, reparse_choice, result),
                        "content_hash": hashlib.sha256(data).hexdigest(),
                        "data": data,
                    }
                    rerun_app()

        st.markdown("---")

        # --- Remove an already-committed report ---
        reports = ingested_report_summary(engine)
        if reports.empty:
            st.caption("No reports have been committed yet.")
            return

        st.caption(f"{len(reports)} report(s) in the database.")
        choice = st.selectbox(
            "Remove a report",
            reports["filename"].tolist(),
            index=None,
            placeholder="Select a file to remove…",
            key="remove_report_choice",
        )
        if not choice:
            return

        row = reports[reports["filename"] == choice].iloc[0]
        st.caption(
            f"{int(row['n_readings'] or 0)} readings · "
            f"{int(row['n_parameters'] or 0)} parameters · "
            f"{int(row['n_mis'] or 0)} MIS rows"
        )
        confirm = st.checkbox(
            f"Confirm permanent removal of **{choice}**",
            key="remove_report_confirm",
        )
        if st.button(
            "Delete report data", type="primary", disabled=not confirm
        ):
            counts = remove_report(engine, APP_DIR, choice)
            load_readings.clear()
            load_parameters.clear()
            load_mis.clear()
            compute_fleet_status.clear()
            st.success(
                f"Removed {choice}: {counts['readings']} readings, "
                f"{counts['parameters']} parameters, {counts['mis']} MIS rows."
            )
            rerun_app()


# --------------------------------------------------------------------------- #
# Scan IMR: handwritten printout photo -> Gemini Vision -> filled template.
# Zero-cost path: a free Google AI Studio key + gemini-2.5-flash. The rate limit
# (~a handful per minute) is fine for one-IMR-at-a-time entry.
# --------------------------------------------------------------------------- #
SCAN_ZONES = ["Ahmedabad", "Vadodara", "Ankleshwar", "Panoli", "Jhagadia", "Dahej", "Vapi"]
SCAN_MODEL = "gemini-2.5-flash"

# The scan page emits its workbook by GENERATING a sheet shaped to the data,
# rather than pouring the data into a fixed template grid. The old fixed grid had
# room for three stages of 23 modules and silently dropped everything past that —
# a IV STAGE outright, and 37 of Superform's 60 third-stage modules.
#
# Generating is safe because the workbook is transport, not an archive: the
# operator reviews and corrects every value in the on-page editor, and the file
# exists only to carry that reviewed data into the SAME block parser that reads
# real IMR workbooks. That parser is driven entirely by cell text and relative
# position (find_header_row -> find_module_blocks -> extract_rows_from_sheet,
# plus read_cover_block and extract_operating_parameters) and never looks at
# fonts, borders or merges — so the sheet is emitted plain, and the only thing
# that must be preserved verbatim is the header/label/unit WORDING below.
SCAN_BLOCK_WIDTH = 7          # 6 columns per stage + a spacer, as on the paper form
SCAN_COVER_ROWS = (3, 4, 5)   # within read_cover_block's first-14-rows search window
SCAN_STAGE_LABEL_ROW = 7      # find_stage_label scans up from the header row
SCAN_HEADER_ROW = 8
SCAN_DATA_ROW = 9

# Column headers, verbatim: is_module_header / is_install_date_header /
# is_time_header / is_flow_header / is_conductivity_header key off these words.
SCAN_BLOCK_HEADERS = [
    "Mo no.", "Inst Date", "Time for ______ ml",
    "Colour/ Non Colour", "Total flow liter/hr.", "Cond.    us/cm",
]

# Presentation, lifted from the printed form so a generated sheet still reads as
# an IMR. Sizing and styling are independent concerns: the grid is built to fit
# the data (any stage count, any module count) and then dressed with the styles
# below, rather than inheriting a look and a 3x23 cage together from a file.
SCAN_COL_WIDTHS = [8.7, 9.7, 13.0, 10.1, 8.7, 10.7]  # per stage block
SCAN_SPACER_WIDTH = 2.0
SCAN_ROW_H = {"title": 40.0, "cover": 21.0, "band": 18.0, "header": 33.75, "data": 18.0}
SCAN_VALUE_FILL = "FFFDE7"  # pale yellow on fill-in cells, as on the form

# extract_operating_parameters identifies a reading by its UNIT cell, so a
# parameter written without one is invisible to the parser. The extractor only
# returns (tag, value), so the unit is restored here — by exact tag where the
# form names it, else by instrument prefix so a tag we haven't seen still lands.
SCAN_PARAM_UNITS = {
    "ph 141": "-", "ti 151": "deg. C", "reject cond": "us/cm",
    "feed flow": "m3/hr", "permeat flow": "m3/hr", "permeate flow": "m3/hr",
}
SCAN_PARAM_UNIT_BY_PREFIX = {
    "pi": "bar", "cis": "us/cm", "ci": "us/cm",
    "fis": "lit/hrs.", "fi": "lit/hrs.", "ti": "deg. C", "ph": "-",
}


def scan_param_unit(tag: object) -> str | None:
    """The unit to write beside an extracted operating parameter, or None if the
    tag is unrecognizable (in which case it can't round-trip and is skipped)."""
    norm = normalize_text(tag)
    if not norm:
        return None
    if norm in SCAN_PARAM_UNITS:
        return SCAN_PARAM_UNITS[norm]
    prefix = norm.split()[0] if norm.split() else ""
    return SCAN_PARAM_UNIT_BY_PREFIX.get(prefix)

EXTRACTION_PROMPT = """You are reading a handwritten "Individual Module Report (IMR)" for a
reverse-osmosis plant. Return ONLY JSON matching exactly this shape:

{
  "plant_sr_no": <integer or null>,
  "report_date": "<YYYY-MM-DD or null>",
  "zone": "<one of: Ahmedabad, Vadodara, Ankleshwar, Panoli, Jhagadia, Dahej, Vapi, or null>",
  "site_name": "<string or null>",
  "plant_capacity": "<string or null>",
  "stages": [
    {"stage_label": "<e.g. I STAGE, II STAGE, III STAGE>",
     "volume_ml": <number or null, the volume printed in the "Time for ___ ml" header, in millilitres; 1 ltr = 1000>,
     "modules": [
       {"mo_no": <int>, "inst_date": "<YYYY-MM-DD or null>",
        "time_sec": <number or null>,
        "flow": <number or null>, "cond": <number or null>,
        "remark": "<string or null, e.g. BY PASS>"}
     ]}
  ],
  "parameters": [ {"tag": "<e.g. CIS - 151 -, FIS - 180 -, PI - 1601 ->", "value": <number>} ]
}

Rules:
- Transcribe digits exactly as written; do not invent or 'correct' values.
- If a cell is blank or unreadable, use null (do not guess).
- DECIMAL POINTS ARE CRITICAL. Many numbers here are decimals — preserve the
  decimal point exactly where written and NEVER drop it or merge the digits.
  A stopwatch time written "11.47" is the number 11.47, NOT 1147; "8.53" is 8.53,
  not 853. If you see a dot (or a small gap) between digits, keep it as a decimal.
- 'time_sec' is the "Time for ___ ml" column (the stopwatch reading in seconds).
  These are SMALL decimal numbers, almost always between about 3 and 90 seconds
  (e.g. 11.47, 8.53, 21.4). A bare 3-4 digit integer like 1147 is a decimal whose
  point was missed — write it as the decimal (11.47).
- 'volume_ml' is the millilitre volume named in that column's header (e.g. 500, 1000).
- 'flow' is the Total flow (liter/hr) column. It is DERIVED, not measured:
  flow = volume_ml / 1000 / (time_sec / 3600) = volume_ml * 3.6 / time_sec. It is
  usually left blank on the form because it's computed from the time — so return
  null for 'flow' unless a number is actually handwritten in that column, and let
  the app compute it from time_sec and volume_ml.
- 'cond' is the Cond. us/cm column.
- Keep stage labels and parameter tags verbatim as printed on the form.
Return only the JSON object, no prose."""


def _secret_key(env_var: str, section: str) -> str | None:
    key = os.environ.get(env_var)
    if not key:
        try:
            key = st.secrets[section]["api_key"]
        except Exception:  # noqa: BLE001 - secrets file / section may be absent
            key = None
    return key or None


def get_gemini_key() -> str | None:
    return _secret_key("GEMINI_API_KEY", "gemini")


def get_groq_key() -> str | None:
    return _secret_key("GROQ_API_KEY", "groq")


class GeminiRateLimited(Exception):
    """Free-tier quota / rate limit hit (429) — retrying soon won't help."""


class GeminiUnavailable(Exception):
    """Model is transiently overloaded (503/500/UNAVAILABLE) after retries."""


class GeminiTruncated(Exception):
    """The reply hit the output ceiling and stopped mid-JSON, so it won't parse.

    Worth its own type because the raw symptom is a json.loads error deep in the
    string ("Expecting property name ... char 3225"), which reads like a model
    quality problem when it is really a budget problem."""


# When the primary model is overloaded, fall back to a second model alias whose
# pool is usually separate. Same family, still free.
SCAN_FALLBACK_MODEL = "gemini-flash-latest"


def _coerce_extracted(obj: object) -> dict:
    """Models sometimes wrap the JSON object in an array (e.g. `[{...}]`) despite
    the prompt. Normalize any list to the first dict element so callers always get
    the single object they expect."""
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, list):
        for item in obj:
            if isinstance(item, dict):
                return item
    return {}


def extract_imr_from_images(
    images: list[tuple[bytes, str]], model: str, api_key: str
) -> dict:
    """Send one or more form photos/PDFs to Gemini and return the parsed JSON dict.

    Resilient to the transient overloads (503 UNAVAILABLE) Gemini throws under load:
    retries the primary model with backoff, then tries a fallback model. Raises
    GeminiRateLimited on a 429 (wait for quota), GeminiTruncated if the reply kept
    running out of output room, and GeminiUnavailable if every attempt is
    overloaded."""
    from google import genai  # lazy: only needed on this page
    from google.genai import types

    client = genai.Client(api_key=api_key)
    parts: list[object] = [
        types.Part.from_bytes(data=data, mime_type=mime) for data, mime in images
    ]
    parts.append(EXTRACTION_PROMPT)

    def build_config(drop_thinking: bool) -> object:
        # A 4-stage, 120-module form runs to ~10k output tokens, and on a thinking
        # model the deliberation is charged against the SAME ceiling (a real run
        # here: 9,167 thinking + 9,575 JSON). Ask for the model's full 65,536 so a
        # long report can't be cut off mid-JSON. Thinking is left ON by default —
        # it earns its keep on smudged handwriting — and only surrendered on a
        # retry that already truncated, where the output room matters more.
        kwargs: dict[str, object] = {
            "response_mime_type": "application/json",
            "temperature": 0,
            "max_output_tokens": 65536,
        }
        if drop_thinking:
            kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=0)
        return types.GenerateContentConfig(**kwargs)

    # (model, seconds-to-wait-before-this-attempt): try primary, retry primary
    # after a pause, then fall back to a second model.
    attempts = [(model, 0.0), (model, 2.0), (SCAN_FALLBACK_MODEL, 3.0)]
    last_error: Exception | None = None
    truncated = False
    for mdl, wait in attempts:
        if wait:
            time.sleep(wait)
        try:
            response = client.models.generate_content(
                model=mdl, contents=parts, config=build_config(truncated)
            )
        except Exception as exc:  # noqa: BLE001
            if _is_quota_error(exc):
                raise GeminiRateLimited(str(exc)) from exc
            if _is_transient_error(exc):
                last_error = exc
                continue  # transient — try again / fall back
            raise  # a real error (bad request, auth, etc.) — surface it

        # A reply that ran out of room stops mid-JSON, so json.loads fails with an
        # offset deep in the string. Name it here instead of letting the parser
        # error stand in for it, and retry with the thinking budget surrendered.
        if _hit_output_ceiling(response):
            truncated = True
            last_error = GeminiTruncated(
                f"{mdl} ran out of output room and stopped mid-JSON "
                "(the form is too dense for one reply)."
            )
            continue
        try:
            return _coerce_extracted(json.loads(response.text or "{}"))
        except json.JSONDecodeError as exc:
            # Malformed (not truncated) JSON is the single most retryable failure
            # there is, and it used to be the one failure that fell straight
            # through to `raise` without ever reaching the fallback model.
            last_error = exc
            continue

    if isinstance(last_error, GeminiTruncated):
        raise GeminiTruncated(str(last_error))
    raise GeminiUnavailable(str(last_error) if last_error else "Gemini unavailable")


def _is_quota_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return bool(re.search(r"\b429\b", text)) or "resource_exhausted" in text or "quota" in text


def _is_transient_error(exc: Exception) -> bool:
    """Server-side blips worth retrying. The status codes are matched on word
    boundaries: a bare `"500" in text` also fires on any error whose message
    happens to contain a number like 3500, which silently miscategorises
    unrelated failures as retryable."""
    text = str(exc).lower()
    if re.search(r"\b(500|502|503|504)\b", text):
        return True
    return any(s in text for s in ("unavailable", "overload", "internal", "deadline"))


def _hit_output_ceiling(response: object) -> bool:
    """True if the reply stopped because it hit max_output_tokens."""
    candidates = getattr(response, "candidates", None) or []
    for cand in candidates:
        if "MAX_TOKENS" in str(getattr(cand, "finish_reason", "")).upper():
            return True
    return False


# Groq backup: separate infrastructure, so it survives a Gemini overload. Qwen3.6
# is multimodal; its vision input is IMAGES ONLY, so PDFs are rasterized to page
# images first. A notch below Gemini on handwriting — fine for a backup.
#
# Groq rotates its catalog and RETIRES models without notice — the previous pick
# (meta-llama/llama-4-scout-17b-16e-instruct) was decommissioned, which surfaces
# as a "model does not exist" 404 at call time, NOT a bad key. If that recurs,
# GET https://api.groq.com/openai/v1/models with the key and pick a current
# vision model (one that accepts an image_url content part).
GROQ_MODEL = "qwen/qwen3.6-27b"


def _to_images(files: list[tuple[bytes, str]]) -> list[tuple[bytes, str]]:
    """Expand any PDFs into per-page PNG images; pass real images through. Used
    for providers (Groq) whose vision input does not accept PDF directly."""
    out: list[tuple[bytes, str]] = []
    for data, mime in files:
        if "pdf" in (mime or "").lower():
            import fitz  # PyMuPDF, lazy

            doc = fitz.open(stream=data, filetype="pdf")
            try:
                for page in doc:
                    pix = page.get_pixmap(dpi=150)
                    out.append((pix.tobytes("png"), "image/png"))
            finally:
                doc.close()
        else:
            out.append((data, mime or "image/jpeg"))
    return out


def extract_imr_via_groq(files: list[tuple[bytes, str]], api_key: str) -> dict:
    """Backup extractor: Groq Qwen3.6 vision -> parsed JSON dict. Qwen vision is
    images-only, so any PDF is rasterized to page images first."""
    import base64

    from groq import Groq  # lazy: only needed on fallback

    content: list[dict] = [{"type": "text", "text": EXTRACTION_PROMPT}]
    for data, mime in _to_images(files):
        b64 = base64.b64encode(data).decode("ascii")
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})

    client = Groq(api_key=api_key)
    response = client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[{"role": "user", "content": content}],
        response_format={"type": "json_object"},
        temperature=0,
        max_tokens=16384,  # headroom so a long report's JSON isn't cut off
    )
    return _coerce_extracted(json.loads(response.choices[0].message.content or "{}"))


class ExtractionFailed(Exception):
    """Every configured provider failed; message lists what happened."""


def _looks_like_overflow(exc: Exception) -> bool:
    """Heuristic: did this failure come from the model's JSON being cut off /
    rejected because the report was too big for one reply? Those are the cases the
    per-page fallback fixes (vs. auth/quota errors, which it wouldn't)."""
    if isinstance(exc, (json.JSONDecodeError, GeminiTruncated)):
        return True
    t = str(exc).lower()
    return any(
        s in t for s in (
            "unterminated", "expecting", "failed to validate json", "json_validate",
            "max_tokens", "maximum context", "too long", "truncat", "finish_reason",
        )
    )


def merge_extracted(parts: list[dict]) -> dict:
    """Combine per-page extractions into one IMR dict: cover fields take the first
    non-empty value; stages merge by label (their module lists concatenate, deduped
    by Mo No); parameters concatenate, deduped by tag."""
    merged: dict = {
        "plant_sr_no": None, "report_date": None, "zone": None,
        "site_name": None, "plant_capacity": None, "stages": [], "parameters": [],
    }
    stage_by_key: dict[str, dict] = {}
    param_by_tag: dict[str, dict] = {}

    def empty(v: object) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    for raw in parts:
        page = _coerce_extracted(raw)
        for field in ("plant_sr_no", "report_date", "zone", "site_name", "plant_capacity"):
            if empty(merged[field]) and not empty(page.get(field)):
                merged[field] = page.get(field)
        for stage in page.get("stages") or []:
            key = str(stage.get("stage_label") or "").strip().upper()
            tgt = stage_by_key.setdefault(
                key, {"stage_label": stage.get("stage_label"), "volume_ml": None, "modules": []}
            )
            if empty(tgt["volume_ml"]) and not empty(stage.get("volume_ml")):
                tgt["volume_ml"] = stage.get("volume_ml")
            seen = {m.get("mo_no") for m in tgt["modules"]}
            for module in stage.get("modules") or []:
                mo = module.get("mo_no")
                if mo is None or mo not in seen:
                    tgt["modules"].append(module)
                    seen.add(mo)
        for param in page.get("parameters") or []:
            tag = normalize_text(param.get("tag"))
            if tag and tag not in param_by_tag:
                param_by_tag[tag] = param

    merged["stages"] = list(stage_by_key.values())
    merged["parameters"] = list(param_by_tag.values())
    return merged


def _extract_paged(call, files: list[tuple[bytes, str]]) -> dict:
    """Run `call` on the whole document at once; if that fails because the reply
    overflowed (crowded/multi-page report), rasterize to pages and extract each
    page separately, then merge. `call` takes a list of (bytes, mime) items — the
    whole document for the one-shot, a single page image for the per-page pass —
    and returns a parsed dict."""
    try:
        return call(files)
    except Exception as exc:  # noqa: BLE001
        if not _looks_like_overflow(exc):
            raise
        pages = _to_images(files)
        if len(pages) <= 1:
            raise
        results, errors = [], []
        for page in pages:
            try:
                results.append(call([page]))
            except Exception as page_exc:  # noqa: BLE001 - skip a bad page, keep the rest
                errors.append(str(page_exc)[:80])
        if not results:
            raise ExtractionFailed("per-page extraction failed: " + " | ".join(errors[:3])) from exc
        return merge_extracted(results)


def extract_imr(
    files: list[tuple[bytes, str]], gemini_key: str | None, groq_key: str | None
) -> tuple[dict, str]:
    """Extract with Gemini (primary); fall back to Groq on any Gemini failure. Each
    provider first tries the whole document in one call, then — if the reply
    overflowed (crowded/multi-page report) — falls back to per-page extraction and
    merges. Returns (data, provider_label); raises ExtractionFailed if all fail."""
    notes: list[str] = []
    if gemini_key:
        try:
            data = _extract_paged(
                lambda imgs: extract_imr_from_images(imgs, SCAN_MODEL, gemini_key), files
            )
            return data, "Gemini (2.5-flash)"
        except GeminiRateLimited:
            notes.append("Gemini hit its free-tier limit (429)")
        except GeminiUnavailable:
            notes.append("Gemini was overloaded (503)")
        except GeminiTruncated:
            notes.append(
                "Gemini ran out of output room on this form (too many modules for "
                "one reply) — split the photo by stage and scan the halves"
            )
        except Exception as exc:  # noqa: BLE001
            notes.append(f"Gemini error: {str(exc)[:120]}")
    if groq_key:
        try:
            data = _extract_paged(lambda imgs: extract_imr_via_groq(imgs, groq_key), files)
            return data, "Groq (Qwen3.6)"
        except Exception as exc:  # noqa: BLE001
            notes.append(f"Groq error: {str(exc)[:120]}")
    if not gemini_key and not groq_key:
        notes.append("no API key configured")
    raise ExtractionFailed(" → ".join(notes))


def compute_scan_flow(time_sec: object, volume_ml: object) -> float | None:
    """Total flow (L/hr) from the timed-volume method the paper form uses:
    flow = volume_ml * 3.6 / time_sec. None if either input is missing or zero."""
    t = pd.to_numeric(time_sec, errors="coerce")
    v = pd.to_numeric(volume_ml, errors="coerce")
    if pd.notna(t) and t > 0 and pd.notna(v) and v > 0:
        return round(float(v) * 3.6 / float(t), 1)
    return None


def _scan_cell(value: object) -> object:
    """Blank cells from the editor arrive as NaN/empty string; keep them empty so
    the sheet never shows a literal 'nan'."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return value


def _scan_date(value: object) -> object:
    """Write dates as real date cells, never as text.

    read_cover_block parses a TEXT date with dayfirst=True, because hand-filled
    forms write dd/mm — which silently turns our own ISO "2026-07-09" into 7 Sep,
    a future month the reader then discards, losing the report date entirely. A
    real date cell takes its unambiguous datetime branch instead.

    ONLY an exact ISO date is converted. The date field is free text in the
    editor, so anything else may well be a hand-typed "09/07/2026" — that has to
    stay text and reach read_cover_block's dayfirst rule, or converting it here
    would read it month-first and land on the wrong month, which is the very bug
    this function exists to prevent."""
    cleaned = _scan_cell(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, (datetime, date)):
        return cleaned
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(cleaned).strip()):
        return cleaned
    parsed = pd.to_datetime(str(cleaned).strip(), errors="coerce")
    return parsed.date() if pd.notna(parsed) else cleaned


_SCAN_BORDERS: dict[tuple[str | None, ...], object] = {}


def _scan_border(left: str, right: str, top: str, bottom: str) -> object:
    """Cached Border — a 120-module sheet touches ~900 cells, and building a fresh
    style object per cell is both slow and bloats the saved workbook."""
    from openpyxl.styles import Border, Side

    key = (left, right, top, bottom)
    if key not in _SCAN_BORDERS:
        _SCAN_BORDERS[key] = Border(
            left=Side(style=left), right=Side(style=right),
            top=Side(style=top), bottom=Side(style=bottom),
        )
    return _SCAN_BORDERS[key]


def build_imr_workbook(extracted: dict) -> bytes:
    """Render an extracted/corrected IMR dict as a workbook the ingest parser reads.

    One stage block per stage found, each exactly as tall as its own module list,
    laid out left to right with SCAN_BLOCK_WIDTH columns apiece — so a IV STAGE or
    a 60-module stage comes through whole, and no plant is a special case. The
    sheet is then styled to match the printed form (see the SCAN_* constants for
    why generating it is safe and which strings the parser depends on)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMR"

    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    centre_flat = Alignment(horizontal="center", vertical="center")
    value_fill = PatternFill(fill_type="solid", fgColor=SCAN_VALUE_FILL)

    stages = [s for s in (extracted.get("stages") or []) if s]
    n_blocks = max(len(stages), 1)
    last_col = n_blocks * SCAN_BLOCK_WIDTH - 1  # drop the trailing spacer

    # ----- title -----
    ws["A1"] = "INDIVIDUAL MODULE REPORT (IMR)"
    ws["A1"].font = Font(name="Calibri", size=24, bold=True)
    ws["A1"].alignment = centre_flat
    ws["A1"].border = _scan_border("medium", "medium", "medium", "medium")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(last_col, 2))
    ws.row_dimensions[1].height = SCAN_ROW_H["title"]

    # ----- cover -----
    # Labels are merged over two columns because a single module column (8.7) is
    # too narrow for "PLANT CAPACITY :-" and the text would spill over its value.
    # read_cover_block finds a label anywhere in the first 14 rows x 12 columns and
    # takes the first NON-EMPTY cell to its right, so the merge — whose trailing
    # cells read as blank — and the skipped spacer column are both invisible to it.
    sr_row, site_row, cap_row = SCAN_COVER_ROWS
    left = Alignment(horizontal="left", vertical="center")

    def cover_field(row: int, label: str, value: object, second: bool) -> None:
        label_col, value_col = (5, 8) if second else (1, 3)
        head = ws.cell(row=row, column=label_col, value=label)
        head.font = Font(name="Calibri", size=10, bold=True)
        head.alignment = left
        ws.merge_cells(start_row=row, start_column=label_col,
                       end_row=row, end_column=label_col + 1)
        target = ws.cell(row=row, column=value_col, value=value)
        target.font = Font(name="Calibri", size=11)
        target.alignment = left
        target.fill = value_fill
        if isinstance(value, (datetime, date)):
            target.number_format = "DD-MM-YYYY"  # otherwise a date renders as ###
        ws.merge_cells(start_row=row, start_column=value_col,
                       end_row=row, end_column=value_col + 1)
        for col in range(label_col, label_col + 2):
            ws.cell(row=row, column=col).border = _scan_border(
                "medium" if col == label_col else "thin", "thin", "thin", "thin")
        for col in range(value_col, value_col + 2):
            cell = ws.cell(row=row, column=col)
            cell.border = _scan_border(
                "thin", "medium" if col == value_col + 1 else "thin", "thin", "thin")
            cell.fill = value_fill

    cover_field(sr_row, "PLANT SR NO :-", _scan_cell(extracted.get("plant_sr_no")), False)
    cover_field(sr_row, "REPORT DATE :-", _scan_date(extracted.get("report_date")), True)
    cover_field(site_row, "SITE NAME :-", _scan_cell(extracted.get("site_name")), False)
    cover_field(site_row, "ZONE :-", _scan_cell(extracted.get("zone")), True)
    cover_field(cap_row, "PLANT CAPACITY :-", _scan_cell(extracted.get("plant_capacity")), False)
    for row in SCAN_COVER_ROWS:
        ws.row_dimensions[row].height = SCAN_ROW_H["cover"]

    # ----- stage blocks, each sized to its own module list -----
    last_data_row = SCAN_DATA_ROW
    for index, stage in enumerate(stages):
        base = 1 + index * SCAN_BLOCK_WIDTH  # 1-based first column of this block
        mo_c, inst_c, time_c, flow_c, cond_c = base, base + 1, base + 2, base + 4, base + 5
        time_letter = get_column_letter(time_c)

        band = ws.cell(row=SCAN_STAGE_LABEL_ROW, column=mo_c,
                       value=stage.get("stage_label") or f"STAGE {index + 1}")
        band.font = Font(name="Calibri", size=11, bold=True)
        band.alignment = centre
        ws.merge_cells(start_row=SCAN_STAGE_LABEL_ROW, start_column=mo_c,
                       end_row=SCAN_STAGE_LABEL_ROW, end_column=cond_c)
        for offset in range(6):
            ws.cell(row=SCAN_STAGE_LABEL_ROW, column=base + offset).border = _scan_border(
                "medium" if offset == 0 else "thin",
                "medium" if offset == 5 else "thin", "medium", "thin",
            )

        for offset, header in enumerate(SCAN_BLOCK_HEADERS):
            cell = ws.cell(row=SCAN_HEADER_ROW, column=base + offset, value=header)
            cell.font = Font(name="Calibri", size=9, bold=True)
            cell.alignment = centre
            cell.border = _scan_border(
                "medium" if offset == 0 else "thin",
                "medium" if offset == 5 else "thin", "thin", "thin",
            )

        # Naming the timed volume in the header is what lets parse_volume_ml pick
        # it up downstream, so flow can be re-derived from the stopwatch reading.
        volume_ml = pd.to_numeric(stage.get("volume_ml"), errors="coerce")
        numerator = None
        if pd.notna(volume_ml) and volume_ml > 0:
            ws.cell(row=SCAN_HEADER_ROW, column=time_c, value=f"Time for {volume_ml:g} ml")
            numerator = float(volume_ml) * 3.6

        row = SCAN_DATA_ROW
        for module in stage.get("modules") or []:
            mo = _scan_cell(module.get("mo_no"))
            if mo is None:
                continue  # blank editor row
            ws.cell(row=row, column=mo_c, value=mo)
            ws.cell(row=row, column=inst_c, value=_scan_date(module.get("inst_date")))
            remark = str(_scan_cell(module.get("remark")) or "")
            if "by pass" in remark.lower() or "bypass" in remark.lower():
                ws.cell(row=row, column=cond_c, value="BY PASS")
            else:
                seconds = pd.to_numeric(module.get("time_sec"), errors="coerce")
                ws.cell(row=row, column=time_c,
                        value=float(seconds) if pd.notna(seconds) else None)
                # With both the timed volume and the reading, write the live
                # formula the paper form uses so editing the time updates the
                # flow; otherwise fall back to the transcribed flow number.
                if numerator is not None and pd.notna(seconds) and seconds != 0:
                    ws.cell(row=row, column=flow_c, value=f"={numerator:g}/{time_letter}{row}")
                else:
                    ws.cell(row=row, column=flow_c, value=_scan_cell(module.get("flow")))
                ws.cell(row=row, column=cond_c, value=_scan_cell(module.get("cond")))
            row += 1

        end_row = row - 1
        for r in range(SCAN_DATA_ROW, row):
            ws.row_dimensions[r].height = SCAN_ROW_H["data"]
            for offset in range(6):
                cell = ws.cell(row=r, column=base + offset)
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = centre_flat
                cell.border = _scan_border(
                    "medium" if offset == 0 else "thin",
                    "medium" if offset == 5 else "thin",
                    "thin", "medium" if r == end_row else "thin",
                )
            # The flow formula divides, so it prints to full float precision
            # (276.9231) unless told otherwise; dates need a format to render.
            ws.cell(row=r, column=flow_c).number_format = "0.00"
            ws.cell(row=r, column=inst_c).number_format = "DD-MM-YYYY"

        # Each block gets its own AVERAGE under its own data — blocks no longer
        # share a row, so a short stage isn't padded out to the tallest one.
        if end_row >= SCAN_DATA_ROW:
            avg_row = row
            label = ws.cell(row=avg_row, column=mo_c, value="AVERAGE")
            label.font = Font(name="Calibri", size=11, bold=True)
            label.alignment = centre_flat
            for col in (flow_c, cond_c):
                letter = get_column_letter(col)
                cell = ws.cell(
                    row=avg_row, column=col,
                    value=f'=IFERROR(ROUND(AVERAGE({letter}{SCAN_DATA_ROW}:{letter}{end_row}),2),"")',
                )
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = centre_flat
                cell.number_format = "0.00"
            for offset in range(6):
                ws.cell(row=avg_row, column=base + offset).border = _scan_border(
                    "medium" if offset == 0 else "thin",
                    "medium" if offset == 5 else "thin", "medium", "medium",
                )
            row += 1
        last_data_row = max(last_data_row, row)

        for offset, width in enumerate(SCAN_COL_WIDTHS):
            ws.column_dimensions[get_column_letter(base + offset)].width = width
        if index < len(stages) - 1:
            ws.column_dimensions[get_column_letter(base + 6)].width = SCAN_SPACER_WIDTH

    ws.row_dimensions[SCAN_STAGE_LABEL_ROW].height = SCAN_ROW_H["band"]
    ws.row_dimensions[SCAN_HEADER_ROW].height = SCAN_ROW_H["header"]

    # ----- operating parameters -----
    # extract_operating_parameters keys off the UNIT cell, reading the value one
    # column left and the tag two left, so the tag must sit at column C or later
    # for that lookback to stay in bounds.
    param_row = last_data_row + 1
    heading = ws.cell(row=param_row, column=3, value="OPERATING PARAMETERS")
    heading.font = Font(name="Calibri", size=12, bold=True)
    heading.alignment = centre
    ws.merge_cells(start_row=param_row, start_column=3, end_row=param_row, end_column=5)
    for col in (3, 4, 5):
        ws.cell(row=param_row, column=col).border = _scan_border(
            "medium" if col == 3 else "thin", "medium" if col == 5 else "thin",
            "medium", "thin",
        )
    param_row += 1

    written = [
        (p, scan_param_unit(p.get("tag")), _scan_cell(p.get("value")))
        for p in extracted.get("parameters") or []
    ]
    written = [(p, u, v) for p, u, v in written if u is not None and v is not None]
    for position, (param, unit, value) in enumerate(written):
        last = position == len(written) - 1
        cells = [
            (3, str(param.get("tag")), True),
            (4, value, False),
            (5, unit, False),
        ]
        for col, content, bold in cells:
            cell = ws.cell(row=param_row, column=col, value=content)
            cell.font = Font(name="Calibri", size=12, bold=bold)
            cell.alignment = centre_flat if col != 3 else Alignment(vertical="center")
            cell.border = _scan_border(
                "medium" if col == 3 else "thin", "medium" if col == 5 else "thin",
                "thin", "medium" if last else "thin",
            )
            if col == 4:
                cell.fill = value_fill
        param_row += 1

    # A 60-module stage scrolls past the headers, so pin them; landscape A3 at
    # fit-to-width keeps a 4-block sheet printable on one page.
    ws.freeze_panes = ws.cell(row=SCAN_DATA_ROW, column=1)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.zoomScale = 100

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_scan_page(engine: Engine) -> None:
    st.title("Scan IMR — photo to Excel")
    st.caption(
        "Photograph or scan (PDF) a handwritten IMR printout; it's read into the "
        "finalized Excel format. Review and fix anything, then download it or commit "
        "it straight to the database. Gemini is primary, Groq is the backup if Gemini "
        "is overloaded. No more retyping."
    )

    gemini_key = get_gemini_key()
    groq_key = get_groq_key()
    if not gemini_key and not groq_key:
        st.info(
            "Add a **free** API key to enable scanning (`.streamlit/secrets.toml`):\n\n"
            "```toml\n[gemini]\napi_key = \"YOUR_FREE_GEMINI_KEY\"   # primary\n\n"
            "[groq]\napi_key = \"YOUR_FREE_GROQ_KEY\"       # backup\n```\n\n"
            "Gemini: https://aistudio.google.com/apikey · Groq: https://console.groq.com/keys "
            "(both free)."
        )
        return

    # ----- capture -----
    up_tab, cam_tab = st.tabs(["Upload photo(s)", "Use camera"])
    images: list[tuple[bytes, str]] = []
    with up_tab:
        files = st.file_uploader(
            "Photo(s) or a scanned PDF of the same form",
            type=["jpg", "jpeg", "png", "webp", "pdf"],
            accept_multiple_files=True,
        )
        for f in files or []:
            mime = f.type or (
                "application/pdf" if str(f.name).lower().endswith(".pdf") else "image/jpeg"
            )
            images.append((f.getvalue(), mime))
    with cam_tab:
        shot = st.camera_input("Snap the form")
        if shot is not None:
            images.append((shot.getvalue(), shot.type or "image/jpeg"))

    if st.button("Extract data", type="primary", disabled=not images):
        with st.spinner(f"Reading {len(images)} file(s)…"):
            try:
                extracted, provider = extract_imr(images, gemini_key, groq_key)
            except ExtractionFailed as exc:
                st.error(
                    f"Couldn't extract — every provider failed: {exc}. "
                    "If both are overloaded, wait a minute and retry."
                )
                return
        st.session_state["scan_extracted"] = extracted
        st.session_state["scan_provider"] = provider

    extracted = st.session_state.get("scan_extracted")
    if not extracted:
        return
    provider = st.session_state.get("scan_provider", "")
    if provider:
        st.success(f"Extracted via {provider}. Review and fix below.")

    # ----- review & correct -----
    st.markdown("---")
    st.subheader("Review & fix")
    st.caption("Correct anything the OCR misread before you download or commit.")

    c = st.columns(5)
    with c[0]:
        sr = st.number_input("Plant SR No", value=int(extracted.get("plant_sr_no") or 0), step=1, min_value=0)
    with c[1]:
        date_str = st.text_input("Report date (YYYY-MM-DD)", value=str(extracted.get("report_date") or ""))
    with c[2]:
        zone_val = extracted.get("zone") if extracted.get("zone") in SCAN_ZONES else None
        zone = st.selectbox("Zone", SCAN_ZONES, index=SCAN_ZONES.index(zone_val) if zone_val else 0)
    with c[3]:
        site = st.text_input("Site name", value=str(extracted.get("site_name") or ""))
    with c[4]:
        cap = st.text_input("Plant capacity", value=str(extracted.get("plant_capacity") or ""))

    corrected_stages = []
    for idx, stage in enumerate(extracted.get("stages") or []):
        st.markdown(f"**{stage.get('stage_label', f'Stage {idx+1}')}**")
        vol_default = pd.to_numeric(stage.get("volume_ml"), errors="coerce")
        volume_ml = st.number_input(
            "Time for ___ ml (volume timed per reading)",
            value=float(vol_default) if pd.notna(vol_default) else 0.0,
            step=100.0, min_value=0.0, key=f"scan_vol_{idx}",
            help="Flow is computed as volume × 3.6 ÷ time. Leave 0 if the form has "
                 "no timed volume — flow is then taken from the Flow column as-is.",
        )
        mod_df = pd.DataFrame(stage.get("modules") or [],
                              columns=["mo_no", "inst_date", "time_sec", "flow", "cond", "remark"])
        # Flow is DERIVED (volume × 3.6 ÷ time), so compute and show it rather than
        # relying on the usually-blank transcribed Flow column. Keep any handwritten
        # flow only where there's no time to compute from.
        if volume_ml and volume_ml > 0 and not mod_df.empty:
            computed = mod_df["time_sec"].map(lambda t: compute_scan_flow(t, volume_ml))
            mod_df["flow"] = computed.where(computed.notna(), mod_df["flow"])
        edited = st.data_editor(
            mod_df, num_rows="dynamic", width="stretch", key=f"scan_stage_{idx}",
            column_config={
                "mo_no": "Mo no.", "inst_date": "Inst date", "time_sec": "Time (sec)",
                "flow": st.column_config.NumberColumn(
                    "Flow (L/hr)", help="Auto-computed as volume × 3.6 ÷ time. Edit the time to change it."
                ),
                "cond": "Cond (uS/cm)", "remark": "Remark",
            },
        )
        # A mis-read time (dropped decimal, e.g. 11.47 → 1147) shows up as an
        # implausibly large time and an absurd flow — flag it for the reviewer.
        secs = pd.to_numeric(edited["time_sec"], errors="coerce")
        suspect = edited.loc[secs >= 200, "mo_no"].dropna().tolist()
        if suspect:
            st.warning(
                f"Time looks too large for module(s) {', '.join(map(str, suspect))} "
                "(≥ 200 s). Handwritten times are usually 3–90 s — check for a missed "
                "decimal point (e.g. 1147 should be 11.47)."
            )
        corrected_stages.append({
            "stage_label": stage.get("stage_label"),
            "volume_ml": volume_ml or None,
            "modules": edited.to_dict("records"),
        })

    params = extracted.get("parameters") or []
    if params:
        st.markdown("**Operating parameters**")
        par_df = pd.DataFrame(params, columns=["tag", "value"])
        par_edit = st.data_editor(par_df, num_rows="dynamic", width="stretch", key="scan_params")
        corrected_params = par_edit.to_dict("records")
    else:
        corrected_params = []

    corrected = {
        "plant_sr_no": int(sr) or None,
        "report_date": date_str.strip() or None,
        "zone": zone,
        "site_name": site.strip() or None,
        "plant_capacity": cap.strip() or None,
        "stages": corrected_stages,
        "parameters": corrected_params,
    }

    xlsx_bytes = build_imr_workbook(corrected)
    base = f"{corrected['plant_sr_no'] or 'imr'}_{corrected['report_date'] or 'report'}"
    filename = safe_upload_name(f"{base}.xlsx")

    # ----- download / commit -----
    st.markdown("---")
    written = sum(
        1 for s in corrected["stages"] for m in (s.get("modules") or [])
        if _scan_cell(m.get("mo_no")) is not None
    )
    st.caption(
        f"Excel: {len(corrected['stages'])} stage(s), {written} module(s) — "
        "the sheet is built to fit, so nothing is dropped."
    )
    dl, cm = st.columns(2)
    with dl:
        st.download_button(
            "⬇ Download Excel", xlsx_bytes, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with cm:
        commit = st.button("✓ Commit to database", type="primary", width="stretch")

    if commit:
        content_hash = hashlib.sha256(xlsx_bytes).hexdigest()
        result = parse_report_bytes(filename, xlsx_bytes)
        issues = validate_parse(engine, filename, result)
        blockers = [i for i in issues if i.level == "ERROR"]
        for issue in issues:
            ISSUE_RENDERERS.get(issue.level, st.info)(issue.message)
        if blockers:
            st.error("Fix the blocking issues above (or correct the data) before committing.")
        elif result.readings.empty:
            st.error("No readings were parsed from the filled form — check the module rows.")
        else:
            store_report_bytes(engine, filename, content_hash, xlsx_bytes, status="committed")
            counts = commit_parse_result(engine, filename, content_hash, result)
            load_readings.clear()
            load_parameters.clear()
            load_mis.clear()
            st.success(
                f"Committed {filename}: {counts['readings']} readings, "
                f"{counts['parameters']} parameters, {counts['mis']} MIS rows."
            )
            st.session_state.pop("scan_extracted", None)


def build_submission_roster(plants: pd.DataFrame) -> pd.DataFrame:
    """The roster of plants expected to report: every register plant with a normal
    (blank) status. A plant flagged with any exception status (Plant Shutdown,
    Stand By RO, Not in Rochem Scope, STRO/SPRO/UF RO) is not expected to submit an
    IMR, so it's excluded here — but kept in the register for history."""
    cols = ["plant_sr_no", "zone", "site_name"]
    if plants is None or plants.empty:
        return pd.DataFrame(columns=cols)
    active = plants.dropna(subset=["plant_sr_no"]).copy()
    if "status" in active:
        active = active[active["status"].map(status_is_reporting)]
    if active.empty:
        return pd.DataFrame(columns=cols)
    active["plant_sr_no"] = active["plant_sr_no"].astype(int)
    active["zone"] = active["zone"].map(canonicalize_zone).fillna("Unknown")
    active["site_name"] = active["site_name"].fillna("").astype(str).str.strip()
    return active.drop_duplicates("plant_sr_no")[cols].reset_index(drop=True)


def submission_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Boolean grid (index plant_sr_no x columns 'YYYY-MM'): True if that plant
    had at least one reading that month."""
    sub = df.dropna(subset=["plant_sr_no"]).copy()
    if sub.empty:
        return pd.DataFrame()
    sub["plant_sr_no"] = sub["plant_sr_no"].astype(int)
    sub["month"] = sub["report_date"].dt.to_period("M").astype(str)
    return sub.groupby(["plant_sr_no", "month"]).size().unstack(fill_value=0) > 0


def plant_search_mask(plants: pd.DataFrame, query: str) -> pd.Series:
    """Rows of `plants` matching a Plant Register lookup. A query that is exactly a
    plant's SR No wins outright (so searching 104 doesn't bury it under 1040-1049);
    otherwise it matches the digits anywhere in an SR No, or the text anywhere in a
    site name — both case-insensitive."""
    q = str(query).strip().lower()
    if not q or plants.empty:
        return pd.Series(True, index=plants.index)

    sr = plants["plant_sr_no"].astype("string").fillna("")
    digits = re.sub(r"\D", "", q)
    if digits:
        exact = sr == digits
        if bool(exact.any()):
            return exact

    name = plants.get("site_name", pd.Series(dtype=object)).reindex(plants.index)
    mask = name.astype("string").fillna("").str.lower().str.contains(q, regex=False)
    if digits:
        mask = mask | sr.str.contains(digits, regex=False)
    return mask.fillna(False)


def render_plant_register(engine: Engine) -> None:
    st.title("Plant Register")
    st.caption(
        "The master list of plants — the single source of truth for each plant's "
        "zone and whether it's still operating. Fix a zone, add a new client's "
        "plant, or mark a shut-down plant Inactive. The IMR Tracker and every zone "
        "view read from this; an Inactive plant drops out of the tracker's expected "
        "list but keeps its history. Use the form below to onboard a plant, the search "
        "to look one up by SR No, or edit any row in the table; delete a row by "
        "selecting it and pressing the trash icon."
    )
    plants = load_plants()

    zones = known_zones(plants)

    # ----- Add a plant (new business) -----
    with st.expander("🏭 Add a plant (onboarding a new client / site)"):
        with st.form("add_plant_form", clear_on_submit=True):
            a1, a2 = st.columns([1, 3])
            new_sr = a1.number_input(
                "SR No", min_value=1, step=1, value=None, format="%d",
                help="The plant's serial number — the key every IMR workbook and "
                     "every zone view joins on. Must be unique.",
            )
            new_site = a2.text_input("Site name", placeholder="e.g. Acme Chemicals, Dahej")
            b1, b2, b3 = st.columns(3)
            new_zone_pick = b1.selectbox(
                "Zone", options=zones, index=None, placeholder="Pick a zone",
            )
            new_capacity = b2.text_input("Capacity", placeholder="e.g. 100 KLD")
            new_type = b3.selectbox(
                "Type", options=PLANT_TYPE_OPTIONS,
                index=PLANT_TYPE_OPTIONS.index(DEFAULT_PLANT_TYPE),
                help="The kind of system at this plant. Leave as the default (PT) if "
                     "you're not sure — it can be changed in the table any time.",
            )
            new_status = st.selectbox("Status", options=PLANT_STATUS_OPTIONS, index=0)
            submitted = st.form_submit_button("Add plant", type="primary")
        if submitted:
            if new_sr is None:
                st.error("SR No is required — it's how readings join to this plant.")
            elif not (new_site or "").strip():
                st.error("Site name is required.")
            else:
                try:
                    add_plant(engine, int(new_sr), new_site, new_zone_pick,
                              new_capacity, new_status, new_type)
                except ValueError as exc:
                    st.error(str(exc))
                else:
                    load_plants.clear()
                    load_mis.clear()
                    compute_fleet_status.clear()
                    st.success(f"Added {new_site.strip()} (SR {int(new_sr)}).")
                    rerun_app()

    # ----- Status filter/slicer (top of page) -----
    status_label = (
        plants["status"].map(lambda v: canonical_status(v) or "Active")
        if not plants.empty else pd.Series([], dtype=str)
    )
    counts = status_label.value_counts()
    present = [s for s in PLANT_STATUS_OPTIONS if counts.get(s, 0) > 0]
    fc1, fc2 = st.columns([4, 1])
    with fc1:
        picked_status = st.multiselect(
            "Show statuses",
            options=PLANT_STATUS_OPTIONS,
            default=present or PLANT_STATUS_OPTIONS,
            format_func=lambda s: f"{s} ({int(counts.get(s, 0))})",
            help="Filter the table to plants with these statuses — e.g. pick only "
                 "Inactive to review shut-down plants. Editing/saving affects just "
                 "the rows shown.",
        )
    with fc2:
        st.metric("In view", f"{int(status_label.isin(picked_status).sum())}/{len(plants)}")

    # ----- Look up a plant by SR No (or name) -----
    query = st.text_input(
        "🔎 Find a plant",
        placeholder="SR No (e.g. 1043) or part of a site name",
        help="Filters the table to the matching plants. An SR No matches exactly first; "
             "if nothing has that SR, the digits are matched anywhere in an SR No. "
             "Editing/saving affects just the rows shown.",
    ).strip()

    status_filtered = bool(picked_status) and set(picked_status) != set(PLANT_STATUS_OPTIONS)
    view = plants[status_label.isin(picked_status)] if status_filtered else plants
    if query and not view.empty:
        view = view[plant_search_mask(view, query)]
    # A save may only delete rows that were on screen, so a search counts as a filter.
    is_full_view = not status_filtered and not query
    # SRs currently on screen — bounds what a save is allowed to delete.
    scope_srs = set(pd.to_numeric(view["plant_sr_no"], errors="coerce").dropna().astype(int)) \
        if not view.empty else set()

    if query:
        if view.empty:
            st.warning(
                f"No plant matches “{query}”. If this is a new site, add it above — "
                "and check the status filter isn't hiding it."
            )
        elif len(view) == 1:
            hit = view.iloc[0]
            st.success(
                f"SR {int(hit['plant_sr_no'])} · {hit.get('site_name') or 'unnamed'} · "
                f"{hit.get('zone') or 'no zone'} · "
                f"{canonical_plant_type(hit.get('plant_type')) or DEFAULT_PLANT_TYPE} · "
                f"{canonical_status(hit.get('status')) or 'Active'}"
                + (f" · {hit['installed_capacity']}" if hit.get("installed_capacity") else "")
            )
        else:
            st.info(f"{len(view)} plants match “{query}”.")

    # Always A-Z by site name (case-insensitive, unnamed plants last) — the editor's
    # headers aren't clickable, so the order is fixed here rather than left to the user.
    if not view.empty and "site_name" in view.columns:
        key = (
            view["site_name"].astype("string").str.strip().str.lower().replace("", pd.NA)
        )
        view = view.loc[key.sort_values(na_position="last").index]

    GRID_COLUMNS = [
        "plant_sr_no", "site_name", "zone", "installed_capacity", "plant_type", "status",
    ]
    grid = view.reindex(columns=GRID_COLUMNS)
    if grid.empty:
        grid = pd.DataFrame(columns=GRID_COLUMNS)
    grid = grid.reset_index(drop=True)
    grid["status"] = grid["status"].map(canonical_status)
    # Everything is PT until someone marks the exceptions.
    grid["plant_type"] = grid["plant_type"].map(
        lambda v: canonical_plant_type(v) or DEFAULT_PLANT_TYPE
    )

    # The editor keys its pending edits by row POSITION, so re-filtering would reapply
    # them to the wrong plants. Fold the filter and the search into the key: changing
    # either starts the editor fresh on the new row set (and discards unsaved edits).
    editor_key = f"plant_editor::{','.join(sorted(picked_status))}::{query.lower()}"
    edited = st.data_editor(
        grid,
        num_rows="dynamic",
        width="stretch",
        height=560,
        key=editor_key,
        column_config={
            "plant_sr_no": st.column_config.NumberColumn("SR No", format="%d", width="small"),
            "site_name": st.column_config.TextColumn("Site name", width="large"),
            "zone": st.column_config.SelectboxColumn("Zone", options=zones, width="medium"),
            "installed_capacity": st.column_config.TextColumn("Capacity", width="small"),
            "plant_type": st.column_config.SelectboxColumn(
                "Type", options=PLANT_TYPE_OPTIONS, width="small", required=True,
                help="The kind of system installed at this plant. Everything starts as "
                     "PT — change the ones that aren't.",
            ),
            "status": st.column_config.SelectboxColumn(
                "Status", options=PLANT_STATUS_OPTIONS, width="medium",
                help="Leave blank for a normal RO plant that reports monthly. Pick a "
                     "flag to mark an exception — any flag excludes the plant from the "
                     "IMR Tracker's expected list (but keeps it in the register).",
            ),
        },
    )

    left, _ = st.columns([1, 3])
    if left.button("💾 Save changes", type="primary"):
        srs = pd.to_numeric(edited.get("plant_sr_no"), errors="coerce")
        valid = srs.dropna()
        if valid.duplicated().any():
            dups = sorted(valid[valid.duplicated(keep=False)].astype(int).unique())
            st.error(
                f"Duplicate SR No(s): {', '.join(map(str, dups))}. "
                "Each plant needs a unique SR No — fix these before saving."
            )
        else:
            # Scope the save to the filtered view so hidden plants aren't deleted;
            # a full-table view (all statuses) does a normal full replace.
            scope = None if is_full_view else scope_srs
            n = save_plants(engine, edited, scope_srs=scope)
            load_plants.clear()
            load_mis.clear()
            compute_fleet_status.clear()
            skipped = int(srs.isna().sum())
            msg = f"Saved {n} plants."
            if skipped:
                msg += f" ({skipped} row(s) without an SR No were skipped.)"
            st.success(msg)
            rerun_app()

    if not plants.empty:
        reporting = (
            int(plants["status"].map(status_is_reporting).sum())
            if "status" in plants else len(plants)
        )
        tally = " · ".join(
            f"{z}: {n}" for z, n in
            plants["zone"].map(canonicalize_zone).fillna("Unknown")
            .value_counts().sort_index().items()
        )
        st.caption(
            f"{len(plants)} plants ({reporting} reporting, "
            f"{len(plants) - reporting} flagged) · {tally}"
        )
        types = (
            plants.get("plant_type", pd.Series(dtype=object))
            .reindex(plants.index)
            .map(lambda v: canonical_plant_type(v) or DEFAULT_PLANT_TYPE)
            .value_counts()
        )
        st.caption("Types · " + " · ".join(
            f"{t}: {int(types[t])}" for t in PLANT_TYPE_OPTIONS if types.get(t, 0) > 0
        ))


def render_input_tracker(df: pd.DataFrame, plants: pd.DataFrame) -> None:
    st.title("IMR Input Tracker")
    st.caption(
        "Who has sent their IMR each month, by zone — the plant register checked "
        "against the readings actually received. Use it to chase the sites that "
        "haven't reported yet. Edit the roster on the Plant Register page."
    )

    roster = build_submission_roster(plants)
    grid = submission_matrix(df)
    if roster.empty or grid.empty:
        st.info(
            "Need both a plant register and some received readings to track "
            "submissions. Add plants on the Plant Register page and ingest at least "
            "one report with plant SR numbers."
        )
        return

    months = sorted(grid.columns.tolist(), reverse=True)
    month_labels = [pd.Period(m).strftime("%b %Y") for m in months]
    picked_label = st.selectbox("Month", month_labels, index=0)
    picked = months[month_labels.index(picked_label)]

    received_sr = set(grid.index[grid[picked]]) if picked in grid.columns else set()
    roster = roster.copy()
    roster["received"] = roster["plant_sr_no"].isin(received_sr)

    # Display name: prefer the readings sheet name, else the MIS site name.
    name_by_sr = (
        df.dropna(subset=["plant_sr_no"])
        .assign(plant_sr_no=lambda d: d["plant_sr_no"].astype(int))
        .sort_values("report_date")
        .drop_duplicates("plant_sr_no", keep="last")
        .set_index("plant_sr_no")["plant"]
        .to_dict()
    )
    roster["site"] = roster.apply(
        lambda r: str(name_by_sr.get(r["plant_sr_no"]) or r["site_name"] or f"SR {r['plant_sr_no']}"),
        axis=1,
    )

    # Last month each plant submitted (for chasing the chronic non-reporters).
    def last_submitted(sr: int) -> str:
        if sr in grid.index:
            hits = [m for m in months if grid.loc[sr, m]]
            if hits:
                return pd.Period(max(hits)).strftime("%b %Y")
        return "Never"
    roster["last_submitted"] = roster["plant_sr_no"].map(last_submitted)

    expected = len(roster)
    got = int(roster["received"].sum())
    missing = expected - got
    pct = got / expected * 100 if expected else 0.0
    zone_grp = roster.groupby("zone")
    zones_total = roster["zone"].nunique()
    zones_done = int((zone_grp["received"].mean() == 1).sum())

    # ----- 1. Headline + KPIs -----
    hero_card(
        "Missing IMRs",
        f"{missing:,}",
        f"{picked_label} — {got}/{expected} sites received ({pct:.0f}%)",
        color="#16a34a" if missing == 0 else "#dc2626",
    )
    st.markdown("")
    k = st.columns(4)
    with k[0]:
        metric_card("Sites expected", f"{expected:,}", "in the register")
    with k[1]:
        metric_card("Received", f"{got:,}", f"{pct:.0f}% this month", "#16a34a")
    with k[2]:
        metric_card("Missing", f"{missing:,}", "not yet received", "#dc2626")
    with k[3]:
        metric_card("Zones complete", f"{zones_done}/{zones_total}", "all sites in")

    # ----- 2. By zone -----
    st.markdown("---")
    st.subheader(f"By zone — {picked_label}")
    zs = (
        zone_grp.agg(expected=("plant_sr_no", "count"), received=("received", "sum"))
        .reset_index()
    )
    zs["missing"] = zs["expected"] - zs["received"]
    zs["completion"] = (zs["received"] / zs["expected"] * 100).round(0)
    zs = zs.sort_values(["missing", "zone"], ascending=[False, True])
    fig = go.Figure(
        go.Bar(
            x=zs["completion"],
            y=zs["zone"],
            orientation="h",
            marker_color=[
                "#16a34a" if c == 100 else "#f59e0b" if c >= 50 else "#dc2626"
                for c in zs["completion"]
            ],
            text=[f"{r}/{e}" for r, e in zip(zs["received"], zs["expected"])],
            textposition="auto",
            hovertemplate="%{y}: %{x:.0f}% received (%{text})<extra></extra>",
        )
    )
    fig.update_layout(
        template="plotly_white",
        height=max(220, 42 * len(zs)),
        xaxis_title="Completion %",
        xaxis_range=[0, 100],
        margin={"l": 10, "r": 10, "t": 10, "b": 10},
    )
    st.plotly_chart(fig, width="stretch")
    ztab = pd.DataFrame(
        {
            "Zone": zs["zone"],
            "Received": zs["received"],
            "Expected": zs["expected"],
            "Missing": zs["missing"],
            "Completion": zs["completion"].map(lambda v: f"{v:.0f}%"),
        }
    )
    st.dataframe(ztab, width="stretch", hide_index=True)

    # ----- 3. Missing this month (the actionable list) -----
    st.markdown("---")
    st.subheader(f"Missing this month — {picked_label}")
    miss = roster[~roster["received"]].sort_values(["zone", "site"])
    if miss.empty:
        st.success("Every registered site has submitted this month. 🎉")
    else:
        for zone, grp in miss.groupby("zone"):
            with st.expander(f"{zone} — {len(grp)} missing", expanded=True):
                st.dataframe(
                    pd.DataFrame(
                        {
                            "Site": grp["site"],
                            "SR No": grp["plant_sr_no"],
                            "Last submitted": grp["last_submitted"],
                        }
                    ),
                    width="stretch",
                    hide_index=True,
                )
        miss_out = pd.DataFrame(
            {
                "Zone": miss["zone"],
                "Site": miss["site"],
                "SR No": miss["plant_sr_no"],
                "Last submitted": miss["last_submitted"],
            }
        )
        st.download_button(
            "Download missing list (CSV)",
            miss_out.to_csv(index=False).encode("utf-8"),
            file_name=f"missing_imr_{picked}.csv",
            mime="text/csv",
        )

    # ----- 4. Submission history matrix -----
    st.markdown("---")
    st.subheader("Submission history")
    st.caption("✓ received · ✗ missing, by zone. Most recent 8 months.")
    zones = sorted(roster["zone"].unique())
    sel_zones = st.multiselect("Zones", zones, default=zones, key="tracker_zones")
    recent = months[:8][::-1]  # oldest -> newest for left-to-right reading
    month_cols = [pd.Period(m).strftime("%b %y") for m in recent]
    view = roster[roster["zone"].isin(sel_zones or zones)].sort_values(["zone", "site"])
    rows = []
    for _, r in view.iterrows():
        sr = r["plant_sr_no"]
        row = {"Zone": r["zone"], "Site": r["site"]}
        for m, label in zip(recent, month_cols):
            received_here = bool(grid.loc[sr, m]) if (sr in grid.index and m in grid.columns) else False
            row[label] = "✓" if received_here else "✗"
        rows.append(row)
    mat = pd.DataFrame(rows)
    if mat.empty:
        st.info("No sites in the selected zones.")
    else:
        def color_cell(v: str) -> str:
            if v == "✓":
                return "color:#16a34a;font-weight:700"
            if v == "✗":
                return "color:#dc2626;font-weight:700"
            return ""
        styler = mat.style
        styler = (styler.map if hasattr(styler, "map") else styler.applymap)(
            color_cell, subset=month_cols
        )
        st.dataframe(
            styler,
            width="stretch",
            hide_index=True,
            height=min(640, 60 + 28 * len(mat)),
        )


def _col_letter(col_idx: int) -> str:
    """Convert a 1-based column index to an Excel column letter (1->A, 27->AA)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def is_dark_hex(rgb: str) -> bool:
    """True if a 6-digit RRGGBB fill is dark enough that text on it must be light.
    Uses perceived luminance (ITU-R BT.601), so a mid yellow counts as light and a
    mid blue as dark, the way the eye reads them."""
    try:
        r, g, b = (int(rgb[i:i + 2], 16) for i in (0, 2, 4))
    except (ValueError, IndexError):
        return False
    return (0.299 * r + 0.587 * g + 0.114 * b) < 140


def _render_sheet_as_excel_html(
    ws, *, max_rows: int = 200, max_cols: int | None = None
) -> str:
    """Render an openpyxl worksheet as an HTML table that looks like MS Excel.

    Handles merged cells, number formatting, and preserves the raw layout. The
    output table has column-letter headers and row-number gutters, exactly like
    a real spreadsheet.
    """
    from openpyxl.utils import get_column_letter  # noqa: F811

    # Determine the extent of the sheet (clip to max_rows/max_cols for speed).
    n_rows = min(ws.max_row or 1, max_rows)
    n_cols = min(ws.max_column or 1, max_cols) if max_cols else (ws.max_column or 1)
    if n_cols > 50:
        n_cols = 50  # safety cap

    # Build a set of merged-cell ranges for quick lookup.
    # merged_top_left: (row, col) -> (row_span, col_span)
    # merged_hidden: set of (row, col) that are swallowed by a merge
    merged_top_left: dict[tuple[int, int], tuple[int, int]] = {}
    merged_hidden: set[tuple[int, int]] = set()
    for merge_range in ws.merged_cells.ranges:
        min_r, min_c = merge_range.min_row, merge_range.min_col
        max_r, max_c = merge_range.max_row, merge_range.max_col
        row_span = max_r - min_r + 1
        col_span = max_c - min_c + 1
        merged_top_left[(min_r, min_c)] = (row_span, col_span)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    merged_hidden.add((r, c))

    # Build HTML rows.
    rows_html: list[str] = []

    # Column header row (A, B, C, …)
    header = ['<th class="xl-corner"></th>']  # empty top-left corner
    for c in range(1, n_cols + 1):
        header.append(f'<th class="xl-col-hdr">{get_column_letter(c)}</th>')
    rows_html.append("<tr>" + "".join(header) + "</tr>")

    for r in range(1, n_rows + 1):
        cells = [f'<td class="xl-row-hdr">{r}</td>']
        for c in range(1, n_cols + 1):
            if (r, c) in merged_hidden:
                continue  # swallowed by a merge — no <td>

            cell = ws.cell(row=r, column=c)
            value = cell.value
            if value is None:
                display = ""
            elif isinstance(value, float):
                display = f"{value:g}"
            elif isinstance(value, (int, np.integer)):
                display = str(int(value))
            else:
                display = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

            # Alignment: numbers right, text left.
            align = "right" if isinstance(value, (int, float, np.integer, np.floating)) else "left"

            # Bold detection.
            bold = ""
            try:
                if cell.font and cell.font.bold:
                    bold = "font-weight:700;"
            except Exception:  # noqa: BLE001
                pass

            # Build cell background from fill. The workbook's own font colours are
            # not carried over, so a dark fill (navy/dark-green banner rows are
            # common in these sheets) would render dark-on-dark — pair it with light
            # text instead.
            bg = ""
            try:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb not in ("00000000", "0",):
                    rgb = str(cell.fill.fgColor.rgb)
                    if len(rgb) == 8:
                        rgb = rgb[2:]  # strip alpha
                    if len(rgb) == 6 and rgb != "000000":
                        bg = f"background:#{rgb};"
                        bg += "color:#ffffff;" if is_dark_hex(rgb) else "color:#111827;"
            except Exception:  # noqa: BLE001
                pass

            style = f'style="text-align:{align};{bold}{bg}"'

            span = merged_top_left.get((r, c))
            if span:
                rs, cs = span
                span_attr = ""
                if rs > 1:
                    span_attr += f' rowspan="{rs}"'
                if cs > 1:
                    span_attr += f' colspan="{cs}"'
                cells.append(f"<td {style}{span_attr}>{display}</td>")
            else:
                cells.append(f"<td {style}>{display}</td>")

        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    clipped_note = ""
    if (ws.max_row or 0) > max_rows:
        clipped_note = (
            f'<tr><td colspan="{n_cols + 1}" style="text-align:center;color:#475569;'
            f'font-style:italic;padding:8px;">Showing first {max_rows} of '
            f'{ws.max_row} rows</td></tr>'
        )

    return (
        '<div class="xl-sheet-wrap"><table class="xl-sheet">'
        + "\n".join(rows_html)
        + clipped_note
        + "</table></div>"
    )


# CSS that makes the HTML table look like an Excel spreadsheet.
EXCEL_PREVIEW_CSS = """
<style>
.xl-sheet-wrap {
    overflow: auto;
    max-height: 70vh;
    border: 1px solid #8eaaaa;
    border-radius: 2px;
    margin-bottom: 1rem;
}
table.xl-sheet {
    border-collapse: collapse;
    font-family: Calibri, "Segoe UI", Arial, sans-serif;
    /* The sheet is painted on a white ground, so its text colour must be pinned
       here — inheriting Streamlit's body colour renders grey (or near-white under a
       dark theme) text on white cells. */
    color: #111827;
    font-size: 12px;
    white-space: nowrap;
    width: max-content;
    min-width: 100%;
}
table.xl-sheet th, table.xl-sheet td {
    border: 1px solid #c6d0d0;
    padding: 2px 6px;
    height: 22px;
    min-width: 64px;
    max-width: 320px;
    overflow: hidden;
    text-overflow: ellipsis;
    vertical-align: middle;
}
/* Column header row (A, B, C …) */
table.xl-sheet th.xl-col-hdr {
    background: linear-gradient(180deg, #f0f5f0 0%, #dce4dc 100%);
    color: #333;
    font-weight: 600;
    text-align: center;
    position: sticky;
    top: 0;
    z-index: 3;
    border-bottom: 2px solid #8eaaaa;
    font-size: 11px;
    letter-spacing: 0.02em;
}
/* Top-left corner cell */
table.xl-sheet th.xl-corner {
    background: linear-gradient(135deg, #e0e8e0 0%, #c8d4c8 100%);
    position: sticky;
    top: 0;
    left: 0;
    z-index: 4;
    min-width: 40px;
    width: 40px;
    border-bottom: 2px solid #8eaaaa;
    border-right: 2px solid #8eaaaa;
}
/* Row-number gutter */
table.xl-sheet td.xl-row-hdr {
    background: linear-gradient(90deg, #f0f5f0 0%, #dce4dc 100%);
    color: #555;
    font-weight: 600;
    text-align: center;
    position: sticky;
    left: 0;
    z-index: 2;
    min-width: 40px;
    width: 40px;
    border-right: 2px solid #8eaaaa;
    font-size: 11px;
}
/* Data cells — white default */
table.xl-sheet td {
    background: #ffffff;
    color: #111827;
}
/* Sheet tab bar */
.xl-tab-bar {
    display: flex;
    gap: 0;
    border-bottom: 2px solid #8eaaaa;
    margin-bottom: 0;
    padding-left: 40px;
    background: #e8ece8;
    overflow-x: auto;
}
.xl-tab {
    padding: 5px 16px;
    font-size: 11.5px;
    font-family: Calibri, "Segoe UI", Arial, sans-serif;
    color: #333;
    border: 1px solid #a8b8a8;
    border-bottom: none;
    border-radius: 4px 4px 0 0;
    background: linear-gradient(180deg, #f6faf6 0%, #e0e8e0 100%);
    cursor: pointer;
    margin-right: -1px;
    white-space: nowrap;
}
.xl-tab.active {
    background: #ffffff;
    font-weight: 700;
    border-bottom: 2px solid #ffffff;
    position: relative;
    top: 1px;
    color: #1a5c1a;
}
/* Formula bar */
.xl-formula-bar {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 4px 8px;
    background: #f8faf8;
    border: 1px solid #c6d0d0;
    border-top: none;
    margin-bottom: 0;
    font-family: Calibri, "Segoe UI", Arial, sans-serif;
    font-size: 12px;
    color: #555;
}
.xl-formula-bar .xl-cell-ref {
    background: #fff;
    border: 1px solid #c6d0d0;
    padding: 1px 8px;
    min-width: 60px;
    font-weight: 600;
    text-align: center;
}
.xl-toolbar {
    display: flex;
    align-items: center;
    background: linear-gradient(180deg, #e8f0e8 0%, #d0dcd0 100%);
    border: 1px solid #a0b0a0;
    border-bottom: none;
    padding: 3px 8px;
    gap: 6px;
    font-family: Calibri, "Segoe UI", Arial, sans-serif;
    font-size: 11px;
    color: #444;
}
.xl-toolbar span {
    padding: 2px 6px;
    background: #f6faf6;
    border: 1px solid #c0ccc0;
    border-radius: 2px;
    font-weight: 600;
    font-size: 10.5px;
}
</style>
"""


def report_plant_index(
    readings: pd.DataFrame, mis: pd.DataFrame, plants: pd.DataFrame
) -> pd.DataFrame:
    """Which plant appears in which stored report: one row per
    (source_file, plant_sr_no, site_name). Names come from three places — the
    workbook's own sheet name, its MIS/cover site name, and the register's name for
    that SR — so a plant is findable by whichever name the user knows it by. Column
    names match `plant_search_mask()` so the same lookup rules apply here."""
    cols = ["source_file", "plant_sr_no", "site_name"]
    parts: list[pd.DataFrame] = []
    if readings is not None and not readings.empty:
        parts.append(
            readings.reindex(columns=["source_file", "plant_sr_no", "plant"])
            .rename(columns={"plant": "site_name"})
        )
    if mis is not None and not mis.empty:
        rows = mis
        if "source_file" in rows:  # drop the synthetic register-authority rows
            rows = rows[rows["source_file"] != REGISTER_SOURCE]
        parts.append(rows.reindex(columns=cols))
    if not parts:
        return pd.DataFrame(columns=cols)

    idx = pd.concat(parts, ignore_index=True).dropna(subset=["source_file"])
    idx["plant_sr_no"] = pd.to_numeric(idx["plant_sr_no"], errors="coerce").astype("Int64")

    if plants is not None and not plants.empty:
        register = plants.dropna(subset=["plant_sr_no"])
        names = dict(zip(
            pd.to_numeric(register["plant_sr_no"], errors="coerce").astype("Int64"),
            register.get("site_name", pd.Series(dtype=object)),
        ))
        aliases = idx[["source_file", "plant_sr_no"]].drop_duplicates().copy()
        aliases["site_name"] = aliases["plant_sr_no"].map(names)
        idx = pd.concat([idx, aliases.dropna(subset=["site_name"])], ignore_index=True)

    return idx.reindex(columns=cols).drop_duplicates().reset_index(drop=True)


def render_imr_preview(engine: Engine) -> None:
    """Render a stored IMR workbook exactly as it appears in Excel — with sheet
    tabs, column headers, row numbers, merged cells, and the classic spreadsheet
    grid. The raw bytes are loaded from the database and read with openpyxl."""
    st.title("📗 IMR Preview")
    st.caption(
        "Select a previously uploaded IMR to view the raw Excel workbook exactly as stored — "
        "each sheet rendered with its original layout, merged cells, and values. Search by "
        "SR No or site name to find the reports a particular plant appears in."
    )

    # A deep link from another page (selecting a module on the Portfolio) hands off
    # the plant to search for and the exact workbook to open. Seeding the widget keys
    # before the widgets are built is what makes them come up pre-filled; popping the
    # request means later reruns on this page behave normally.
    request = st.session_state.pop("preview_open_request", None)
    if request:
        st.session_state["preview_plant_query"] = str(request.get("query") or "")
    requested_file = str(request.get("source_file")) if request else None

    stored = report_files_index(engine)
    if not stored:
        st.info(
            "No reports have been stored yet. Upload reports through the "
            "**Manage Data** panel in the sidebar."
        )
        return

    # Also get ingested_files summary for the landing table.
    reports = ingested_report_summary(engine)

    # ----- Look up which reports contain a given plant -----
    query = st.text_input(
        "🔎 Find the reports for a plant",
        placeholder="SR No (e.g. 1043) or part of a site name",
        help="Narrows the report list below to the workbooks that contain that plant. "
             "Matches the SR No, the workbook's own sheet/site name, or the name the "
             "plant has in the Plant Register.",
        key="preview_plant_query",
    ).strip()

    if requested_file and requested_file not in stored:
        st.warning(
            f"The raw file for **{requested_file}** is not in the report store, so it "
            "cannot be previewed. It was likely ingested before reports were kept "
            "byte-for-byte — re-upload it to make it previewable. Showing the other "
            "reports for this plant instead."
        )
        requested_file = None

    options = stored
    if query:
        index = report_plant_index(load_readings(), load_mis(), load_plants())
        hits = index[plant_search_mask(index, query)] if not index.empty else index
        files = {f for f in hits["source_file"] if f in set(stored)}
        if requested_file:
            # The caller already knows this report holds the plant; never let the
            # name-matching drop the very file we were asked to open.
            files.add(requested_file)
        if not files:
            st.warning(
                f"No stored report contains a plant matching “{query}”. Showing every "
                "report instead — the plant may not have been ingested yet."
            )
        else:
            options = [f for f in stored if f in files]
            found = (
                hits.dropna(subset=["plant_sr_no"])
                # Names are appended sheet name -> MIS -> register, so the last one
                # per SR is the tidiest name we have for it.
                .drop_duplicates("plant_sr_no", keep="last")
                .sort_values("plant_sr_no")
            )
            named = " · ".join(
                f"SR {int(r['plant_sr_no'])}"
                + (f" {r['site_name']}" if pd.notna(r["site_name"]) else "")
                for _, r in found.head(5).iterrows()
            )
            st.success(
                f"{len(options)} of {len(stored)} reports contain "
                + (named or f"“{query}”")
                + (" …" if len(found) > 5 else "")
            )

    # The query is part of the key so narrowing the list can't leave a stale
    # selection pointing at a report that is no longer an option.
    choice_key = f"preview_report_choice::{query.lower()}"
    if requested_file in options:
        # Clear anything this key is holding from an earlier visit, so the widget
        # initializes from `index` below and the deep link wins.
        st.session_state.pop(choice_key, None)
        default_index = options.index(requested_file)
    elif query and len(options) == 1 and options is not stored:
        default_index = 0  # a search that lands on a single report opens it
    else:
        default_index = None
    choice = st.selectbox(
        "Select a report to preview",
        options,
        index=default_index,
        placeholder="Pick a committed report…",
        key=choice_key,
    )

    if not choice:
        if not reports.empty:
            st.markdown("---")
            st.subheader("Committed reports")
            summary_table = pd.DataFrame(
                {
                    "Report": reports["filename"],
                    "Readings": reports["n_readings"].fillna(0).astype(int),
                    "Parameters": reports["n_parameters"].fillna(0).astype(int),
                    "MIS rows": reports["n_mis"].fillna(0).astype(int),
                    "Ingested at": reports["ingested_at"].map(
                        lambda v: pd.Timestamp(v).strftime("%d %b %Y %H:%M") if pd.notna(v) else "—"
                    ),
                }
            )
            st.dataframe(summary_table, width="stretch", hide_index=True)
        return

    # Load the raw Excel bytes.
    data = load_report_bytes(engine, choice)
    if data is None:
        st.error(f"No stored bytes found for **{choice}**. The report may have been removed.")
        return

    # Download button for the original file.
    ext = ".xlsx" if choice.lower().endswith(".xlsx") else ".xls"
    st.download_button(
        f"⬇ Download original file",
        data,
        file_name=choice,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext == ".xlsx" else "application/vnd.ms-excel",
        key="preview_download_original",
    )

    render_excel_workbook(data, choice, key="imr_preview")


def render_excel_workbook(data: bytes, label: str, key: str) -> None:
    """Render raw Excel bytes as an Excel-like preview — a spreadsheet grid with
    sheet tabs, column-letter headers, a row-number gutter, merged cells and
    gridlines. Shared by the IMR Preview page and the upload-review panel."""
    from openpyxl import load_workbook as _load_wb
    try:
        wb = _load_wb(io.BytesIO(data), read_only=False, data_only=True)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not open **{label}** as an Excel workbook: {exc}")
        return

    sheet_names = wb.sheetnames
    if not sheet_names:
        st.warning("The workbook has no sheets.")
        wb.close()
        return

    st.markdown(EXCEL_PREVIEW_CSS, unsafe_allow_html=True)
    st.markdown(
        f'<div class="xl-toolbar">'
        f'<span>📗 {label}</span>'
        f'<span>{len(sheet_names)} sheet{"s" if len(sheet_names) != 1 else ""}</span>'
        f'<span>Read-only preview</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    picked_sheet = st.radio(
        "Sheet", sheet_names, horizontal=True,
        key=f"{key}_sheet_tab", label_visibility="collapsed",
    )
    ws = wb[picked_sheet]
    dims = f"{ws.max_row or 0} rows × {ws.max_column or 0} cols"
    st.markdown(
        f'<div class="xl-formula-bar">'
        f'<div class="xl-cell-ref">A1</div>'
        f'<div style="flex:1;color:#888;">{picked_sheet} — {dims}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        _render_sheet_as_excel_html(ws, max_rows=200, max_cols=None),
        unsafe_allow_html=True,
    )
    wb.close()


def main() -> None:
    st.set_page_config(
        page_title="RO Membrane Health Dashboard",
        page_icon=":bar_chart:",
        layout="wide",
    )
    st.markdown(
        """
        <style>
        .block-container { padding-top: 1.8rem; }
        .metric-card {
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            background: #ffffff;
            padding: 16px 18px;
            min-height: 118px;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06);
        }
        /* A metric card that links to its own detail section. Block-level so the
           whole tile is the hit area, and it must not pick up link underline/blue —
           the inner rows set their own colours. */
        a.metric-card-link {
            display: block;
            text-decoration: none;
            color: inherit;
            cursor: pointer;
            transition: border-color .12s ease, box-shadow .12s ease, transform .12s ease;
        }
        a.metric-card-link:hover {
            border-color: #bfdbfe;
            box-shadow: 0 2px 8px rgba(15, 23, 42, 0.12);
            transform: translateY(-1px);
        }
        .metric-title {
            color: #64748b;
            font-size: 0.82rem;
            font-weight: 700;
            letter-spacing: 0;
            text-transform: uppercase;
        }
        .metric-value {
            color: #0f172a;
            font-size: clamp(1.35rem, 2vw, 2rem);
            line-height: 1.15;
            font-weight: 800;
            margin-top: 0.42rem;
            overflow-wrap: anywhere;
        }
        .metric-subtitle {
            color: #64748b;
            font-size: 0.86rem;
            margin-top: 0.45rem;
            min-height: 1.1rem;
        }
        .hero-card {
            border: 1px solid #fecaca;
            border-radius: 12px;
            background: linear-gradient(180deg, #fff5f5 0%, #ffffff 70%);
            padding: 22px 26px;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08);
        }
        .hero-title {
            color: #64748b;
            font-size: 0.95rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.02em;
        }
        .hero-value {
            font-size: clamp(3rem, 7vw, 5rem);
            line-height: 1.05;
            font-weight: 800;
            margin-top: 0.2rem;
        }
        .hero-subtitle {
            color: #64748b;
            font-size: 1rem;
            margin-top: 0.35rem;
        }
        .nav-grid {
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            margin: 6px 0 2px;
        }
        .nav-card {
            flex: 1 1 150px;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            background: #f8fafc;
            padding: 10px 14px;
            text-decoration: none;
            display: flex;
            flex-direction: column;
            gap: 2px;
            transition: background .12s ease, border-color .12s ease, box-shadow .12s ease;
        }
        .nav-card:hover {
            background: #eff6ff;
            border-color: #bfdbfe;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08);
        }
        .nav-card .nav-label {
            color: #0f172a;
            font-weight: 700;
            font-size: 0.9rem;
        }
        .nav-card .nav-sub {
            color: #64748b;
            font-size: 0.78rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    engine = get_engine()
    if engine is None:
        st.error(
            "No database is configured. Set a `DATABASE_URL` environment variable, "
            "or add a `.streamlit/secrets.toml` with:\n\n"
            "```toml\n[postgres]\nurl = \"postgresql://USER:PASSWORD@HOST:5432/DBNAME\"\n```"
        )
        return

    try:
        init_db(engine)
        seeded = seed_plants_if_empty(engine, str(APP_DIR))
        summary = ingest_reports(engine, str(APP_DIR))
        # Rebuild from durable bytes anything whose derived rows went missing
        # (e.g. Cloud restart wiped disk but report_files survived). Usually a no-op.
        recovered = ingest_report_files(engine)
        summary["ingested"].extend(recovered["ingested"])
        summary["failed"].extend(recovered["failed"])
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not reach the database: {exc}")
        return

    # New/changed reports were parsed this run — drop the cached query results so
    # the fresh rows show up. Seeding the register also affects the folded-in zones.
    if summary["ingested"] or seeded:
        load_readings.clear()
        load_parameters.clear()
        load_plants.clear()
        load_mis.clear()
        compute_fleet_status.clear()

    df = load_readings()
    if df.empty:
        st.error("No RO module readings are in the database yet.")
        if summary["failed"]:
            st.write(summary["failed"])
        return

    if summary["failed"]:
        with st.sidebar.expander("Skipped inputs"):
            for item in summary["failed"]:
                st.write(item)

    params = load_parameters()
    mis = load_mis()
    plants = load_plants()

    render_data_manager(engine)

    PAGES.clear()
    PAGES["preview"] = st.Page(
        lambda: render_imr_preview(engine),
        title="IMR Preview",
        icon="🔍",
        url_path="preview",
    )

    pages = st.navigation(
        [
            st.Page(
                lambda: render_portfolio_page(df, params, mis),
                title="Portfolio",
                icon="🌐",
                url_path="portfolio",
                default=True,
            ),
            st.Page(
                lambda: render_input_tracker(df, plants),
                title="IMR Tracker",
                icon="📥",
                url_path="tracker",
            ),
            st.Page(
                lambda: render_plant_register(engine),
                title="Plant Register",
                icon="🏭",
                url_path="register",
            ),
            PAGES["preview"],
            st.Page(
                lambda: render_scan_page(engine),
                title="Scan IMR",
                icon="📷",
                url_path="scan",
            ),
            st.Page(
                lambda: render_dashboard(df, params),
                title="Dashboard",
                icon="📊",
                url_path="dashboard",
            ),
            st.Page(
                lambda: render_replacement_page(df),
                title="Replacement Candidates",
                icon="🛠️",
                url_path="replacement",
            ),
        ]
    )
    # A just-uploaded report takes over the main panel for review (Excel preview +
    # Confirm/Discard); the sidebar nav stays put. Once nothing is pending, the
    # selected page renders normally.
    if st.session_state.get("staged"):
        render_staged_review(engine)
    else:
        pages.run()


if __name__ == "__main__":
    main()
