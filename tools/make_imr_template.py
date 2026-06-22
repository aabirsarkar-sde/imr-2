"""Generate the hardened IMR data-entry template.

Run:  python3 tools/make_imr_template.py
Output: templates/IMR_template.xlsx

This is the "harden the existing grid" template: it keeps the site managers'
familiar wide stage-block layout, but makes the sheet machine-trustworthy:

  * A locked COVER block with real, machine-read cells (PLANT SR NO, REPORT DATE,
    ZONE, SITE NAME, CAPACITY) so plant identity + date come from the sheet, not
    the filename.
  * Stage labels are DROPDOWNS (I/II/III/IV STAGE) — the operator cannot invent a
    second "II STAGE", which closes the stage-mislabel hole.
  * Per-cell data validation: module no. (whole >= 1), install date (date),
    flow (number >= 0), conductivity (0..50000 us/cm), colour (dropdown), and the
    parameter values (number >= 0).
  * Sheet protection: every structural cell is locked; only the input cells are
    editable, so inserting/shifting columns (which breaks the positional block
    parser) is blocked.

The parameter block uses the clean adjacent (tag | value | unit) layout the
reader prefers — no gap column.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---- shared style constants -------------------------------------------------
TITLE_FONT = Font(bold=True, size=14, color="0F172A")
LABEL_FONT = Font(bold=True, size=10, color="334155")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
STAGE_FONT = Font(bold=True, size=11, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")        # pale yellow = "type here"
HEADER_FILL = PatternFill("solid", fgColor="2563EB")       # blue header
STAGE_FILL = PatternFill("solid", fgColor="0F172A")        # dark stage band
COVER_FILL = PatternFill("solid", fgColor="EFF6FF")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
UNLOCK = Protection(locked=False)

# Edit these to match the real zone register.
ZONES = ["Ankleshwar", "Jhagadia", "Panoli", "Dahej"]
STAGES = ["I STAGE", "II STAGE", "III STAGE", "IV STAGE"]

# Grid geometry.
GRID_COLUMNS = ["Mo no.", "Inst Date", "Total flow liter/hr.", "Cond.    us/cm"]
N_STAGE_BLOCKS = 3            # I / II / III laid out side by side
ROWS_PER_STAGE = 25          # generous; leave extras blank
BLOCK_WIDTH = len(GRID_COLUMNS) + 1   # 4 data cols + 1 spacer

# Parameter block: (tag, unit). Mirrors the instruments the app reads.
PARAMETERS = [
    ("PI - 121 -", "bar"), ("PI - 141 -", "bar"), ("PI - 151 -", "bar"),
    ("PI - 1601 -", "bar"), ("PI - 1602 -", "bar"), ("PI - 171 -", "bar"),
    ("PI - 180 -", "bar"),
    ("CIS - 151 -", "us/cm"), ("CIS - 180 -", "us/cm"),
    ("pH - 141 -", "-"), ("TI - 151 -", "deg. C"),
    ("FI - 181 -", "lit/hrs."), ("FI - 182 -", "lit/hrs."), ("FIS - 180 -", "lit/hrs."),
    ("Reject cond", "us/cm"), ("Feed Flow", "m3/hr"), ("Permeat Flow", "m3/hr"),
]


def input_cell(ws, coord, *, fill=INPUT_FILL):
    c = ws[coord]
    c.protection = UNLOCK
    c.fill = fill
    c.border = BORDER
    return c


def build_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    grid_top = 9                              # stage-band row
    header_row = grid_top + 1                 # column-header row
    data_top = header_row + 1
    data_bottom = data_top + ROWS_PER_STAGE - 1

    # one reusable validation per kind (Excel wants them added once, ranged many)
    dv_stage = DataValidation(type="list", formula1='"%s"' % ",".join(STAGES),
                              allow_blank=True, showErrorMessage=True,
                              errorTitle="Pick a stage",
                              error="Choose I/II/III/IV STAGE from the list.")
    dv_zone = DataValidation(type="list", formula1='"%s"' % ",".join(ZONES),
                             allow_blank=False, showErrorMessage=True,
                             errorTitle="Pick a zone", error="Choose a zone from the list.")
    dv_module = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1",
                               allow_blank=True, showErrorMessage=True,
                               errorTitle="Module number",
                               error="Module number must be a whole number >= 1.")
    dv_date = DataValidation(type="date", operator="greaterThan", formula1="DATE(2000,1,1)",
                             allow_blank=True, showErrorMessage=True,
                             errorTitle="Date", error="Enter a real date.")
    dv_flow = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                             allow_blank=True, showErrorMessage=True,
                             errorTitle="Flow", error="Flow must be a number >= 0.")
    dv_cond = DataValidation(type="decimal", operator="between", formula1="0", formula2="50000",
                             allow_blank=True, showErrorMessage=True, errorTitle="Conductivity",
                             error="Conductivity must be 0..50000 us/cm.")
    dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                            allow_blank=True, showErrorMessage=True, errorTitle="Value",
                            error="Value must be a number >= 0.")
    dv_srno = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1",
                             allow_blank=False, showErrorMessage=True, errorTitle="Plant SR No",
                             error="Plant SR No must be a whole number >= 1.")
    for dv in (dv_stage, dv_zone, dv_module, dv_date, dv_flow, dv_cond, dv_pos, dv_srno):
        ws.add_data_validation(dv)

    # ---- title + COVER block (machine-read identity / date) -----------------
    ws["A1"] = "INDIVIDUAL MODULE REPORT (IMR)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "Fill ONLY the pale-yellow cells. The rest of the sheet is locked."
    ws["A2"].font = Font(italic=True, size=9, color="64748B")
    ws.merge_cells("A2:K2")

    cover = [
        ("A3", "PLANT SR NO :-", "C3", dv_srno),
        ("A4", "REPORT DATE :-", "C4", dv_date),
        ("A5", "ZONE :-", "C5", dv_zone),
        ("A6", "SITE NAME :-", "C6", None),
        ("A7", "PLANT CAPACITY :-", "C7", None),
    ]
    for label_coord, label, in_coord, dv in cover:
        ws[label_coord] = label
        ws[label_coord].font = LABEL_FONT
        ws[label_coord].alignment = LEFT
        ws.merge_cells(f"{label_coord}:B{label_coord[1:]}")
        cell = input_cell(ws, in_coord, fill=COVER_FILL)
        cell.alignment = LEFT
        ws.merge_cells(f"{in_coord}:E{in_coord[1:]}")
        if dv is not None:
            dv.add(ws[in_coord])
    ws["C4"].number_format = "dd-mmm-yyyy"
    ws["G3"] = "REPORT TIME :-"
    ws["G3"].font = LABEL_FONT
    input_cell(ws, "H3").alignment = LEFT

    # ---- stage grid blocks --------------------------------------------------
    for block in range(N_STAGE_BLOCKS):
        start_col = 1 + block * BLOCK_WIDTH
        # stage band (dropdown-controlled label)
        band = ws.cell(row=grid_top, column=start_col)
        end_col = start_col + len(GRID_COLUMNS) - 1
        ws.merge_cells(start_row=grid_top, start_column=start_col,
                       end_row=grid_top, end_column=end_col)
        band.fill = STAGE_FILL
        band.font = STAGE_FONT
        band.alignment = CENTER
        band.protection = UNLOCK            # the stage label itself is an input...
        band.value = STAGES[block] if block < len(STAGES) else ""
        dv_stage.add(band)                  # ...but constrained to the dropdown
        for r in range(grid_top, data_bottom + 1):
            band_fill_border(ws, r, start_col, end_col)

        # column headers
        for j, name in enumerate(GRID_COLUMNS):
            c = ws.cell(row=header_row, column=start_col + j, value=name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER

        # data input cells + validation
        for r in range(data_top, data_bottom + 1):
            mod = input_cell(ws, ws.cell(row=r, column=start_col).coordinate)
            inst = input_cell(ws, ws.cell(row=r, column=start_col + 1).coordinate)
            flow = input_cell(ws, ws.cell(row=r, column=start_col + 2).coordinate)
            cond = input_cell(ws, ws.cell(row=r, column=start_col + 3).coordinate)
            inst.number_format = "dd-mmm-yyyy"
            dv_module.add(mod)
            dv_date.add(inst)
            dv_flow.add(flow)
            dv_cond.add(cond)

    # ---- parameter block (clean adjacent tag | value | unit) ---------------
    p_label_col = 1 + N_STAGE_BLOCKS * BLOCK_WIDTH + 1
    p_value_col = p_label_col + 1
    p_unit_col = p_label_col + 2
    ws.cell(row=grid_top, column=p_label_col, value="OPERATING PARAMETERS").font = STAGE_FONT
    ws.cell(row=grid_top, column=p_label_col).fill = STAGE_FILL
    ws.merge_cells(start_row=grid_top, start_column=p_label_col,
                   end_row=grid_top, end_column=p_unit_col)
    ws.cell(row=grid_top, column=p_label_col).alignment = CENTER
    for i, (tag, unit) in enumerate(PARAMETERS):
        r = header_row + i
        lc = ws.cell(row=r, column=p_label_col, value=tag)
        lc.font = LABEL_FONT
        lc.border = BORDER
        val = input_cell(ws, ws.cell(row=r, column=p_value_col).coordinate)
        dv_pos.add(val)
        uc = ws.cell(row=r, column=p_unit_col, value=unit)
        uc.alignment = CENTER
        uc.border = BORDER
        uc.font = Font(size=9, color="64748B")

    # ---- column widths + protection ----------------------------------------
    for block in range(N_STAGE_BLOCKS):
        sc = 1 + block * BLOCK_WIDTH
        ws.column_dimensions[get_column_letter(sc)].width = 8
        ws.column_dimensions[get_column_letter(sc + 1)].width = 12
        ws.column_dimensions[get_column_letter(sc + 2)].width = 14
        ws.column_dimensions[get_column_letter(sc + 3)].width = 12
        ws.column_dimensions[get_column_letter(sc + 4)].width = 2   # spacer
    ws.column_dimensions[get_column_letter(p_label_col)].width = 16
    ws.column_dimensions[get_column_letter(p_value_col)].width = 12
    ws.column_dimensions[get_column_letter(p_unit_col)].width = 9

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def band_fill_border(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.border.left is None or cell.border.left.style is None:
            cell.border = BORDER


def main() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Plant (rename to NAME + SR NO)"
    build_sheet(ws)
    out_dir = Path(__file__).resolve().parent.parent / "templates"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "IMR_template.xlsx"
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
